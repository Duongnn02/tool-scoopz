# -*- coding: utf-8 -*-

import os
import sys
import threading
import json
import csv
import subprocess
import shutil
import time
import math
import re
import random
import sqlite3
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed

# Ensure Tk uses bundled Tcl/Tk when frozen
if getattr(sys, "frozen", False):
    _MEI = getattr(sys, "_MEIPASS", "")
    _tcl = os.path.join(_MEI, "_tcl_data")
    _tk = os.path.join(_MEI, "_tk_data")
    if os.path.isdir(_tcl):
        os.environ.setdefault("TCL_LIBRARY", _tcl)
    if os.path.isdir(_tk):
        os.environ.setdefault("TK_LIBRARY", _tk)

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE_DIR = os.path.dirname(_THIS_DIR)
if _THIS_DIR not in sys.path:
    sys.path.insert(0, _THIS_DIR)
if _BASE_DIR not in sys.path:
    sys.path.insert(1, _BASE_DIR)

def _resource_path(name: str) -> str:
    base = getattr(sys, "_MEIPASS", _THIS_DIR)
    return os.path.join(base, name)

from gpm_client import create_profile, start_profile, close_profile, delete_profile, extract_driver_info
from login_scoopz import login_scoopz, login_scoopz_profile, open_profile_in_scoopz
from config import (
    SCOOPZ_URL,
    SCOOPZ_UPLOAD_URL,
    COOKIES_FILE,
    COOKIES_FILE_FALLBACK,
    DATA_DIR,
    LICENSE_SERVER_ENABLED,
    LICENSE_SERVER_HOST,
    LICENSE_SERVER_PORT,
    LICENSE_SERVER_URL,
    LICENSE_ADMIN_TOKEN,
)
from license_server import start_license_server
from yt_simple_download import download_one
from fb_simple_download import download_one_facebook
from shorts_csv_store import get_next_unuploaded, mark_uploaded, update_title_if_empty
from scoopz_uploader import upload_prepare, upload_post_async
from followers_fetcher import fetch_followers
from profile_updater import fetch_youtube_profile_assets_local, fetch_facebook_profile_assets_local, update_profile_from_assets
from shorts_scanner import scan_shorts_for_email
from fb_reels_scanner import scan_facebook_reels_for_email, scan_facebook_reels_multi
from monetization_utils import normalize_payment_status, format_stats_count_map
from threading_utils import ResourcePool, RetryHelper, ThreadSafeCounter
from logging_config import initialize_logger
from rate_limiter import initialize_rate_limiting, get_operation_delayer
from operation_orchestrator import initialize_orchestrator


_PACER_LOCK = threading.Lock()
_PACER_LAST_TS = {
    "create_profile": 0.0,
    "start_profile": 0.0,
    "login": 0.0,
}


def _humanized_action_pace(action: str, min_gap_s: float, jitter_min_s: float, jitter_max_s: float) -> None:
    """Global pacing gate to avoid bursty multi-thread browser actions."""
    with _PACER_LOCK:
        now = time.time()
        last = _PACER_LAST_TS.get(action, 0.0)
        wait_base = max(0.0, min_gap_s - (now - last))
        _PACER_LAST_TS[action] = now + wait_base
    if wait_base > 0:
        time.sleep(wait_base)
    if jitter_max_s > 0:
        time.sleep(random.uniform(jitter_min_s, jitter_max_s))


_raw_create_profile = create_profile
_raw_start_profile = start_profile
_raw_login_scoopz = login_scoopz


def _paced_create_profile(*args, **kwargs):
    _humanized_action_pace("create_profile", min_gap_s=1.2, jitter_min_s=0.25, jitter_max_s=0.8)
    return _raw_create_profile(*args, **kwargs)


def _paced_start_profile(*args, **kwargs):
    _humanized_action_pace("start_profile", min_gap_s=1.5, jitter_min_s=0.3, jitter_max_s=0.9)
    return _raw_start_profile(*args, **kwargs)


def _paced_login_scoopz(*args, **kwargs):
    _humanized_action_pace("login", min_gap_s=1.8, jitter_min_s=0.35, jitter_max_s=1.0)
    return _raw_login_scoopz(*args, **kwargs)


# Rebind locally so all existing flows automatically use paced operations.
create_profile = _paced_create_profile
start_profile = _paced_start_profile
login_scoopz = _paced_login_scoopz


ACCOUNTS = []
PROFILE_BATCH_SIZE = 100
PROFILE_BATCH_STAGGER_SEC = 0.15
SKIP_DOWNLOAD_UPLOAD = False  # TEMP: skip download/upload, only follow + creator fund check
CLEAR_CREATOR_FUND_ON_START = False  # TEMP: clear creator_fund_status in DB on startup
FIXED_SCAN_FOLDERS = {
    "alfreorasoly26_at_hotmail_com",
    "driterlaruu_at_hotmail_com",
    "dwengahuiju_at_hotmail_com",
    "eduanyadzia_at_hotmail_com",
    "janioanshaa11_at_hotmail_com",
    "opendauria_at_hotmail_com",
    "shueybrwamo_at_hotmail_com",
    "jassornivar_at_hotmail_com",
    "navudagishan_at_hotmail_com",
    "ussmanentang_at_hotmail_com",
    "yhamirchhaya_at_hotmail_com",
    "kaisynadena_at_hotmail_com",
    "nkhusikpatsa_at_hotmail_com",
    "tybenduviol_at_hotmail_com",
    "adlaheok_at_hotmail_com",
    "norbilagami_at_hotmail_com",
    "tekanalenart_at_hotmail_com",
    "utasirme_at_hotmail_com",
    "ibadetmorsin9684_at_hotmail_com",
    "csutatabong_at_hotmail_com",
    "curagariba_at_hotmail_com",
}
FIXED_SCAN_EMAILS = {
    f.replace("_at_", "@").replace("_", ".") for f in FIXED_SCAN_FOLDERS
}


class App:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        try:
            self.root.withdraw()
        except Exception:
            pass
        self.root.title("GPM Multi-Profile Suite")
        self.root.geometry("1280x720")
        self.root.minsize(1180, 640)
        try:
            icon_path = _resource_path("logo.ico")
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except Exception:
            pass
        self._cache_db = os.path.join(DATA_DIR, "cache.db")
        self._cache_db_lock = threading.Lock()
        self._extra_proxy_file = os.path.join(DATA_DIR, "extra_proxies.txt")
        self._default_gpm_root = os.path.join(DATA_DIR, "gpm_profiles")
        try:
            os.makedirs(self._default_gpm_root, exist_ok=True)
        except Exception:
            pass
        
        # Initialize logger
        log_dir = os.path.join(DATA_DIR, "logs")
        self.error_logger = initialize_logger(log_dir)
        self.error_logger.log_info("SYSTEM", "START", "Application started")
        try:
            self.error_logger.log_info("SYSTEM", "DB", f"Cache DB: {self._cache_db}")
        except Exception:
            pass
        
        # Initialize orchestrator with CONSERVATIVE mode
        # This coordinates all operations: login delays, sequential downloads, serial uploads
        self.orchestrator = initialize_orchestrator("balanced", logger=self.error_logger.main_logger.info)
        
        # Initialize rate limiter with conservative strategy
        initialize_rate_limiting("conservative")
        self.operation_delayer = get_operation_delayer()

        self.stop_event = threading.Event()
        self.executor = None
        self.executor_lock = threading.Lock()  # Lock for executor access
        self.active_profiles = {}
        self.active_lock = threading.Lock()
        self.created_profiles = set()
        self.create_lock = threading.Lock()
        # Resource management
        self.post_button_semaphore = threading.BoundedSemaphore(1)  # ⭐ CRITICAL: Only 1 POST button click at a time!
        self.upload_retry_semaphore = threading.BoundedSemaphore(2)  # Max 2 concurrent uploads
        self.login_semaphore = threading.BoundedSemaphore(2)  # Max 2 concurrent logins
        self.active_drivers = {}  # Track active drivers per thread
        self.active_drivers_lock = threading.Lock()
        self.profile_active_drivers = {}
        self.profile_active_drivers_lock = threading.Lock()
        self.profile_paths = {}
        self.profile_paths_lock = threading.Lock()
        self.profile_paths_used = set()
        self._gpm_cleanup_count = 0
        self._gpm_cleanup_lock = threading.Lock()
        self._gpm_cleanup_running = False
        self.profile_created_profiles = set()
        self._upload_queue_lock = threading.Lock()
        self._upload_queue_cond = threading.Condition(self._upload_queue_lock)
        self._upload_queue = []
        self._upload_queue_seq = 0
        self.profile_semaphore = None
        self.profile_update_lock = threading.Lock()
        self.csv_lock = threading.Lock()  # CSV atomic operations
        
        # Track failed accounts for retry after completion
        self.failed_accounts = []
        self.failed_accounts_lock = threading.Lock()
        self.profile_failed_accounts = []
        self.profile_failed_lock = threading.Lock()
        
        self._log_lock = threading.Lock()
        self._dragging = False
        self._drag_start = None
        self._context_item = None
        self._profile_dragging = False
        self._profile_drag_start = None
        self._profile_context_item = None
        self._fb_dragging = False
        self._fb_drag_start = None
        self._fb_context_item = None
        self._fb_profile_dragging = False
        self._fb_profile_drag_start = None
        self._all_dragging = False
        self._all_drag_start = None
        self._all_context_item = None
        self._from_all_tab = False
        self._all_pending_fb_emails = set()
        self._all_repeat_snapshot = None
        self._all_filter_active = False
        self._all_retry_round = 0
        self._fallback_caption_file = os.path.join(DATA_DIR, "fallback_captions.txt")
        self._fallback_captions = []
        self._fallback_caption_idx = 0
        self._fallback_caption_lock = threading.Lock()
        self._creator_fund_checked = set()
        self._creator_fund_lock = threading.Lock()
        self._extra_proxies = []
        self._extra_proxy_idx = 0
        self._extra_proxy_lock = threading.Lock()
        self._cell_editor = None
        self._manage_email = ""
        self._manage_rows = []
        self._manage_fieldnames = []
        self._manage_csv_path = ""
        self._manage_email_map = []
        self._manage_email_all = []
        self._manage_search_var = tk.StringVar()
        self._cookie_status_var = tk.StringVar(value="Cookie: -")
        self._license_server = None
        self._license_key_cache = ""
        self._license_valid = False
        self.repeat_var = tk.BooleanVar(value=True)
        self._advanced_var = tk.BooleanVar(value=False)
        self._busy = False
        self._repeat_after_id = None
        self._repeat_countdown_after_id = None
        self._repeat_enabled = False
        self._repeat_delay_sec = 0
        self._repeat_cycle_pending = False
        self._cycle_count = 0
        self._run_upload_after_fb = False
        self._force_upload_only = False
        self._fixed_threads = None
        self._max_retry_rounds = 1
        self._upload_retry_rounds = 0
        self._retry_round = 0
        self._profile_retry_round = 0
        self._profile_batch_running = False
        self._run_counts = {
            "upload": {"done": 0, "total": 0, "emails": set()},
            "profile": {"done": 0, "total": 0, "emails": set()},
            "fb": {"done": 0, "total": 0, "emails": set()},
            "fb_profile": {"done": 0, "total": 0, "emails": set()},
        }
        self._run_counts_lock = threading.Lock()
        self._runtime_var = tk.StringVar(value="Runtime: 00:00:00")
        self._perf_var = tk.StringVar(value="Speed: 0.0 acc/min | ETA: -")
        self._upload_stats_var = tk.StringVar(value="Run: processed 0 | ok 0 | no video 0 | err 0")
        self._runtime_accum_sec = 0.0
        self._runtime_started_at = None
        self._runtime_running = False
        self._runtime_after_id = None
        self._upload_outcomes = {}
        self._upload_stats_lock = threading.Lock()
        self._resume_pending = {
            "upload": set(),
            "profile": set(),
            "fb": set(),
            "fb_profile": set(),
        }
        self._count_var = tk.StringVar(value="Total: 0 | YTB: 0 | FB: 0")
        self._profile_count_var = tk.StringVar(value="YTB Profile: 0")
        self._fb_profile_count_var = tk.StringVar(value="FB Profile: 0")
        self._stats_total_var = tk.StringVar(value="Total: 0")
        self._stats_breakdown_var = tk.StringVar(value="Creator: - | Payment: -")
        self._follow_sort_after_id = None
        self._job_item_email_map = {}
        self._job_item_email_lock = threading.Lock()
        self._transient_statuses = {
            "START...",
            "STARTED",
            "STARTED (no debug)",
            "DOWNLOAD...",
            "DOWNLOAD OK",
            "POSTING...",
            "LOGIN...",
            "RESTART...",
        }
        self._auto_scroll_block_until = {}
        self._auto_scroll_catchup_after_id = {}
        self._last_active_item = {}
        self._pulse_tokens = {}
        self._search_last_query = ""
        self._search_last_tab = ""
        self._search_last_matches = []
        self._search_cycle_index = -1
        self.active_channel = tk.StringVar(value="ytb")
        self._sidebar_buttons = {}
        self._active_sidebar_key = "overview"

        self._apply_theme()
        self._migrate_legacy_data()
        self._build_ui()
        self.root.after(300, self._show_intro)
        self._load_fallback_captions()
        self._load_extra_proxy_list()
        self._start_license_server()
        self._clear_not_applied_status_cache()
        self._normalize_payment_status_cache()
        self._reset_operational_status_cache()
        if CLEAR_CREATOR_FUND_ON_START:
            self._clear_creator_fund_status_cache()
        self.accounts = self._load_accounts_cache() or ACCOUNTS
        self._load_rows()
        self.profile_accounts = self._load_profile_accounts_cache()
        self._load_profile_rows()
        self.fb_accounts = self._load_fb_accounts_cache()
        self._load_fb_rows()
        self.fb_profile_accounts = self._load_fb_profile_accounts_cache()
        self._load_fb_profile_rows()
        self._refresh_manage_emails()
        self._load_cookie_into_form()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _apply_theme(self) -> None:
        palette = {
            "bg": "#F1F5F9",
            "panel": "#FFFFFF",
            "border": "#D9E2EC",
            "text": "#0F172A",
            "muted": "#64748B",
            "accent": "#0E7490",
            "accent_dark": "#155E75",
            "danger": "#B91C1C",
            "danger_dark": "#991B1B",
            "select_bg": "#E0F2FE",
        }
        self._palette = palette

        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        self.root.configure(bg=palette["bg"])

        base_font = ("Segoe UI", 10)
        title_font = ("Segoe UI Semibold", 14)
        subtitle_font = ("Segoe UI", 9)
        tab_font = ("Segoe UI Semibold", 10)

        style.configure("TFrame", background=palette["bg"])
        style.configure("Panel.TFrame", background=palette["panel"])
        style.configure("TLabel", background=palette["bg"], foreground=palette["text"], font=base_font)
        style.configure("Header.TLabel", background=palette["bg"], foreground=palette["text"], font=title_font)
        style.configure("Subtle.TLabel", background=palette["bg"], foreground=palette["muted"], font=subtitle_font)
        style.configure("SidebarTitle.TLabel", background=palette["bg"], foreground=palette["muted"], font=("Segoe UI Semibold", 9))
        style.configure("Panel.TLabel", background=palette["panel"], foreground=palette["text"], font=base_font)

        style.configure("TEntry", padding=6, fieldbackground=palette["panel"], foreground=palette["text"])
        style.configure("TButton", padding=(10, 6))
        style.configure(
            "Accent.TButton",
            background=palette["accent"],
            foreground="#FFFFFF",
            font=("Segoe UI Semibold", 10),
        )
        style.map("Accent.TButton", background=[("active", palette["accent_dark"])])
        style.configure(
            "Danger.TButton",
            background=palette["danger"],
            foreground="#FFFFFF",
            font=("Segoe UI Semibold", 10),
        )
        style.map("Danger.TButton", background=[("active", palette["danger_dark"])])

        style.configure("TNotebook", background=palette["bg"], tabmargins=(8, 4, 8, 0))
        style.configure("TNotebook.Tab", padding=(12, 6), font=tab_font)
        style.map(
            "TNotebook.Tab",
            background=[("selected", palette["panel"])],
            foreground=[("selected", palette["text"])],
        )
        style.configure("Hidden.TNotebook", background=palette["bg"], tabmargins=0)
        style.layout("Hidden.TNotebook.Tab", [])

        style.configure("TLabelframe", background=palette["bg"], bordercolor=palette["border"])
        style.configure(
            "TLabelframe.Label",
            background=palette["bg"],
            foreground=palette["muted"],
            font=("Segoe UI Semibold", 9),
        )

        style.configure("Status.TLabel", background=palette["bg"], foreground=palette["muted"], font=subtitle_font)
        style.configure(
            "Accent.Horizontal.TProgressbar",
            troughcolor=palette["border"],
            background=palette["accent"],
            bordercolor=palette["border"],
            lightcolor=palette["accent"],
            darkcolor=palette["accent_dark"],
        )

        style.configure(
            "Treeview",
            background=palette["panel"],
            fieldbackground=palette["panel"],
            foreground=palette["text"],
            rowheight=24,
            font=base_font,
            bordercolor=palette["border"],
            borderwidth=1,
        )
        style.map(
            "Treeview",
            background=[("selected", palette["select_bg"])],
            foreground=[("selected", palette["text"])],
        )
        style.configure(
            "Treeview.Heading",
            background=palette["bg"],
            foreground=palette["text"],
            font=("Segoe UI Semibold", 9),
            padding=(6, 4),
        )
        style.map("Treeview.Heading", background=[("active", palette["border"])])

        style.configure(
            "Sidebar.TButton",
            background=palette["panel"],
            foreground=palette["text"],
            font=("Segoe UI Semibold", 10),
            padding=(12, 9),
            anchor="w",
        )
        style.map(
            "Sidebar.TButton",
            background=[("active", palette["select_bg"])],
            foreground=[("active", palette["text"])],
        )
        style.configure(
            "SidebarActive.TButton",
            background=palette["accent"],
            foreground="#FFFFFF",
            font=("Segoe UI Semibold", 10),
            padding=(12, 9),
            anchor="w",
        )
        style.map("SidebarActive.TButton", background=[("active", palette["accent_dark"])])
        style.configure(
            "Channel.TRadiobutton",
            background=palette["panel"],
            foreground=palette["text"],
            font=("Segoe UI Semibold", 10),
        )
        style.map(
            "Channel.TRadiobutton",
            foreground=[("selected", palette["accent_dark"]), ("!selected", palette["text"])],
            background=[("active", palette["panel"])],
        )

    def _build_ui(self) -> None:
        self._ui_paned = tk.PanedWindow(
            self.root,
            orient="vertical",
            sashwidth=6,
            sashrelief="raised",
            bd=0,
            bg=self._palette.get("border", "#E5E7EB"),
        )
        self._ui_paned.pack(fill="both", expand=True)

        main_container = ttk.Frame(self._ui_paned)
        log_container = ttk.Frame(self._ui_paned)
        self._ui_paned.add(main_container, stretch="always")
        self._ui_paned.add(log_container, minsize=80)

        control_panel = ttk.Frame(main_container)
        control_panel.pack(fill="x", padx=12, pady=(10, 6))
        control_panel.columnconfigure(0, weight=1)
        control_panel.columnconfigure(1, weight=2)
        control_panel.columnconfigure(2, weight=1)

        run_frame = ttk.LabelFrame(control_panel, text="Run Settings")
        run_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        run_frame.columnconfigure(1, weight=1)

        ttk.Label(run_frame, text="Threads:").grid(row=0, column=0, sticky="w", padx=8, pady=(8, 4))
        self.entry_threads = ttk.Entry(run_frame, width=6)
        self.entry_threads.insert(0, "5")
        self.entry_threads.grid(row=0, column=1, sticky="w", padx=(0, 12), pady=(8, 4))

        ttk.Label(run_frame, text="Videos:").grid(row=0, column=2, sticky="w", padx=(0, 6), pady=(8, 4))
        self.entry_videos = ttk.Entry(run_frame, width=6)
        self.entry_videos.insert(0, "1")
        self.entry_videos.grid(row=0, column=3, sticky="w", padx=(0, 8), pady=(8, 4))

        self.chk_repeat = ttk.Checkbutton(run_frame, text="Repeat", variable=self.repeat_var)
        self.chk_repeat.grid(row=1, column=0, sticky="w", padx=8, pady=(0, 8))
        ttk.Label(run_frame, text="Delay (min):").grid(row=1, column=1, sticky="w", padx=(0, 6), pady=(0, 8))
        self.entry_repeat_delay = ttk.Entry(run_frame, width=6)
        self.entry_repeat_delay.insert(0, "5")
        self.entry_repeat_delay.grid(row=1, column=2, sticky="w", padx=(0, 8), pady=(0, 8))

        path_frame = ttk.LabelFrame(control_panel, text="Paths & Search")
        path_frame.grid(row=0, column=1, sticky="nsew", padx=(0, 8))
        path_frame.columnconfigure(1, weight=1)

        ttk.Label(path_frame, text="GPM Path:").grid(row=0, column=0, sticky="w", padx=8, pady=(8, 4))
        self.entry_gpm_path = ttk.Entry(path_frame, width=28)
        self.entry_gpm_path.insert(0, self._default_gpm_root)
        self.entry_gpm_path.grid(row=0, column=1, sticky="ew", padx=(0, 8), pady=(8, 4))

        self._search_placeholder = "Search email..."
        self.entry_search_email = ttk.Entry(path_frame, width=30)
        self.entry_search_email.insert(0, self._search_placeholder)
        self.entry_search_email.grid(row=1, column=0, columnspan=2, sticky="ew", padx=8, pady=(0, 8))
        try:
            self.entry_search_email.configure(foreground="gray")
        except Exception:
            pass
        self.entry_search_email.bind("<FocusIn>", self._search_focus_in)
        self.entry_search_email.bind("<FocusOut>", self._search_focus_out)
        self.entry_search_email.bind("<Return>", self._search_email)
        self.entry_search_email.bind("<Shift-Return>", self._search_email_prev)
        self.btn_search_email = ttk.Button(path_frame, text="FIND", command=self._search_email)
        self.btn_search_email.grid(row=1, column=2, sticky="w", padx=(0, 8), pady=(0, 8))

        self._sort_state = {}
        self.btn_sort_follow_all = ttk.Button(path_frame, text="SORT FOLLOW ALL", command=self._toggle_followers_sort_all)
        self.btn_sort_follow_all.grid(row=0, column=2, sticky="w", padx=(0, 8), pady=(8, 4))

        action_frame = ttk.LabelFrame(control_panel, text="Actions")
        action_frame.grid(row=0, column=2, sticky="nsew")

        self.btn_start = ttk.Button(action_frame, text="START", command=self.start_jobs, style="Accent.TButton")
        self.btn_start.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 4))
        self.btn_stop = ttk.Button(action_frame, text="STOP", command=self.stop_jobs, style="Danger.TButton")
        self.btn_stop.grid(row=0, column=1, sticky="ew", padx=8, pady=(8, 4))
        self.btn_reload = ttk.Button(action_frame, text="RELOAD", command=self.reload_app)
        self.btn_reload.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 8))

        self.chk_advanced = ttk.Checkbutton(
            action_frame,
            text="Show advanced",
            variable=self._advanced_var,
            command=self._toggle_advanced,
        )
        self.chk_advanced.grid(row=1, column=1, sticky="w", padx=8, pady=(0, 8))

        self._advanced_frame = ttk.Frame(action_frame)
        self._advanced_frame.grid(row=2, column=0, columnspan=2, sticky="ew", padx=8, pady=(0, 8))
        self._advanced_frame.columnconfigure(1, weight=1)
        self.btn_clear_videos = ttk.Button(self._advanced_frame, text="CLEAR VIDEOS", command=self.clear_all_email_videos)
        self.btn_clear_videos.grid(row=0, column=0, columnspan=2, sticky="ew")
        self.btn_clear_gpm = ttk.Button(self._advanced_frame, text="CLEAR GPM PROFILES", command=self._clear_all_gpm_profiles)
        self.btn_clear_gpm.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(6, 0))
        self._advanced_frame.grid_remove()

        status_frame = ttk.Frame(main_container)
        status_frame.pack(fill="x", padx=12, pady=(0, 8))
        status_frame.columnconfigure(5, weight=1)

        self.lbl_total = ttk.Label(status_frame, textvariable=self._count_var, style="Status.TLabel")
        self.lbl_total.grid(row=0, column=0, sticky="w", padx=(0, 12))
        self.lbl_profile_total = ttk.Label(status_frame, textvariable=self._profile_count_var, style="Status.TLabel")
        self.lbl_profile_total.grid(row=0, column=1, sticky="w", padx=(0, 12))
        self.lbl_fb_profile_total = ttk.Label(status_frame, textvariable=self._fb_profile_count_var, style="Status.TLabel")
        self.lbl_fb_profile_total.grid(row=0, column=2, sticky="w", padx=(0, 12))

        self._busy_bar = ttk.Progressbar(status_frame, mode="indeterminate", style="Accent.Horizontal.TProgressbar")
        self._busy_bar.grid(row=0, column=4, sticky="w", padx=(6, 12))

        self._cycle_var = tk.StringVar(value="Cycles: 0")
        self._next_cycle_var = tk.StringVar(value="Next cycle: -")

        self.lbl_cycle = ttk.Label(status_frame, textvariable=self._cycle_var, style="Status.TLabel")
        self.lbl_cycle.grid(row=1, column=0, sticky="w", pady=(4, 0))
        self.lbl_next_cycle = ttk.Label(status_frame, textvariable=self._next_cycle_var, style="Status.TLabel")
        self.lbl_next_cycle.grid(row=1, column=1, sticky="w", pady=(4, 0), padx=(6, 0))
        self.lbl_runtime = ttk.Label(status_frame, textvariable=self._runtime_var, style="Status.TLabel")
        self.lbl_runtime.grid(row=1, column=2, sticky="w", pady=(4, 0), padx=(10, 0))
        self.lbl_perf = ttk.Label(status_frame, textvariable=self._perf_var, style="Status.TLabel")
        self.lbl_perf.grid(row=1, column=3, sticky="w", pady=(4, 0), padx=(10, 0))
        self.lbl_upload_stats = ttk.Label(status_frame, textvariable=self._upload_stats_var, style="Status.TLabel")
        self.lbl_upload_stats.grid(row=1, column=4, sticky="w", pady=(4, 0), padx=(10, 0))

        ttk.Separator(main_container, orient="horizontal").pack(fill="x", padx=12, pady=(2, 8))

        main_body = ttk.Frame(main_container)
        main_body.pack(fill="both", expand=True, padx=8, pady=8)

        sidebar = ttk.Frame(main_body, width=210)
        sidebar.pack(side="left", fill="y", padx=(0, 8))

        content = ttk.Frame(main_body)
        content.pack(side="right", fill="both", expand=True)

        self.notebook = ttk.Notebook(content, style="Hidden.TNotebook")
        self.tab_all = ttk.Frame(self.notebook)
        self.tab_upload = ttk.Frame(self.notebook)
        self.tab_profile = ttk.Frame(self.notebook)
        self.tab_fb = ttk.Frame(self.notebook)
        self.tab_fb_profile = ttk.Frame(self.notebook)
        self.tab_interact = ttk.Frame(self.notebook)
        self.tab_stats = ttk.Frame(self.notebook)
        self.tab_manage = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_all, text="Overview")
        self.notebook.add(self.tab_upload, text="YOUTUBE")
        self.notebook.add(self.tab_profile, text="PROFILE")
        self.notebook.add(self.tab_fb, text="FACEBOOK")
        self.notebook.add(self.tab_fb_profile, text="FB PROFILE")
        self.notebook.add(self.tab_interact, text="INTERACT")
        self.notebook.add(self.tab_stats, text="Monetization")
        self.notebook.add(self.tab_manage, text="MANAGEMENT")
        self.notebook.pack(fill="both", expand=True)

        # Logo
        self._logo_img = None
        try:
            logo_path = _resource_path("logo.png")
            if os.path.exists(logo_path):
                self._logo_img = tk.PhotoImage(file=logo_path)
                ttk.Label(sidebar, image=self._logo_img).pack(anchor="center", pady=(4, 8))
        except Exception:
            self._logo_img = None

        ttk.Label(sidebar, text="MENU", style="SidebarTitle.TLabel").pack(anchor="w", padx=8, pady=(2, 8))

        nav_items = [
            ("overview", "Overview", self.tab_all),
            ("upvideo", "UpVideo", self._open_upload_panel),
            ("profile", "Profile", self._open_profile_panel),
            ("interact", "Interact", self.tab_interact),
            ("monetization", "Monetization", self.tab_stats),
            ("manage", "Manage", self.tab_manage),
        ]
        self._sidebar_buttons = {}
        for key, label, target in nav_items:
            btn = ttk.Button(
                sidebar,
                text=label,
                style="Sidebar.TButton",
                command=lambda k=key, t=target: self._sidebar_go(k, t),
            )
            btn.pack(fill="x", padx=6, pady=4)
            self._sidebar_buttons[key] = btn
        self._set_sidebar_active("overview")

        all_top = ttk.Frame(self.tab_all)
        all_top.pack(fill="x", padx=8, pady=(8, 4))
        ttk.Button(all_top, text="Select All", command=self._select_all_all_accounts).pack(side="left", padx=(0, 4))
        ttk.Button(all_top, text="Deselect All", command=self._deselect_all_all_accounts).pack(side="left")
        ttk.Button(all_top, text="Filter errors", command=self._filter_all_errors).pack(side="left", padx=(8, 4))
        ttk.Button(all_top, text="Show all", command=self._clear_all_filter).pack(side="left")
        ttk.Button(all_top, text="SCAN", command=self.start_scan).pack(side="left", padx=(8, 0))

        all_table = ttk.Frame(self.tab_all)
        all_table.pack(fill="both", expand=True, padx=8, pady=8)
        self.all_tree = ttk.Treeview(
            all_table,
            columns=(
                "chk",
                "stt",
                "social",
                "email",
                "pass",
                "status",
                "posts",
                "followers",
                "proxy",
                "link",
                "profile_url",
                "profile_id",
            ),
            show="headings",
            selectmode="extended",
        )
        self.all_tree.heading("chk", text="v")
        self.all_tree.column("chk", width=40, anchor="center")
        self.all_tree.heading("stt", text="STT")
        self.all_tree.column("stt", width=50, anchor="center")
        self.all_tree.heading("social", text="NETWORK")
        self.all_tree.column("social", width=70, anchor="center")
        self.all_tree.heading("email", text="EMAIL")
        self.all_tree.column("email", width=240)
        self.all_tree.heading("pass", text="PASS")
        self.all_tree.column("pass", width=130)
        self.all_tree.heading("status", text="STATUS")
        self.all_tree.column("status", width=200)
        self.all_tree.heading("posts", text="POSTS")
        self.all_tree.column("posts", width=70, anchor="center")
        self.all_tree.heading("followers", text="FOLLOWERS")
        self.all_tree.column("followers", width=90, anchor="center")
        self.all_tree.heading("proxy", text="PROXY")
        self.all_tree.column("proxy", width=260)
        self.all_tree.heading("link", text="LINK")
        self.all_tree.column("link", width=280)
        self.all_tree.heading("profile_url", text="PROFILE URL")
        self.all_tree.column("profile_url", width=260)
        self.all_tree.heading("profile_id", text="PROFILE ID")
        self.all_tree.column("profile_id", width=240)

        def _on_all_scroll(*args):
            self._mark_user_scroll(self.all_tree)
            self.all_tree.yview(*args)

        all_scroll = ttk.Scrollbar(all_table, orient="vertical", command=_on_all_scroll)
        all_scroll_x = ttk.Scrollbar(all_table, orient="horizontal", command=self.all_tree.xview)
        self.all_tree.configure(yscrollcommand=all_scroll.set, xscrollcommand=all_scroll_x.set)
        self.all_tree.grid(row=0, column=0, sticky="nsew")
        all_scroll.grid(row=0, column=1, sticky="ns")
        all_scroll_x.grid(row=1, column=0, sticky="ew")
        all_table.grid_rowconfigure(0, weight=1)
        all_table.grid_columnconfigure(0, weight=1)
        self.all_tree.tag_configure("status_ok", foreground="green")
        self.all_tree.tag_configure("status_err", foreground="red")
        self.all_tree.tag_configure("status_work", foreground="#0F766E")
        self.all_tree.tag_configure("status_warn", foreground="#B45309")
        self.all_tree.tag_configure("status_flash", background="#E8F5FF")
        self.all_tree.tag_configure("status_pulse_a", background="#E8F5FF")
        self.all_tree.tag_configure("status_pulse_b", background="#E4FBEA")
        self.all_tree.bind("<Button-1>", self._on_all_tree_click)
        self.all_tree.bind("<B1-Motion>", self._on_all_tree_drag)
        self.all_tree.bind("<ButtonRelease-1>", self._on_all_tree_release)
        self.all_tree.bind("<Button-3>", self._on_all_tree_right_click)
        self.all_tree.bind("<MouseWheel>", lambda e, t=self.all_tree: self._mark_user_scroll(t))
        self.all_tree.bind("<Button-4>", lambda e, t=self.all_tree: self._mark_user_scroll(t))
        self.all_tree.bind("<Button-5>", lambda e, t=self.all_tree: self._mark_user_scroll(t))

        # Add Select All / Deselect All buttons for tab_upload
        btn_frame_upload = ttk.Frame(self.tab_upload)
        btn_frame_upload.pack(fill="x", padx=8, pady=(8, 4))

        ttk.Button(btn_frame_upload, text="IMPORT", command=self._import_active_upload_accounts).pack(side="left")
        ttk.Button(btn_frame_upload, text="Select All", command=self._select_all_active_upload_accounts).pack(side="left", padx=(8, 4))
        ttk.Button(btn_frame_upload, text="Deselect All", command=self._deselect_all_active_upload_accounts).pack(side="left")
        ttk.Button(btn_frame_upload, text="EXPORT", command=self._export_active_upload_accounts).pack(side="left", padx=(8, 4))
        ttk.Button(btn_frame_upload, text="SCAN", command=self._scan_active_upload_accounts).pack(side="left")
        ttk.Label(btn_frame_upload, text="Channel:", style="Subtle.TLabel").pack(side="left", padx=(14, 6))
        ttk.Radiobutton(
            btn_frame_upload,
            text="YTB",
            value="ytb",
            style="Channel.TRadiobutton",
            variable=self.active_channel,
            command=self._on_active_channel_changed,
        ).pack(side="left", padx=(0, 10))
        ttk.Radiobutton(
            btn_frame_upload,
            text="FB",
            value="fb",
            style="Channel.TRadiobutton",
            variable=self.active_channel,
            command=self._on_active_channel_changed,
        ).pack(side="left")

        upload_table = ttk.Frame(self.tab_upload)
        upload_table.pack(fill="both", expand=True, padx=8, pady=8)
        self.tree = ttk.Treeview(
            upload_table,
            columns=("chk", "stt", "email", "pass", "status", "posts", "followers", "proxy", "youtube", "profile_url", "profile_id"),
            show="headings",
            selectmode="extended",
        )
        self.tree.heading("chk", text="v")
        self.tree.column("chk", width=40, anchor="center")
        self.tree.heading("stt", text="STT")
        self.tree.column("stt", width=50, anchor="center")
        self.tree.heading("email", text="EMAIL")
        self.tree.column("email", width=240)
        self.tree.heading("pass", text="PASS")
        self.tree.column("pass", width=130)
        self.tree.heading("status", text="STATUS")
        self.tree.column("status", width=200)
        self.tree.heading("posts", text="POSTS", command=lambda: self._toggle_upload_sort("posts"))
        self.tree.column("posts", width=70, anchor="center")
        self.tree.heading("followers", text="FOLLOWERS", command=lambda: self._toggle_upload_sort("followers"))
        self.tree.column("followers", width=90, anchor="center")
        self.tree.heading("proxy", text="PROXY")
        self.tree.column("proxy", width=260)
        self.tree.heading("youtube", text="YOUTUBE")
        self.tree.column("youtube", width=280)
        self.tree.heading("profile_url", text="PROFILE URL")
        self.tree.column("profile_url", width=260)
        self.tree.heading("profile_id", text="PROFILE ID")
        self.tree.column("profile_id", width=240)

        def _on_upload_scroll(*args):
            self._mark_user_scroll(self.tree)
            self.tree.yview(*args)

        upload_scroll = ttk.Scrollbar(upload_table, orient="vertical", command=_on_upload_scroll)
        upload_scroll_x = ttk.Scrollbar(upload_table, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=upload_scroll.set, xscrollcommand=upload_scroll_x.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        upload_scroll.grid(row=0, column=1, sticky="ns")
        upload_scroll_x.grid(row=1, column=0, sticky="ew")
        upload_table.grid_rowconfigure(0, weight=1)
        upload_table.grid_columnconfigure(0, weight=1)
        self.tree.tag_configure("status_ok", foreground="green")
        self.tree.tag_configure("status_err", foreground="red")
        self.tree.tag_configure("status_work", foreground="#0F766E")
        self.tree.tag_configure("status_warn", foreground="#B45309")
        self.tree.tag_configure("status_flash", background="#E8F5FF")
        self.tree.tag_configure("status_pulse_a", background="#E8F5FF")
        self.tree.tag_configure("status_pulse_b", background="#E4FBEA")

        self.tree.bind("<Button-1>", self._on_tree_click)
        self.tree.bind("<B1-Motion>", self._on_tree_drag)
        self.tree.bind("<ButtonRelease-1>", self._on_tree_release)
        self.tree.bind("<Button-3>", self._on_tree_right_click)
        self.tree.bind("<MouseWheel>", lambda e, t=self.tree: self._mark_user_scroll(t))
        self.tree.bind("<Button-4>", lambda e, t=self.tree: self._mark_user_scroll(t))
        self.tree.bind("<Button-5>", lambda e, t=self.tree: self._mark_user_scroll(t))

        profile_top = ttk.Frame(self.tab_profile)
        profile_top.pack(fill="x", padx=8, pady=(8, 0))
        self.btn_import_profile = ttk.Button(profile_top, text="IMPORT PROFILE", command=self._import_active_profile_accounts)
        self.btn_import_profile.pack(side="left")
        ttk.Button(profile_top, text="CHUYEN DU LIEU", command=self._transfer_active_profile_accounts).pack(side="left", padx=(8, 4))
        ttk.Button(profile_top, text="Select All", command=self._select_all_active_profile_accounts).pack(side="left", padx=(8, 4))
        ttk.Button(profile_top, text="Deselect All", command=self._deselect_all_active_profile_accounts).pack(side="left")
        ttk.Label(profile_top, text="Channel:", style="Subtle.TLabel").pack(side="left", padx=(14, 6))
        ttk.Radiobutton(
            profile_top,
            text="YTB",
            value="ytb",
            style="Channel.TRadiobutton",
            variable=self.active_channel,
            command=self._on_active_channel_changed,
        ).pack(side="left", padx=(0, 10))
        ttk.Radiobutton(
            profile_top,
            text="FB",
            value="fb",
            style="Channel.TRadiobutton",
            variable=self.active_channel,
            command=self._on_active_channel_changed,
        ).pack(side="left")

        profile_table = ttk.Frame(self.tab_profile)
        profile_table.pack(fill="both", expand=True, padx=8, pady=8)
        self.profile_tree = ttk.Treeview(
            profile_table,
            columns=("chk", "stt", "email", "pass", "proxy", "youtube", "status"),
            show="headings",
            selectmode="extended",
        )
        self.profile_tree.heading("chk", text="v")
        self.profile_tree.column("chk", width=40, anchor="center")
        self.profile_tree.heading("stt", text="STT")
        self.profile_tree.column("stt", width=50, anchor="center")
        self.profile_tree.heading("email", text="EMAIL")
        self.profile_tree.column("email", width=240)
        self.profile_tree.heading("pass", text="PASS")
        self.profile_tree.column("pass", width=130)
        self.profile_tree.heading("proxy", text="PROXY")
        self.profile_tree.column("proxy", width=260)
        self.profile_tree.heading("youtube", text="YOUTUBE")
        self.profile_tree.column("youtube", width=280)
        self.profile_tree.heading("status", text="STATUS")
        self.profile_tree.column("status", width=200)

        def _on_profile_scroll(*args):
            self._mark_user_scroll(self.profile_tree)
            self.profile_tree.yview(*args)

        profile_scroll = ttk.Scrollbar(profile_table, orient="vertical", command=_on_profile_scroll)
        self.profile_tree.configure(yscrollcommand=profile_scroll.set)
        self.profile_tree.grid(row=0, column=0, sticky="nsew")
        profile_scroll.grid(row=0, column=1, sticky="ns")
        profile_table.grid_rowconfigure(0, weight=1)
        profile_table.grid_columnconfigure(0, weight=1)
        self.profile_tree.tag_configure("status_ok", foreground="green")
        self.profile_tree.tag_configure("status_err", foreground="red")
        self.profile_tree.tag_configure("status_work", foreground="#0F766E")
        self.profile_tree.tag_configure("status_warn", foreground="#B45309")
        self.profile_tree.tag_configure("status_flash", background="#E8F5FF")
        self.profile_tree.tag_configure("status_pulse_a", background="#E8F5FF")
        self.profile_tree.tag_configure("status_pulse_b", background="#E4FBEA")

        self.profile_tree.bind("<Button-1>", self._on_profile_tree_click)
        self.profile_tree.bind("<B1-Motion>", self._on_profile_tree_drag)
        self.profile_tree.bind("<ButtonRelease-1>", self._on_profile_tree_release)
        self.profile_tree.bind("<Button-3>", self._on_profile_tree_right_click)
        self.profile_tree.bind("<MouseWheel>", lambda e, t=self.profile_tree: self._mark_user_scroll(t))
        self.profile_tree.bind("<Button-4>", lambda e, t=self.profile_tree: self._mark_user_scroll(t))
        self.profile_tree.bind("<Button-5>", lambda e, t=self.profile_tree: self._mark_user_scroll(t))

        fb_top = ttk.Frame(self.tab_fb)
        fb_top.pack(fill="x", padx=8, pady=(8, 0))
        self.btn_import_fb = ttk.Button(fb_top, text="IMPORT", command=self._import_active_upload_accounts)
        self.btn_import_fb.pack(side="left")
        ttk.Button(fb_top, text="Select All", command=self._select_all_active_upload_accounts).pack(side="left", padx=(8, 4))
        ttk.Button(fb_top, text="Deselect All", command=self._deselect_all_active_upload_accounts).pack(side="left")
        ttk.Button(fb_top, text="EXPORT", command=self._export_active_upload_accounts).pack(side="left", padx=(8, 4))
        ttk.Button(fb_top, text="SCAN", command=self._scan_active_upload_accounts).pack(side="left")
        ttk.Label(fb_top, text="Channel:", style="Subtle.TLabel").pack(side="left", padx=(14, 6))
        ttk.Radiobutton(
            fb_top,
            text="YTB",
            value="ytb",
            style="Channel.TRadiobutton",
            variable=self.active_channel,
            command=self._on_active_channel_changed,
        ).pack(side="left", padx=(0, 10))
        ttk.Radiobutton(
            fb_top,
            text="FB",
            value="fb",
            style="Channel.TRadiobutton",
            variable=self.active_channel,
            command=self._on_active_channel_changed,
        ).pack(side="left")

        fb_table = ttk.Frame(self.tab_fb)
        fb_table.pack(fill="both", expand=True, padx=8, pady=8)
        self.fb_tree = ttk.Treeview(
            fb_table,
            columns=("chk", "stt", "email", "pass", "status", "posts", "followers", "proxy", "facebook", "profile_url", "profile_id"),
            show="headings",
            selectmode="extended",
        )
        self.fb_tree.heading("chk", text="v")
        self.fb_tree.column("chk", width=40, anchor="center")
        self.fb_tree.heading("stt", text="STT")
        self.fb_tree.column("stt", width=50, anchor="center")
        self.fb_tree.heading("email", text="EMAIL")
        self.fb_tree.column("email", width=240)
        self.fb_tree.heading("pass", text="PASS")
        self.fb_tree.column("pass", width=130)
        self.fb_tree.heading("status", text="STATUS")
        self.fb_tree.column("status", width=200)
        self.fb_tree.heading("posts", text="POSTS", command=lambda: self._toggle_fb_sort("posts"))
        self.fb_tree.column("posts", width=70, anchor="center")
        self.fb_tree.heading("followers", text="FOLLOWERS", command=lambda: self._toggle_fb_sort("followers"))
        self.fb_tree.column("followers", width=90, anchor="center")
        self.fb_tree.heading("proxy", text="PROXY")
        self.fb_tree.column("proxy", width=260)
        self.fb_tree.heading("facebook", text="FB REELS")
        self.fb_tree.column("facebook", width=320)
        self.fb_tree.heading("profile_url", text="PROFILE URL")
        self.fb_tree.column("profile_url", width=260)
        self.fb_tree.heading("profile_id", text="PROFILE ID")
        self.fb_tree.column("profile_id", width=240)

        def _on_fb_scroll(*args):
            self._mark_user_scroll(self.fb_tree)
            self.fb_tree.yview(*args)

        fb_scroll = ttk.Scrollbar(fb_table, orient="vertical", command=_on_fb_scroll)
        all_scroll_x = ttk.Scrollbar(all_table, orient="horizontal", command=self.all_tree.xview)
        fb_scroll_x = ttk.Scrollbar(fb_table, orient="horizontal", command=self.fb_tree.xview)
        self.all_tree.configure(xscrollcommand=all_scroll_x.set)
        self.fb_tree.configure(yscrollcommand=fb_scroll.set, xscrollcommand=fb_scroll_x.set)
        self.fb_tree.grid(row=0, column=0, sticky="nsew")
        all_scroll_x.grid(row=1, column=0, sticky="ew")
        fb_scroll.grid(row=0, column=1, sticky="ns")
        fb_scroll_x.grid(row=1, column=0, sticky="ew")
        fb_table.grid_rowconfigure(0, weight=1)
        fb_table.grid_columnconfigure(0, weight=1)
        self.fb_tree.tag_configure("status_ok", foreground="green")
        self.fb_tree.tag_configure("status_err", foreground="red")
        self.fb_tree.tag_configure("status_work", foreground="#0F766E")
        self.fb_tree.tag_configure("status_warn", foreground="#B45309")
        self.fb_tree.tag_configure("status_flash", background="#E8F5FF")
        self.fb_tree.tag_configure("status_pulse_a", background="#E8F5FF")
        self.fb_tree.tag_configure("status_pulse_b", background="#E4FBEA")

        self.fb_tree.bind("<Button-1>", self._on_fb_tree_click)
        self.fb_tree.bind("<B1-Motion>", self._on_fb_tree_drag)
        self.fb_tree.bind("<ButtonRelease-1>", self._on_fb_tree_release)
        self.fb_tree.bind("<Button-3>", self._on_fb_tree_right_click)
        self.fb_tree.bind("<MouseWheel>", lambda e, t=self.fb_tree: self._mark_user_scroll(t))
        self.fb_tree.bind("<Button-4>", lambda e, t=self.fb_tree: self._mark_user_scroll(t))
        self.fb_tree.bind("<Button-5>", lambda e, t=self.fb_tree: self._mark_user_scroll(t))

        fb_profile_top = ttk.Frame(self.tab_fb_profile)
        fb_profile_top.pack(fill="x", padx=8, pady=(8, 0))
        self.btn_import_fb_profile = ttk.Button(fb_profile_top, text="IMPORT PROFILE", command=self._import_active_profile_accounts)
        self.btn_import_fb_profile.pack(side="left")
        ttk.Button(fb_profile_top, text="CHUYEN DU LIEU", command=self._transfer_active_profile_accounts).pack(side="left", padx=(8, 4))
        ttk.Button(fb_profile_top, text="Select All", command=self._select_all_active_profile_accounts).pack(side="left", padx=(8, 4))
        ttk.Button(fb_profile_top, text="Deselect All", command=self._deselect_all_active_profile_accounts).pack(side="left")
        ttk.Label(fb_profile_top, text="Channel:", style="Subtle.TLabel").pack(side="left", padx=(14, 6))
        ttk.Radiobutton(
            fb_profile_top,
            text="YTB",
            value="ytb",
            style="Channel.TRadiobutton",
            variable=self.active_channel,
            command=self._on_active_channel_changed,
        ).pack(side="left", padx=(0, 10))
        ttk.Radiobutton(
            fb_profile_top,
            text="FB",
            value="fb",
            style="Channel.TRadiobutton",
            variable=self.active_channel,
            command=self._on_active_channel_changed,
        ).pack(side="left")

        fb_profile_table = ttk.Frame(self.tab_fb_profile)
        fb_profile_table.pack(fill="both", expand=True, padx=8, pady=8)
        self.fb_profile_tree = ttk.Treeview(
            fb_profile_table,
            columns=("chk", "stt", "email", "pass", "proxy", "facebook", "status"),
            show="headings",
            selectmode="extended",
        )
        self.fb_profile_tree.heading("chk", text="v")
        self.fb_profile_tree.column("chk", width=40, anchor="center")
        self.fb_profile_tree.heading("stt", text="STT")
        self.fb_profile_tree.column("stt", width=50, anchor="center")
        self.fb_profile_tree.heading("email", text="EMAIL")
        self.fb_profile_tree.column("email", width=240)
        self.fb_profile_tree.heading("pass", text="PASS")
        self.fb_profile_tree.column("pass", width=130)
        self.fb_profile_tree.heading("proxy", text="PROXY")
        self.fb_profile_tree.column("proxy", width=260)
        self.fb_profile_tree.heading("facebook", text="FB PROFILE LINK")
        self.fb_profile_tree.column("facebook", width=320)
        self.fb_profile_tree.heading("status", text="STATUS")
        self.fb_profile_tree.column("status", width=200)

        def _on_fb_profile_scroll(*args):
            self._mark_user_scroll(self.fb_profile_tree)
            self.fb_profile_tree.yview(*args)

        fb_profile_scroll = ttk.Scrollbar(fb_profile_table, orient="vertical", command=_on_fb_profile_scroll)
        self.fb_profile_tree.configure(yscrollcommand=fb_profile_scroll.set)
        self.fb_profile_tree.grid(row=0, column=0, sticky="nsew")
        fb_profile_scroll.grid(row=0, column=1, sticky="ns")
        fb_profile_table.grid_rowconfigure(0, weight=1)
        fb_profile_table.grid_columnconfigure(0, weight=1)
        self.fb_profile_tree.tag_configure("status_ok", foreground="green")
        self.fb_profile_tree.tag_configure("status_err", foreground="red")
        self.fb_profile_tree.tag_configure("status_work", foreground="#0F766E")
        self.fb_profile_tree.tag_configure("status_warn", foreground="#B45309")
        self.fb_profile_tree.tag_configure("status_flash", background="#E8F5FF")
        self.fb_profile_tree.tag_configure("status_pulse_a", background="#E8F5FF")
        self.fb_profile_tree.tag_configure("status_pulse_b", background="#E4FBEA")

        self.fb_profile_tree.bind("<Button-1>", self._on_fb_profile_tree_click)
        self.fb_profile_tree.bind("<B1-Motion>", self._on_fb_profile_tree_drag)
        self.fb_profile_tree.bind("<ButtonRelease-1>", self._on_fb_profile_tree_release)
        self.fb_profile_tree.bind("<Button-3>", self._on_fb_profile_tree_right_click)
        self.fb_profile_tree.bind("<MouseWheel>", lambda e, t=self.fb_profile_tree: self._mark_user_scroll(t))
        self.fb_profile_tree.bind("<Button-4>", lambda e, t=self.fb_profile_tree: self._mark_user_scroll(t))
        self.fb_profile_tree.bind("<Button-5>", lambda e, t=self.fb_profile_tree: self._mark_user_scroll(t))

        interact_top = ttk.Frame(self.tab_interact)
        interact_top.pack(fill="x", padx=8, pady=(8, 0))
        ttk.Label(interact_top, text="Like comments:").pack(side="left")
        self.entry_like_min = ttk.Entry(interact_top, width=4)
        self.entry_like_min.insert(0, "1")
        self.entry_like_min.pack(side="left", padx=(5, 5))
        ttk.Label(interact_top, text="to").pack(side="left")
        self.entry_like_max = ttk.Entry(interact_top, width=4)
        self.entry_like_max.insert(0, "10")
        self.entry_like_max.pack(side="left", padx=(5, 15))
        self.chk_reply = ttk.Checkbutton(interact_top, text="Reply", variable=tk.BooleanVar(value=True))
        self.chk_reply.pack(side="left", padx=(0, 10))
        ttk.Label(interact_top, text="Watch videos:").pack(side="left")
        self.entry_watch_min = ttk.Entry(interact_top, width=4)
        self.entry_watch_min.insert(0, "1")
        self.entry_watch_min.pack(side="left", padx=(5, 5))
        ttk.Label(interact_top, text="to").pack(side="left")
        self.entry_watch_max = ttk.Entry(interact_top, width=4)
        self.entry_watch_max.insert(0, "3")
        self.entry_watch_max.pack(side="left", padx=(5, 15))
        ttk.Label(interact_top, text="Join circles max:").pack(side="left")
        self.entry_join_max = ttk.Entry(interact_top, width=4)
        self.entry_join_max.insert(0, "10")
        self.entry_join_max.pack(side="left", padx=(5, 15))
        self.btn_start_interact = ttk.Button(interact_top, text="START INTERACT", command=self._interact_not_ready)
        self.btn_start_interact.pack(side="left")
        self.btn_start_join = ttk.Button(interact_top, text="START JOIN", command=self.start_join_circles)
        self.btn_start_join.pack(side="left", padx=(8, 0))

        interact_body = ttk.Frame(self.tab_interact)
        interact_body.pack(fill="both", expand=True, padx=8, pady=8)
        ttk.Label(interact_body, text="Paste URLs (one per line):").pack(anchor="w")
        self.interact_urls = tk.Text(interact_body, height=12)
        self.interact_urls.pack(fill="both", expand=True)

        stats_top = ttk.Frame(self.tab_stats)
        stats_top.pack(fill="x", padx=8, pady=(8, 0))
        ttk.Label(stats_top, text="Creator Fund Status", style="Subtle.TLabel").pack(side="left")
        ttk.Label(stats_top, textvariable=self._stats_total_var, style="Status.TLabel").pack(side="left", padx=(18, 0))
        self.btn_stats_refresh = ttk.Button(stats_top, text="REFRESH", command=self._refresh_stats)
        self.btn_stats_refresh.pack(side="right")
        self.btn_stats_check_payment = ttk.Button(
            stats_top, text="CHECK PAYMENT", command=self._stats_check_payment_setup
        )
        self.btn_stats_check_payment.pack(side="right", padx=(0, 8))
        self.btn_stats_check = ttk.Button(stats_top, text="CHECK CREATOR FUND", command=self._stats_check_creator_fund)
        self.btn_stats_check.pack(side="right", padx=(0, 8))

        stats_summary = ttk.Frame(self.tab_stats)
        stats_summary.pack(fill="x", padx=8, pady=(2, 0))
        ttk.Label(stats_summary, textvariable=self._stats_breakdown_var, style="Status.TLabel").pack(side="left")

        stats_table = ttk.Frame(self.tab_stats)
        stats_table.pack(fill="both", expand=True, padx=8, pady=8)
        self.stats_tree = ttk.Treeview(
            stats_table,
            columns=("stt", "source", "email", "followers", "status", "payment"),
            show="headings",
            selectmode="browse",
        )
        self.stats_tree.heading("stt", text="STT")
        self.stats_tree.column("stt", width=50, anchor="center")
        self.stats_tree.heading("source", text="SOURCE")
        self.stats_tree.column("source", width=90, anchor="center")
        self.stats_tree.heading("email", text="EMAIL")
        self.stats_tree.column("email", width=260)
        self.stats_tree.heading("followers", text="FOLLOWERS")
        self.stats_tree.column("followers", width=100, anchor="center")
        self.stats_tree.heading("status", text="STATUS")
        self.stats_tree.column("status", width=160, anchor="center")
        self.stats_tree.heading("payment", text="PAYMENT")
        self.stats_tree.column("payment", width=140, anchor="center")
        self.stats_tree.tag_configure("status_joined", foreground="green")
        self.stats_tree.tag_configure("status_pending", foreground="orange")
        self.stats_tree.tag_configure("status_not_applied", foreground="#6B7280")
        stats_scroll = ttk.Scrollbar(stats_table, orient="vertical", command=self.stats_tree.yview)
        self.stats_tree.configure(yscrollcommand=stats_scroll.set)
        self.stats_tree.grid(row=0, column=0, sticky="nsew")
        stats_scroll.grid(row=0, column=1, sticky="ns")
        stats_table.grid_rowconfigure(0, weight=1)
        stats_table.grid_columnconfigure(0, weight=1)

        # === MANAGE TAB ===
        manage_tabs = ttk.Notebook(self.tab_manage)
        manage_tabs.pack(fill="both", expand=True, padx=8, pady=8)
        self.tab_manage_video = ttk.Frame(manage_tabs)
        self.tab_manage_cookie = ttk.Frame(manage_tabs)
        self.tab_manage_keys = ttk.Frame(manage_tabs)
        manage_tabs.add(self.tab_manage_video, text="VIDEO")
        manage_tabs.add(self.tab_manage_cookie, text="COOKIE")
        manage_tabs.add(self.tab_manage_keys, text="KEYS")

        # --- Manage > Video ---
        manage_main = ttk.Frame(self.tab_manage_video)
        manage_main.pack(fill="both", expand=True, padx=8, pady=8)
        manage_main.columnconfigure(1, weight=1)
        manage_main.rowconfigure(0, weight=1)

        manage_left = ttk.LabelFrame(manage_main, text="Video Emails", width=260)
        manage_left.grid(row=0, column=0, sticky="ns", padx=(0, 8))
        manage_left.columnconfigure(0, weight=1)
        manage_left.rowconfigure(2, weight=1)
        manage_left.grid_propagate(False)

        manage_left_top = ttk.Frame(manage_left)
        manage_left_top.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 6))
        manage_left_top.columnconfigure(0, weight=1)
        ttk.Entry(manage_left_top, textvariable=self._manage_search_var).grid(row=0, column=0, sticky="ew")
        manage_left_actions = ttk.Frame(manage_left_top)
        manage_left_actions.grid(row=0, column=1, padx=(6, 0))
        ttk.Button(
            manage_left_actions,
            text="Clear",
            command=lambda: self._manage_search_var.set(""),
            width=6,
        ).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(
            manage_left_actions,
            text="Refresh",
            command=self._refresh_manage_emails,
            width=7,
        ).grid(row=0, column=1)

        self._manage_email_count_var = tk.StringVar(value="Emails: 0")
        ttk.Label(manage_left, textvariable=self._manage_email_count_var, style="Subtle.TLabel").grid(
            row=1, column=0, sticky="w", padx=8, pady=(0, 6)
        )

        manage_list_frame = ttk.Frame(manage_left)
        manage_list_frame.grid(row=2, column=0, sticky="nsew", padx=8, pady=(0, 8))
        manage_list_frame.columnconfigure(0, weight=1)
        manage_list_frame.rowconfigure(0, weight=1)

        self.manage_email_list = tk.Listbox(manage_list_frame, activestyle="none")
        self.manage_email_list.grid(row=0, column=0, sticky="nsew")
        manage_email_scroll = ttk.Scrollbar(manage_list_frame, orient="vertical", command=self.manage_email_list.yview)
        self.manage_email_list.configure(yscrollcommand=manage_email_scroll.set)
        manage_email_scroll.grid(row=0, column=1, sticky="ns")
        self.manage_email_list.bind("<<ListboxSelect>>", self._on_manage_email_select)
        self.manage_email_list.bind("<KeyRelease>", self._filter_manage_emails)
        self._manage_search_var.trace_add("write", lambda *_: self._filter_manage_emails())

        manage_right = ttk.Frame(manage_main)
        manage_right.grid(row=0, column=1, sticky="nsew")
        manage_right.columnconfigure(0, weight=1)
        manage_right.rowconfigure(1, weight=1)

        manage_top = ttk.Frame(manage_right)
        manage_top.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        manage_top.columnconfigure(0, weight=1)
        self._manage_selected_var = tk.StringVar(value="Email: -")
        ttk.Label(manage_top, textvariable=self._manage_selected_var, style="Status.TLabel").pack(
            side="left", padx=(0, 16)
        )
        self._manage_total_var = tk.StringVar(value="Total: 0")
        self._manage_uploaded_var = tk.StringVar(value="Uploaded: 0")
        self._manage_remaining_var = tk.StringVar(value="Remaining: 0")
        ttk.Label(manage_top, textvariable=self._manage_total_var, style="Status.TLabel").pack(side="left", padx=(0, 12))
        ttk.Label(manage_top, textvariable=self._manage_uploaded_var, style="Status.TLabel").pack(side="left", padx=(0, 12))
        ttk.Label(manage_top, textvariable=self._manage_remaining_var, style="Status.TLabel").pack(side="left")
        ttk.Button(manage_top, text="Reload CSV", command=self._reload_manage_csv).pack(side="right", padx=(6, 0))
        ttk.Button(manage_top, text="Save CSV", command=self._save_manage_csv).pack(side="right")

        manage_table = ttk.Frame(manage_right)
        manage_table.grid(row=1, column=0, sticky="nsew")
        manage_table.rowconfigure(0, weight=1)
        manage_table.columnconfigure(0, weight=1)

        self.manage_tree = ttk.Treeview(
            manage_table,
            columns=("stt", "video_id", "title", "url", "status"),
            show="headings",
            selectmode="extended",
        )
        self.manage_tree.heading("stt", text="STT")
        self.manage_tree.column("stt", width=50, anchor="center")
        self.manage_tree.heading("video_id", text="VIDEO ID")
        self.manage_tree.column("video_id", width=160)
        self.manage_tree.heading("title", text="TITLE")
        self.manage_tree.column("title", width=240)
        self.manage_tree.heading("url", text="URL")
        self.manage_tree.column("url", width=360)
        self.manage_tree.heading("status", text="STATUS")
        self.manage_tree.column("status", width=90, anchor="center")

        manage_scroll = ttk.Scrollbar(manage_table, orient="vertical", command=self.manage_tree.yview)
        self.manage_tree.configure(yscrollcommand=manage_scroll.set)
        self.manage_tree.grid(row=0, column=0, sticky="nsew")
        manage_scroll.grid(row=0, column=1, sticky="ns")
        self.manage_tree.bind("<Button-1>", self._on_manage_tree_click)

        # --- Manage > Cookie ---
        cookie_main = ttk.Frame(self.tab_manage_cookie)
        cookie_main.pack(fill="both", expand=True, padx=8, pady=8)
        cookie_main.columnconfigure(0, weight=1)
        cookie_main.rowconfigure(1, weight=1)

        cookie_top = ttk.Frame(cookie_main)
        cookie_top.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        cookie_top.columnconfigure(0, weight=1)
        ttk.Label(cookie_top, text="Paste cookies (same format as cookies.txt):").grid(
            row=0, column=0, sticky="w"
        )
        ttk.Label(cookie_top, textvariable=self._cookie_status_var, style="Status.TLabel").grid(
            row=0, column=1, sticky="e", padx=(8, 0)
        )

        cookie_actions = ttk.Frame(cookie_top)
        cookie_actions.grid(row=1, column=0, columnspan=2, sticky="e", pady=(6, 0))
        ttk.Button(cookie_actions, text="Save", command=self._save_cookie_from_form).pack(side="right")
        ttk.Button(cookie_actions, text="Reload", command=self._load_cookie_into_form).pack(
            side="right", padx=(0, 6)
        )
        ttk.Button(cookie_actions, text="Clear", command=self._clear_cookie_form).pack(
            side="right", padx=(0, 6)
        )

        self.cookie_text = tk.Text(cookie_main, height=18)
        self.cookie_text.grid(row=1, column=0, sticky="nsew")

        # --- Manage > Keys ---
        keys_main = ttk.Frame(self.tab_manage_keys)
        keys_main.pack(fill="both", expand=True, padx=8, pady=8)
        keys_main.columnconfigure(0, weight=1)
        keys_main.rowconfigure(1, weight=1)

        keys_top = ttk.Frame(keys_main)
        keys_top.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        ttk.Label(keys_top, text="License Key Manager", style="Subtle.TLabel").pack(side="left")
        ttk.Button(keys_top, text="REFRESH", command=self._refresh_license_keys).pack(side="right")
        ttk.Button(keys_top, text="REVOKE", command=self._revoke_selected_license_key).pack(
            side="right", padx=(0, 6)
        )
        ttk.Button(keys_top, text="COPY", command=self._copy_selected_license_key).pack(
            side="right", padx=(0, 6)
        )
        ttk.Button(keys_top, text="CREATE", command=self._create_license_key).pack(
            side="right", padx=(0, 6)
        )

        keys_table = ttk.Frame(keys_main)
        keys_table.grid(row=1, column=0, sticky="nsew")
        keys_table.rowconfigure(0, weight=1)
        keys_table.columnconfigure(0, weight=1)
        self.keys_tree = ttk.Treeview(
            keys_table,
            columns=("key", "status", "hwid", "created", "activated", "last_check"),
            show="headings",
            selectmode="browse",
        )
        self.keys_tree.heading("key", text="KEY")
        self.keys_tree.column("key", width=220)
        self.keys_tree.heading("status", text="STATUS")
        self.keys_tree.column("status", width=90, anchor="center")
        self.keys_tree.heading("hwid", text="HWID")
        self.keys_tree.column("hwid", width=220)
        self.keys_tree.heading("created", text="CREATED")
        self.keys_tree.column("created", width=130, anchor="center")
        self.keys_tree.heading("activated", text="ACTIVATED")
        self.keys_tree.column("activated", width=130, anchor="center")
        self.keys_tree.heading("last_check", text="LAST CHECK")
        self.keys_tree.column("last_check", width=130, anchor="center")
        keys_scroll = ttk.Scrollbar(keys_table, orient="vertical", command=self.keys_tree.yview)
        self.keys_tree.configure(yscrollcommand=keys_scroll.set)
        self.keys_tree.grid(row=0, column=0, sticky="nsew")
        keys_scroll.grid(row=0, column=1, sticky="ns")

        # --- Manage > Sync ---
        sync_main = ttk.Frame(self.tab_manage)
        sync_main.pack_forget()
        self.tab_manage_sync = ttk.Frame(manage_tabs)
        manage_tabs.add(self.tab_manage_sync, text="SYNC")
        sync_body = ttk.Frame(self.tab_manage_sync)
        sync_body.pack(fill="both", expand=True, padx=12, pady=12)
        ttk.Label(sync_body, text="Sync Folder: ScoopzSync (near exe)", style="Subtle.TLabel").pack(anchor="w")
        btn_row = ttk.Frame(sync_body)
        btn_row.pack(anchor="w", pady=(8, 0))
        ttk.Button(btn_row, text="EXPORT SYNC", command=self._export_sync_bundle).pack(side="left")
        ttk.Button(btn_row, text="IMPORT SYNC", command=self._import_sync_bundle).pack(side="left", padx=(8, 0))

        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        self.menu = tk.Menu(self.root, tearoff=0)
        self.menu.add_command(label="Tick selected", command=lambda: self._set_checked_selected(True))
        self.menu.add_command(label="Untick selected", command=lambda: self._set_checked_selected(False))
        self.menu.add_command(label="Tick all", command=self._select_all_accounts)
        self.menu.add_command(label="Untick all", command=self._deselect_all_accounts)
        self.menu.add_separator()
        self.menu.add_command(label="Login selected", command=self.menu_login_selected)
        self.menu.add_command(label="Upload selected", command=self.menu_upload_selected)
        self.menu.add_command(label="Get followers", command=self.menu_follow_selected)
        self.menu.add_separator()
        self.menu.add_command(label="Replace proxy errors", command=lambda: self._replace_proxy_errors("upload"))

        self.profile_menu = tk.Menu(self.root, tearoff=0)
        self.profile_menu.add_command(label="Tick selected", command=lambda: self._set_checked_selected_profile(True))
        self.profile_menu.add_command(label="Untick selected", command=lambda: self._set_checked_selected_profile(False))
        self.profile_menu.add_command(label="Tick all", command=self._select_all_profile_accounts)
        self.profile_menu.add_command(label="Untick all", command=self._deselect_all_profile_accounts)
        self.profile_menu.add_separator()
        self.profile_menu.add_command(label="Open YouTube", command=self.menu_profile_selected)

        self.fb_menu = tk.Menu(self.root, tearoff=0)
        self.fb_menu.add_command(label="Tick selected", command=lambda: self._set_checked_selected_fb(True))
        self.fb_menu.add_command(label="Untick selected", command=lambda: self._set_checked_selected_fb(False))
        self.fb_menu.add_command(label="Tick all", command=self._select_all_fb_accounts)
        self.fb_menu.add_command(label="Untick all", command=self._deselect_all_fb_accounts)
        self.fb_menu.add_separator()
        self.fb_menu.add_command(label="Login selected", command=self.menu_fb_login_selected)
        self.fb_menu.add_command(label="Upload selected", command=self.menu_fb_upload_selected)
        self.fb_menu.add_command(label="Get followers", command=self.menu_fb_follow_selected)
        self.fb_menu.add_command(label="Replace proxy errors", command=lambda: self._replace_proxy_errors("fb"))

        self.fb_profile_menu = tk.Menu(self.root, tearoff=0)
        self.fb_profile_menu.add_command(label="Tick selected", command=lambda: self._set_checked_selected_fb_profile(True))
        self.fb_profile_menu.add_command(label="Untick selected", command=lambda: self._set_checked_selected_fb_profile(False))
        self.fb_profile_menu.add_command(label="Tick all", command=self._select_all_fb_profile_accounts)
        self.fb_profile_menu.add_command(label="Untick all", command=self._deselect_all_fb_profile_accounts)

        self.all_menu = tk.Menu(self.root, tearoff=0)
        self.all_menu.add_command(label="Tick selected", command=lambda: self._set_checked_selected_all(True))
        self.all_menu.add_command(label="Untick selected", command=lambda: self._set_checked_selected_all(False))
        self.all_menu.add_command(label="Tick all", command=self._select_all_all_accounts)
        self.all_menu.add_command(label="Untick all", command=self._deselect_all_all_accounts)
        self.all_menu.add_separator()
        self.all_menu.add_command(label="Login selected", command=self.menu_all_login_selected)
        self.all_menu.add_command(label="Upload selected", command=self.menu_all_upload_selected)
        self.all_menu.add_command(label="Get followers", command=self.menu_all_follow_selected)
        self.all_menu.add_separator()
        self.all_menu.add_command(label="Replace proxy errors", command=self.menu_all_replace_proxy_errors)

        self.log_box = tk.Text(
            log_container,
            height=8,
            state="disabled",
            font=("Consolas", 10),
        )
        self.log_box.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        self._ui_paned.forget(log_container)

    def _show_intro(self) -> None:
        try:
            splash = tk.Toplevel(self.root)
            splash.overrideredirect(True)
            splash.attributes("-alpha", 0.0)
            splash.configure(bg="#0B1220")
            splash.lift()
            splash.attributes("-topmost", True)

            width, height = 520, 260
            try:
                sw = self.root.winfo_screenwidth()
                sh = self.root.winfo_screenheight()
                x = int((sw - width) / 2)
                y = int((sh - height) / 2)
            except Exception:
                x, y = 200, 200
            splash.geometry(f"{width}x{height}+{x}+{y}")

            canvas = tk.Canvas(splash, width=width, height=height, highlightthickness=0)
            canvas.pack(fill="both", expand=True)

            # Subtle gradient background
            for i in range(0, height, 4):
                ratio = i / max(1, height)
                r = int(11 + (20 - 11) * ratio)
                g = int(18 + (30 - 18) * ratio)
                b = int(32 + (45 - 32) * ratio)
                color = f"#{r:02x}{g:02x}{b:02x}"
                canvas.create_rectangle(0, i, width, i + 4, outline=color, fill=color)

            # Decorative rings
            ring1 = canvas.create_oval(-40, 30, 220, 290, outline="#132238", width=2)
            ring2 = canvas.create_oval(300, -30, 620, 290, outline="#102033", width=2)

            # Title block
            canvas.create_text(
                width / 2,
                84,
                text="GPM Multi-Profile Suite",
                fill="#E5ECF5",
                font=("Segoe UI Semibold", 16),
            )
            canvas.create_text(
                width / 2,
                112,
                text="Workspace initializing",
                fill="#94A3B8",
                font=("Segoe UI", 10),
            )

            # Accent line
            canvas.create_line(width / 2 - 60, 132, width / 2 + 60, 132, fill="#1E293B", width=2)

            # SCOOPZ glow
            glow = canvas.create_oval(
                width / 2 - 92,
                150,
                width / 2 + 92,
                214,
                outline="",
                fill="#0B2A3A",
            )
            # SCOOPZ label
            scoopz = canvas.create_text(
                width / 2,
                182,
                text="SCOOPZ",
                fill="#22D3EE",
                font=("Segoe UI Semibold", 24),
            )

            # Floating dots
            dots = [
                canvas.create_oval(70, 170, 76, 176, fill="#22D3EE", outline=""),
                canvas.create_oval(430, 150, 435, 155, fill="#38BDF8", outline=""),
                canvas.create_oval(260, 46, 264, 50, fill="#7DD3FC", outline=""),
            ]

            def _float_phase(phase: float = 0.0):
                if not canvas.winfo_exists():
                    return
                dy = math.sin(phase) * 3
                canvas.move(glow, 0, dy * 0.08)
                canvas.move(scoopz, 0, dy * 0.12)
                canvas.move(dots[0], 0, dy * 0.25)
                canvas.move(dots[1], 0, -dy * 0.2)
                canvas.move(dots[2], 0, dy * 0.18)
                canvas.move(ring1, dy * 0.02, 0)
                canvas.move(ring2, -dy * 0.02, 0)
                self.root.after(30, lambda: _float_phase(phase + 0.18))

            def _fade_in(alpha: float = 0.0):
                alpha = min(alpha + 0.05, 0.95)
                splash.attributes("-alpha", alpha)
                if alpha < 0.95:
                    self.root.after(30, lambda: _fade_in(alpha))
                else:
                    self.root.after(2600, _fade_out)

            def _fade_out(alpha: float = 0.95):
                alpha = max(alpha - 0.05, 0.0)
                splash.attributes("-alpha", alpha)
                if alpha > 0.0:
                    self.root.after(30, lambda: _fade_out(alpha))
                else:
                    try:
                        splash.destroy()
                    except Exception:
                        pass
                    try:
                        self.root.deiconify()
                    except Exception:
                        pass
                    self.root.after(100, self._ensure_license_valid)

            _float_phase()
            _fade_in()
        except Exception:
            try:
                self.root.deiconify()
            except Exception:
                pass
            try:
                self.root.after(100, self._ensure_license_valid)
            except Exception:
                pass

    def _migrate_legacy_data(self) -> None:
        try:
            os.makedirs(DATA_DIR, exist_ok=True)
        except Exception:
            return
        def _db_has_data(path: str) -> bool:
            try:
                if not os.path.exists(path):
                    return False
                conn = sqlite3.connect(path, timeout=2)
                try:
                    cur = conn.execute("SELECT data FROM cache")
                    rows = cur.fetchall()
                    if not rows:
                        return False
                    for (payload,) in rows:
                        try:
                            loaded = json.loads(payload or "[]")
                            if isinstance(loaded, list) and len(loaded) > 0:
                                return True
                        except Exception:
                            continue
                    return False
                finally:
                    conn.close()
            except Exception:
                return False

        def _db_has_any_rows(path: str, table: str) -> bool:
            try:
                if not os.path.exists(path):
                    return False
                conn = sqlite3.connect(path, timeout=2)
                try:
                    cur = conn.execute(f"SELECT 1 FROM {table} LIMIT 1")
                    return cur.fetchone() is not None
                finally:
                    conn.close()
            except Exception:
                return False

        legacy_candidates = [_THIS_DIR, os.getcwd()]
        files_to_copy = [
            "cache.db",
            "license.db",
            "cookies.txt",
            "cookiefb.txt",
            "extra_proxies.txt",
            "fallback_captions.txt",
            "savef_api_config.json",
        ]
        dirs_to_copy = ["logs", "video", "profile_images"]
        for legacy in legacy_candidates:
            if not legacy or not os.path.isdir(legacy):
                continue
            for name in files_to_copy:
                src = os.path.join(legacy, name)
                dst = os.path.join(DATA_DIR, name)
                if not os.path.exists(src):
                    continue
                if name == "cache.db":
                    dst_ok = _db_has_data(dst)
                    if dst_ok:
                        continue
                if name == "license.db":
                    dst_ok = _db_has_any_rows(dst, "keys")
                    try:
                        src_mtime = os.path.getmtime(src)
                        dst_mtime = os.path.getmtime(dst) if os.path.exists(dst) else -1
                    except Exception:
                        src_mtime = -1
                        dst_mtime = -1
                    if dst_ok and dst_mtime >= src_mtime:
                        continue
                if os.path.exists(src):
                    try:
                        shutil.copy2(src, dst)
                    except Exception:
                        pass
            for name in dirs_to_copy:
                src = os.path.join(legacy, name)
                dst = os.path.join(DATA_DIR, name)
                if os.path.isdir(src) and not os.path.exists(dst):
                    try:
                        shutil.copytree(src, dst, dirs_exist_ok=True)
                    except Exception:
                        pass

    def _interact_not_ready(self) -> None:
        self._log("[INTERACT] UI ready. Logic will be added next step.")

    def _get_join_max(self) -> int:
        try:
            val = int(self.entry_join_max.get())
            if val > 0:
                return val
        except Exception:
            pass
        return 10

    def _on_tab_changed(self, _evt=None) -> None:
        try:
            current = self.notebook.nametowidget(self.notebook.select())
            if current in (self.tab_upload, self.tab_profile):
                self.active_channel.set("ytb")
            elif current in (self.tab_fb, self.tab_fb_profile):
                self.active_channel.set("fb")

            if current == self.tab_all:
                self._set_sidebar_active("overview")
            elif current in (self.tab_upload, self.tab_fb):
                self._set_sidebar_active("upvideo")
            elif current in (self.tab_profile, self.tab_fb_profile):
                self._set_sidebar_active("profile")
            elif current == self.tab_interact:
                self._set_sidebar_active("interact")
            elif current == self.tab_stats:
                self._set_sidebar_active("monetization")
            elif current == self.tab_manage:
                self._set_sidebar_active("manage")

            if current == self.tab_stats:
                self._refresh_stats()
            elif current == self.tab_manage:
                self._refresh_manage_emails()
        except Exception:
            pass

    def _set_sidebar_active(self, key: str) -> None:
        self._active_sidebar_key = key
        for btn_key, btn in (self._sidebar_buttons or {}).items():
            try:
                btn.configure(style="SidebarActive.TButton" if btn_key == key else "Sidebar.TButton")
            except Exception:
                pass

    def _sidebar_go(self, key: str, target) -> None:
        self._set_sidebar_active(key)
        if callable(target):
            target()
        else:
            self._select_tab(target)

    def _select_tab(self, tab) -> None:
        try:
            if tab in (self.tab_upload, self.tab_profile):
                self.active_channel.set("ytb")
            elif tab in (self.tab_fb, self.tab_fb_profile):
                self.active_channel.set("fb")
            self.notebook.select(tab)
        except Exception:
            pass

    def _channel(self) -> str:
        ch = (self.active_channel.get() or "ytb").strip().lower()
        return "fb" if ch == "fb" else "ytb"

    def _import_active_upload_accounts(self) -> None:
        if self._channel() == "fb":
            self._select_tab(self.tab_fb)
            self.import_fb_accounts()
        else:
            self._select_tab(self.tab_upload)
            self.import_accounts()

    def _select_all_active_upload_accounts(self) -> None:
        if self._channel() == "fb":
            self._select_all_fb_accounts()
        else:
            self._select_all_accounts()

    def _deselect_all_active_upload_accounts(self) -> None:
        if self._channel() == "fb":
            self._deselect_all_fb_accounts()
        else:
            self._deselect_all_accounts()

    def _import_active_profile_accounts(self) -> None:
        if self._channel() == "fb":
            self._select_tab(self.tab_fb_profile)
            self.import_fb_profile_accounts()
        else:
            self._select_tab(self.tab_profile)
            self.import_profile_accounts()

    def _scan_active_upload_accounts(self) -> None:
        if self._channel() == "fb":
            self._select_tab(self.tab_fb)
        else:
            self._select_tab(self.tab_upload)
        self.start_scan()

    def _export_active_upload_accounts(self) -> None:
        try:
            import openpyxl  # type: ignore
        except Exception:
            messagebox.showerror("Export", "Can phai cai dat openpyxl de xuat file .xlsx")
            return

        channel = self._channel()
        default_name = "fb_accounts_export.xlsx" if channel == "fb" else "ytb_accounts_export.xlsx"
        path = filedialog.asksaveasfilename(
            title="Export accounts to Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return

        def _to_text(v):
            if v is None:
                return ""
            return str(v)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            if channel == "fb":
                ws.title = "FACEBOOK"
                headers = ["STT", "EMAIL", "PASS", "POSTS", "FOLLOWERS", "PROXY", "FACEBOOK", "PROFILE_URL", "PROFILE_ID"]
                ws.append(headers)
                items = list(self.fb_tree.get_children()) if hasattr(self, "fb_tree") else []
                for idx, iid in enumerate(items, start=1):
                    ws.append(
                        [
                            idx,
                            _to_text(self.fb_tree.set(iid, "email")),
                            _to_text(self.fb_tree.set(iid, "pass")),
                            _to_text(self.fb_tree.set(iid, "posts")),
                            _to_text(self.fb_tree.set(iid, "followers")),
                            _to_text(self.fb_tree.set(iid, "proxy")),
                            _to_text(self.fb_tree.set(iid, "facebook")),
                            _to_text(self.fb_tree.set(iid, "profile_url")),
                            _to_text(self.fb_tree.set(iid, "profile_id")),
                        ]
                    )
            else:
                ws.title = "YOUTUBE"
                headers = ["STT", "EMAIL", "PASS", "POSTS", "FOLLOWERS", "PROXY", "YOUTUBE", "PROFILE_URL", "PROFILE_ID"]
                ws.append(headers)
                items = list(self.tree.get_children()) if hasattr(self, "tree") else []
                for idx, iid in enumerate(items, start=1):
                    ws.append(
                        [
                            idx,
                            _to_text(self.tree.set(iid, "email")),
                            _to_text(self.tree.set(iid, "pass")),
                            _to_text(self.tree.set(iid, "posts")),
                            _to_text(self.tree.set(iid, "followers")),
                            _to_text(self.tree.set(iid, "proxy")),
                            _to_text(self.tree.set(iid, "youtube")),
                            _to_text(self.tree.set(iid, "profile_url")),
                            _to_text(self.tree.set(iid, "profile_id")),
                        ]
                    )

            wb.save(path)
            self._log(f"[EXPORT] Saved {len(items)} {channel.upper()} accounts -> {path}")
            messagebox.showinfo("Export", f"Da xuat xong file:\n{path}")
        except Exception as e:
            messagebox.showerror("Export", f"Loi xuat file: {e}")

    def _select_all_active_profile_accounts(self) -> None:
        if self._channel() == "fb":
            self._select_all_fb_profile_accounts()
        else:
            self._select_all_profile_accounts()

    def _deselect_all_active_profile_accounts(self) -> None:
        if self._channel() == "fb":
            self._deselect_all_fb_profile_accounts()
        else:
            self._deselect_all_profile_accounts()

    def _open_upload_panel(self) -> None:
        self._set_sidebar_active("upvideo")
        if self._channel() == "fb":
            self._select_tab(self.tab_fb)
        else:
            self._select_tab(self.tab_upload)

    def _open_profile_panel(self) -> None:
        self._set_sidebar_active("profile")
        if self._channel() == "fb":
            self._select_tab(self.tab_fb_profile)
        else:
            self._select_tab(self.tab_profile)

    def _on_active_channel_changed(self) -> None:
        try:
            current = self.notebook.nametowidget(self.notebook.select())
        except Exception:
            return
        if current in (self.tab_upload, self.tab_fb):
            self._open_upload_panel()
        elif current in (self.tab_profile, self.tab_fb_profile):
            self._open_profile_panel()

    def _followers_to_int(self, val) -> int:
        txt = str(val or "").strip().upper().replace(",", "")
        if not txt:
            return 0
        m = re.match(r"^([0-9]+(?:\.[0-9]+)?)([KMB])?$", txt)
        if m:
            num = float(m.group(1))
            suffix = m.group(2)
            if suffix == "K":
                num *= 1_000
            elif suffix == "M":
                num *= 1_000_000
            elif suffix == "B":
                num *= 1_000_000_000
            return int(num)
        digits = re.sub(r"[^0-9]", "", txt)
        return int(digits) if digits else 0

    def _refresh_stats(self) -> None:
        self._normalize_payment_status_cache()
        # Always reload from DB to avoid stale in-memory status
        try:
            fresh_accounts = self._load_cache_list("accounts")
            fresh_fb_accounts = self._load_cache_list("fb_accounts")
            if isinstance(fresh_accounts, list):
                self.accounts = fresh_accounts
            if isinstance(fresh_fb_accounts, list):
                self.fb_accounts = fresh_fb_accounts
        except Exception:
            pass

        try:
            self.stats_tree.delete(*self.stats_tree.get_children())
        except Exception:
            pass

        rows = []
        for acc in self.accounts:
            status = (acc.get("creator_fund_status") or "").strip().upper() or "UNCHECKED"
            followers = self._followers_to_int(acc.get("followers"))
            if followers < 1000 and status not in {"PENDING", "JOINED"}:
                continue
            payment = self._normalize_payment_status(acc.get("payment_status"))
            email = (acc.get("uid") or "").strip()
            if email:
                rows.append(("YTB", email, followers, status, payment))

        for acc in self.fb_accounts:
            status = (acc.get("creator_fund_status") or "").strip().upper() or "UNCHECKED"
            followers = self._followers_to_int(acc.get("followers"))
            if followers < 1000 and status not in {"PENDING", "JOINED"}:
                continue
            payment = self._normalize_payment_status(acc.get("payment_status"))
            email = (acc.get("uid") or "").strip()
            if email:
                rows.append(("FB", email, followers, status, payment))

        rows.sort(key=lambda r: (r[0], -int(r[2]), r[1]))

        for idx, (source, email, followers, status, payment) in enumerate(rows, start=1):
            tags = ()
            if status == "JOINED":
                tags = ("status_joined",)
            elif status == "PENDING":
                tags = ("status_pending",)
            elif status in ("NOT_APPLIED", "UNCHECKED"):
                tags = ("status_not_applied",)
            self.stats_tree.insert(
                "",
                "end",
                iid=str(idx),
                values=(idx, source, email, followers, status, payment),
                tags=tags,
            )
        self._update_stats_summary()

    def _format_stats_count_map(self, counts: dict, preferred_order: list | None = None) -> str:
        return format_stats_count_map(counts, preferred_order)

    def _update_stats_summary(self) -> None:
        try:
            total = 0
            status_counts = {}
            payment_counts = {}
            for iid in self.stats_tree.get_children():
                total += 1
                status = (self.stats_tree.set(iid, "status") or "").strip().upper() or "UNKNOWN"
                payment = self._normalize_payment_status(self.stats_tree.set(iid, "payment"))
                status_counts[status] = int(status_counts.get(status, 0)) + 1
                payment_counts[payment] = int(payment_counts.get(payment, 0)) + 1

            status_text = self._format_stats_count_map(
                status_counts,
                preferred_order=["JOINED", "PENDING", "UNCHECKED", "NOT_APPLIED", "CHECKING...", "UNKNOWN"],
            )
            payment_text = self._format_stats_count_map(
                payment_counts,
                preferred_order=["SETUP", "NOT_SETUP", "CHECK_ERR"],
            )
            self._stats_total_var.set(f"Total: {total}")
            self._stats_breakdown_var.set(f"Creator: {status_text} | Payment: {payment_text}")
        except Exception:
            pass

    def _find_tree_item_by_email(self, tree: ttk.Treeview, email: str) -> str | None:
        try:
            for iid in tree.get_children():
                if (tree.set(iid, "email") or "").strip() == email:
                    return iid
        except Exception:
            pass
        return None

    def _set_stats_row_checking(self, source: str, email: str, checking: bool, status_text: str = "") -> None:
        def _update() -> None:
            try:
                src = (source or "").strip().upper()
                em = (email or "").strip()
                if not src or not em:
                    return
                target_iid = None
                for iid in self.stats_tree.get_children():
                    row_src = (self.stats_tree.set(iid, "source") or "").strip().upper()
                    row_email = (self.stats_tree.set(iid, "email") or "").strip()
                    if row_src == src and row_email == em:
                        target_iid = iid
                        break
                if not target_iid:
                    return
                if checking:
                    # Runtime progress text while checking creator fund.
                    try:
                        self.stats_tree.set(target_iid, "status", (status_text or "CHECKING...").strip())
                    except Exception:
                        pass
                    self._update_stats_summary()
                    try:
                        self.stats_tree.see(target_iid)
                    except Exception:
                        pass
                else:
                    self._refresh_stats()
            except Exception:
                pass
        try:
            self.root.after(0, _update)
        except Exception:
            pass

    def _stats_check_creator_fund(self) -> None:
        if self.executor is not None:
            self._log("[STATS] Dang chay job, hay STOP truoc.")
            return

        # Always work from Monetization rows so behavior matches what user sees.
        selected = self.stats_tree.selection()
        rows = []
        items = selected if selected else self.stats_tree.get_children()
        for iid in items:
            try:
                source = (self.stats_tree.set(iid, "source") or "").strip().upper()
                email = (self.stats_tree.set(iid, "email") or "").strip()
                status = (self.stats_tree.set(iid, "status") or "").strip().upper()
                # Only run check for UNCHECKED/PENDING rows as requested.
                if status not in {"UNCHECKED", "PENDING"}:
                    continue
                if source and email:
                    rows.append((source, email))
            except Exception:
                continue

        if not rows:
            self._log("[STATS] No UNCHECKED/PENDING row to check.")
            return

        try:
            max_threads = max(1, int(self.entry_threads.get() or 1))
        except Exception:
            max_threads = 1

        tasks = []

        # Use same window tiling strategy as upload/payment flows.
        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080
        gap = 6
        taskbar_h = 40
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = max(1, min(len(rows), max_threads))
        cols = min(5, active_count)
        rows_layout = min(2, max(1, math.ceil(active_count / cols)))
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))

        for idx, (source, email) in enumerate(rows):
            acc = None
            item_id = None
            if source == "FB":
                for a in self.fb_accounts:
                    if (a.get("uid") or "").strip() == email:
                        acc = a
                        break
                item_id = self._find_tree_item_by_email(self.fb_tree, email)
                if acc is None or item_id is None:
                    self._log(f"[{email}] CREATOR FUND: skip (not found in FB list)")
                    continue
                pos = idx % (cols * rows_layout)
                col = pos % cols
                row = pos // cols
                x = gap + col * (win_w + gap)
                y = gap + row * (win_h + gap)
                win_pos = f"{x},{y}"
                win_size = f"{win_w},{win_h}"
                tasks.append(("FB", item_id, acc, email, win_pos, win_size))
            else:
                for a in self.accounts:
                    if (a.get("uid") or "").strip() == email:
                        acc = a
                        break
                item_id = self._find_tree_item_by_email(self.tree, email)
                if acc is None or item_id is None:
                    self._log(f"[{email}] CREATOR FUND: skip (not found in YTB list)")
                    continue
                pos = idx % (cols * rows_layout)
                col = pos % cols
                row = pos // cols
                x = gap + col * (win_w + gap)
                y = gap + row * (win_h + gap)
                win_pos = f"{x},{y}"
                win_size = f"{win_w},{win_h}"
                tasks.append(("YTB", item_id, acc, email, win_pos, win_size))

        if not tasks:
            self._log("[STATS] No valid row to check.")
            return

        def _run_creator_one(source: str, item_id: str, acc: dict, email: str, win_pos: str, win_size: str) -> None:
            if self.stop_event.is_set():
                return
            self._set_stats_row_checking(source, email, True, "LOGIN...")
            try:
                if source == "FB":
                    self._fb_follow_only_worker(
                        item_id,
                        acc,
                        win_pos=win_pos,
                        win_size=win_size,
                        perform_creator_fund_check=True,
                        force_creator_fund_check=True,
                        stats_source="FB",
                        stats_email=email,
                        skip_follow_fetch=True,
                    )
                else:
                    self._follow_only_worker(
                        item_id,
                        acc,
                        perform_creator_fund_check=True,
                        force_creator_fund_check=True,
                        stats_source="YTB",
                        stats_email=email,
                        win_pos=win_pos,
                        win_size=win_size,
                        skip_follow_fetch=True,
                    )
            except Exception as e:
                self._log(f"[{email}] CREATOR FUND ERR: {e}")
                self._set_stats_row_checking(source, email, False)

        def _run_creator_tasks() -> None:
            with ThreadPoolExecutor(max_workers=max_threads) as pool:
                futures = [
                    pool.submit(_run_creator_one, source, item_id, acc, email, win_pos, win_size)
                    for source, item_id, acc, email, win_pos, win_size in tasks
                ]
                for fut in as_completed(futures):
                    if self.stop_event.is_set():
                        break
                    try:
                        fut.result()
                    except Exception:
                        pass

        threading.Thread(target=_run_creator_tasks, daemon=True).start()

    def _stats_check_payment_setup(self) -> None:
        if self.executor is not None:
            self._log("[PAYMENT] Dang chay job, hay STOP truoc.")
            return

        try:
            fresh_accounts = self._load_cache_list("accounts")
            fresh_fb_accounts = self._load_cache_list("fb_accounts")
            if isinstance(fresh_accounts, list):
                self.accounts = fresh_accounts
            if isinstance(fresh_fb_accounts, list):
                self.fb_accounts = fresh_fb_accounts
        except Exception:
            pass

        selected = self.stats_tree.selection()
        rows = []
        if selected:
            items = selected
        else:
            items = self.stats_tree.get_children()

        for iid in items:
            try:
                source = (self.stats_tree.set(iid, "source") or "").strip().upper()
                email = (self.stats_tree.set(iid, "email") or "").strip()
                if source and email:
                    rows.append((source, email))
            except Exception:
                continue

        if not rows:
            return

        try:
            max_threads = max(1, int(self.entry_threads.get() or 1))
        except Exception:
            max_threads = 1

        tasks = []

        # Use same window tiling strategy as upload/follow flows.
        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080
        gap = 6
        taskbar_h = 40
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = max(1, min(len(rows), max_threads))
        cols = min(5, active_count)
        rows_layout = min(2, max(1, math.ceil(active_count / cols)))
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))

        for idx, (source, email) in enumerate(rows):
            acc = None
            item_id = None
            if source == "FB":
                for a in self.fb_accounts:
                    if (a.get("uid") or "").strip() == email:
                        acc = a
                        break
                item_id = self._find_tree_item_by_email(self.fb_tree, email)
                if acc is None or item_id is None:
                    self._log(f"[{email}] PAYMENT: skip (not found in FB list)")
                    continue
            else:
                for a in self.accounts:
                    if (a.get("uid") or "").strip() == email:
                        acc = a
                        break
                item_id = self._find_tree_item_by_email(self.tree, email)
                if acc is None or item_id is None:
                    self._log(f"[{email}] PAYMENT: skip (not found in YTB list)")
                    continue

            if (acc.get("payment_status") or "").strip().upper() == "SETUP":
                self._log(f"[{email}] PAYMENT: already SETUP, skip")
                continue

            pos = idx % (cols * rows_layout)
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            win_pos = f"{x},{y}"
            win_size = f"{win_w},{win_h}"

            tasks.append((item_id, acc, source, win_pos, win_size))

        if not tasks:
            self._log("[PAYMENT] No valid row to check.")
            return

        def _run_payment_one(item_id: str, acc: dict, source: str, win_pos: str, win_size: str) -> None:
            if self.stop_event.is_set():
                return
            try:
                self._payment_check_worker(item_id, acc, source, win_pos, win_size)
            except Exception as e:
                email = (acc.get("uid") or "") if isinstance(acc, dict) else ""
                self._log(f"[{email}] PAYMENT ERR: {e}")

        def _run_payment_tasks() -> None:
            with ThreadPoolExecutor(max_workers=max_threads) as pool:
                futures = [
                    pool.submit(_run_payment_one, item_id, acc, source, win_pos, win_size)
                    for item_id, acc, source, win_pos, win_size in tasks
                ]
                for fut in as_completed(futures):
                    if self.stop_event.is_set():
                        break
                    try:
                        fut.result()
                    except Exception:
                        pass

        threading.Thread(target=_run_payment_tasks, daemon=True).start()

    # === Manage tab helpers ===
    def _email_to_folder(self, email: str) -> str:
        return (
            (email or "unknown")
            .strip()
            .replace("@", "_at_")
            .replace(".", "_")
            .replace(":", "_")
            .replace("/", "_")
            .replace("\\", "_")
        )

    def _folder_to_email(self, folder: str) -> str:
        name = (folder or "").strip()
        return name.replace("_at_", "@").replace("_", ".")

    def _refresh_manage_emails(self) -> None:
        video_root = os.path.join(DATA_DIR, "video")
        items = []
        if os.path.isdir(video_root):
            for entry in sorted(os.listdir(video_root)):
                folder_path = os.path.join(video_root, entry)
                if not os.path.isdir(folder_path):
                    continue
                csv_path = os.path.join(folder_path, "shorts.csv")
                if not os.path.exists(csv_path):
                    continue
                email = self._folder_to_email(entry)
                items.append({"email": email, "folder": entry, "csv_path": csv_path})

        self._manage_email_all = items
        self._apply_manage_filter(select_first=True)

    def _on_manage_email_select(self, _evt=None) -> None:
        try:
            sel = self.manage_email_list.curselection()
            if not sel:
                return
            idx = int(sel[0])
        except Exception:
            return
        if idx < 0 or idx >= len(self._manage_email_map):
            return
        item = self._manage_email_map[idx]
        self._manage_email = item["email"]
        self._manage_csv_path = item["csv_path"]
        self._load_manage_csv()
        self._manage_selected_var.set(f"Email: {self._manage_email}")

    def _reload_manage_csv(self) -> None:
        self._load_manage_csv()

    def _load_manage_csv(self) -> None:
        self.manage_tree.delete(*self.manage_tree.get_children())
        self._manage_rows = []
        self._manage_fieldnames = []
        if not self._manage_csv_path or not os.path.exists(self._manage_csv_path):
            self._update_manage_counts()
            return
        try:
            with open(self._manage_csv_path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                self._manage_fieldnames = reader.fieldnames or ["video_id", "title", "url", "status"]
                for row in reader:
                    self._manage_rows.append(
                        {
                            "video_id": (row.get("video_id") or "").strip(),
                            "title": (row.get("title") or "").strip(),
                            "url": (row.get("url") or "").strip(),
                            "status": (row.get("status") or "").strip(),
                        }
                    )
        except Exception as e:
            self._log(f"[MANAGE] Read CSV error: {e}")
            self._update_manage_counts()
            return

        for idx, row in enumerate(self._manage_rows, start=1):
            self.manage_tree.insert(
                "",
                "end",
                iid=str(idx),
                values=(idx, row["video_id"], row["title"], row["url"], row["status"]),
            )
        self._update_manage_counts()

    def _save_manage_csv(self) -> None:
        if not self._manage_csv_path:
            return
        fieldnames = self._manage_fieldnames or ["video_id", "title", "url", "status"]
        try:
            with open(self._manage_csv_path, "w", encoding="utf-8", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                for row in self._manage_rows:
                    writer.writerow({k: row.get(k, "") for k in fieldnames})
            self._log(f"[MANAGE] Saved CSV: {os.path.basename(self._manage_csv_path)}")
        except Exception as e:
            self._log(f"[MANAGE] Save CSV error: {e}")

    def _update_manage_counts(self) -> None:
        total = len(self._manage_rows)
        uploaded = 0
        for row in self._manage_rows:
            if (row.get("status") or "").strip().lower() == "true":
                uploaded += 1
        remaining = total - uploaded
        self._manage_total_var.set(f"Total: {total}")
        self._manage_uploaded_var.set(f"Uploaded: {uploaded}")
        self._manage_remaining_var.set(f"Remaining: {remaining}")

    def _apply_manage_filter(self, select_first: bool = False) -> None:
        query = (self._manage_search_var.get() or "").strip().lower()
        if query:
            filtered = [item for item in self._manage_email_all if query in item["email"].lower()]
        else:
            filtered = list(self._manage_email_all)
        self._manage_email_map = filtered
        self.manage_email_list.delete(0, tk.END)
        for item in filtered:
            self.manage_email_list.insert(tk.END, item["email"])
        self._manage_email_count_var.set(f"Emails: {len(filtered)}")

        if select_first and filtered:
            self.manage_email_list.selection_set(0)
            self._on_manage_email_select()
        elif not filtered:
            self._manage_email = ""
            self._manage_rows = []
            self._manage_fieldnames = []
            self._manage_csv_path = ""
            self._manage_selected_var.set("Email: -")
            self._update_manage_counts()
            self.manage_tree.delete(*self.manage_tree.get_children())

    def _filter_manage_emails(self, _evt=None) -> None:
        self._apply_manage_filter(select_first=False)

    def _on_manage_tree_click(self, event) -> None:
        region = self.manage_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        column = self.manage_tree.identify_column(event.x)
        row = self.manage_tree.identify_row(event.y)
        if row:
            try:
                col_idx = int(column[1:]) - 1
                col_name = self.manage_tree["columns"][col_idx]
            except Exception:
                col_name = ""
            if col_name in {"video_id", "title", "url", "status"}:
                self._begin_cell_edit(self.manage_tree, row, col_name)

    # === Cookie helpers ===
    def _cookie_db_key(self) -> str:
        return "cookies_text"

    def _get_cookies_file_path(self) -> str:
        return os.path.join(DATA_DIR, COOKIES_FILE)

    def _read_cookies_file(self) -> str:
        path = self._get_cookies_file_path()
        if not os.path.exists(path):
            return ""
        try:
            with open(path, "r", encoding="utf-8") as f:
                return f.read()
        except Exception:
            return ""

    def _write_cookies_file(self, content: str) -> None:
        path = self._get_cookies_file_path()
        try:
            with open(path, "w", encoding="utf-8") as f:
                f.write(content or "")
        except Exception:
            pass

    def _load_cookie_from_db(self) -> str:
        try:
            with self._cache_db_lock:
                conn = self._open_cache_db()
                if not conn:
                    return ""
                try:
                    cur = conn.execute("SELECT data FROM cache WHERE key = ?", (self._cookie_db_key(),))
                    row = cur.fetchone()
                    if not row or not row[0]:
                        return ""
                    try:
                        payload = json.loads(row[0])
                        if isinstance(payload, dict):
                            return str(payload.get("value") or "")
                    except Exception:
                        return str(row[0] or "")
                finally:
                    conn.close()
        except Exception:
            return ""
        return ""

    def _save_cookie_to_db(self, content: str) -> None:
        try:
            payload = json.dumps({"value": content or ""}, ensure_ascii=False)
            with self._cache_db_lock:
                conn = self._open_cache_db()
                if not conn:
                    return
                try:
                    conn.execute(
                        "INSERT OR REPLACE INTO cache (key, data, updated_at) VALUES (?, ?, ?)",
                        (self._cookie_db_key(), payload, time.time()),
                    )
                    conn.commit()
                finally:
                    conn.close()
        except Exception:
            pass

    def _load_cookie_into_form(self) -> None:
        if not hasattr(self, "cookie_text"):
            return
        text = self._load_cookie_from_db()
        source = "DB"
        if not text:
            text = self._read_cookies_file()
            source = "FILE" if text else "-"
        self.cookie_text.delete("1.0", tk.END)
        if text:
            self.cookie_text.insert("1.0", text)
        self._cookie_status_var.set(f"Cookie: {source}")

    def _save_cookie_from_form(self) -> None:
        if not hasattr(self, "cookie_text"):
            return
        content = self.cookie_text.get("1.0", tk.END).strip()
        self._save_cookie_to_db(content)
        self._write_cookies_file(content)
        self._cookie_status_var.set("Cookie: SAVED")
        self._log("[COOKIE] Saved to DB and cookies.txt")

    def _clear_cookie_form(self) -> None:
        if not hasattr(self, "cookie_text"):
            return
        self.cookie_text.delete("1.0", tk.END)
        self._cookie_status_var.set("Cookie: -")

    # === License server / key helpers ===
    def _start_license_server(self) -> None:
        if not LICENSE_SERVER_ENABLED:
            return
        if self._license_server is not None:
            return
        try:
            db_path = os.path.join(DATA_DIR, "license.db")
            self._license_server = start_license_server(
                LICENSE_SERVER_HOST,
                int(LICENSE_SERVER_PORT),
                db_path,
                LICENSE_ADMIN_TOKEN,
                log_func=lambda m: self._log(m),
            )
            self._log(f"[LICENSE] Server started on {LICENSE_SERVER_HOST}:{LICENSE_SERVER_PORT}")
        except Exception as e:
            self._log(f"[LICENSE] Server start failed: {e}")

    def _get_hwid(self) -> str:
        try:
            import winreg  # type: ignore

            with winreg.OpenKey(
                winreg.HKEY_LOCAL_MACHINE,
                r"SOFTWARE\Microsoft\Cryptography",
                0,
                winreg.KEY_READ | winreg.KEY_WOW64_64KEY,
            ) as key:
                guid, _ = winreg.QueryValueEx(key, "MachineGuid")
                return str(guid)
        except Exception:
            try:
                import uuid

                return str(uuid.getnode())
            except Exception:
                return ""

    def _license_request(self, method: str, path: str, payload: dict | None = None, admin: bool = False) -> dict:
        url = LICENSE_SERVER_URL.rstrip("/") + path
        headers = {"Content-Type": "application/json"}
        if admin and LICENSE_ADMIN_TOKEN:
            headers["X-Admin-Token"] = LICENSE_ADMIN_TOKEN
        try:
            if method.upper() == "GET":
                r = requests.get(url, headers=headers, timeout=4)
            else:
                r = requests.post(url, headers=headers, json=payload or {}, timeout=6)
            if r.status_code >= 400:
                return {"ok": False, "reason": f"HTTP_{r.status_code}"}
            return r.json()
        except Exception as e:
            return {"ok": False, "reason": str(e)}

    def _load_license_key(self) -> str:
        try:
            with self._cache_db_lock:
                conn = self._open_cache_db()
                if not conn:
                    return ""
                try:
                    cur = conn.execute("SELECT data FROM cache WHERE key = ?", ("license_key",))
                    row = cur.fetchone()
                    if not row or not row[0]:
                        return ""
                    payload = json.loads(row[0])
                    if isinstance(payload, dict):
                        return str(payload.get("value") or "")
                    return str(row[0] or "")
                finally:
                    conn.close()
        except Exception:
            return ""

    def _save_license_key(self, key: str) -> None:
        try:
            payload = json.dumps({"value": key or ""}, ensure_ascii=False)
            with self._cache_db_lock:
                conn = self._open_cache_db()
                if not conn:
                    return
                try:
                    conn.execute(
                        "INSERT OR REPLACE INTO cache (key, data, updated_at) VALUES (?, ?, ?)",
                        ("license_key", payload, time.time()),
                    )
                    conn.commit()
                finally:
                    conn.close()
        except Exception:
            pass

    def _clear_license_key(self) -> None:
        try:
            with self._cache_db_lock:
                conn = self._open_cache_db()
                if not conn:
                    return
                try:
                    conn.execute("DELETE FROM cache WHERE key = ?", ("license_key",))
                    conn.commit()
                finally:
                    conn.close()
        except Exception:
            pass

    def _verify_license_key(self, key: str) -> tuple[bool, str]:
        hwid = self._get_hwid()
        resp = self._license_request("POST", "/verify", {"key": key, "hwid": hwid})
        if resp.get("ok"):
            return True, "OK"
        return False, resp.get("reason") or "INVALID"

    def _ensure_license_valid(self) -> None:
        if self._is_admin_machine():
            self._license_valid = True
            return
        key = self._load_license_key()
        if key:
            ok, _ = self._verify_license_key(key)
            if ok:
                self._license_valid = True
                return
        self._show_license_prompt()

    def _show_license_prompt(self) -> None:
        win = tk.Toplevel(self.root)
        win.title("License Key")
        w, h = 420, 210
        try:
            sw = self.root.winfo_screenwidth()
            sh = self.root.winfo_screenheight()
            x = int((sw - w) / 2)
            y = int((sh - h) / 2)
        except Exception:
            x, y = 200, 200
        win.geometry(f"{w}x{h}+{x}+{y}")
        win.resizable(False, False)
        win.grab_set()
        win.transient(self.root)

        ttk.Label(win, text="Enter license key:", style="Subtle.TLabel").pack(
            anchor="w", padx=16, pady=(16, 6)
        )
        entry = ttk.Entry(win, width=44)
        entry.pack(padx=16, pady=(0, 10))
        msg_var = tk.StringVar(value="")
        ttk.Label(win, textvariable=msg_var, style="Status.TLabel").pack(
            anchor="w", padx=16, pady=(0, 6)
        )

        btn_row = ttk.Frame(win)
        btn_row.pack(fill="x", padx=16, pady=(6, 12))

        def _on_verify():
            k = (entry.get() or "").strip()
            if not k:
                msg_var.set("Key is required.")
                return
            ok, reason = self._verify_license_key(k)
            if ok:
                self._save_license_key(k)
                self._license_valid = True
                msg_var.set("Key verified.")
                win.after(200, win.destroy)
            else:
                msg_var.set(f"Invalid key: {reason}")

        def _on_clear():
            self._clear_license_key()
            entry.delete(0, tk.END)
            msg_var.set("Local key cleared.")

        ttk.Button(btn_row, text="VERIFY", command=_on_verify).pack(side="right")
        ttk.Button(btn_row, text="CLEAR", command=_on_clear).pack(side="right", padx=(0, 6))

        entry.focus_set()

    def _require_license_or_warn(self) -> bool:
        if self._is_admin_machine():
            return True
        if self._license_valid:
            return True
        key = self._load_license_key()
        if not key:
            messagebox.showwarning("License", "Chua co key. Vui long nhap key.")
            self._show_license_prompt()
            return False
        ok, reason = self._verify_license_key(key)
        if ok:
            self._license_valid = True
            return True
        messagebox.showwarning("License", f"Key khong hop le: {reason}")
        self._show_license_prompt()
        return False

    def _is_admin_machine(self) -> bool:
        try:
            if not LICENSE_SERVER_ENABLED:
                return False
            url = (LICENSE_SERVER_URL or "").lower()
            return "127.0.0.1" in url or "localhost" in url
        except Exception:
            return False

    def _format_ts(self, ts: float | None) -> str:
        if not ts:
            return "-"
        try:
            return time.strftime("%Y-%m-%d %H:%M", time.localtime(ts))
        except Exception:
            return "-"

    def _refresh_license_keys(self) -> None:
        if not hasattr(self, "keys_tree"):
            return
        self.keys_tree.delete(*self.keys_tree.get_children())

        def _task():
            resp = self._license_request("GET", "/keys", admin=True)
            if not resp.get("ok"):
                self._log(f"[LICENSE] List keys failed: {resp.get('reason')}")
                return
            rows = resp.get("keys", [])

            def _update():
                try:
                    self.keys_tree.delete(*self.keys_tree.get_children())
                    for item in rows:
                        self.keys_tree.insert(
                            "",
                            "end",
                            values=(
                                item.get("key", ""),
                                item.get("status", ""),
                                item.get("hwid", ""),
                                self._format_ts(item.get("created_at")),
                                self._format_ts(item.get("activated_at")),
                                self._format_ts(item.get("last_check_at")),
                            ),
                        )
                except Exception:
                    pass

            self.root.after(0, _update)

        threading.Thread(target=_task, daemon=True).start()

    def _create_license_key(self) -> None:
        def _task():
            resp = self._license_request("POST", "/keys/create", admin=True, payload={"note": ""})
            if not resp.get("ok"):
                self._log(f"[LICENSE] Create key failed: {resp.get('reason')}")
                return
            key = resp.get("key", "")
            def _update():
                try:
                    self.root.clipboard_clear()
                    self.root.clipboard_append(key)
                except Exception:
                    pass
                self._refresh_license_keys()
            self.root.after(0, _update)
            self._log(f"[LICENSE] Created key: {key}")

        threading.Thread(target=_task, daemon=True).start()

    def _revoke_selected_license_key(self) -> None:
        if not hasattr(self, "keys_tree"):
            return
        sel = self.keys_tree.selection()
        if not sel:
            return
        key = (self.keys_tree.set(sel[0], "key") or "").strip()
        if not key:
            return
        def _task():
            resp = self._license_request("POST", "/keys/revoke", admin=True, payload={"key": key})
            if not resp.get("ok"):
                self._log(f"[LICENSE] Revoke failed: {resp.get('reason')}")
                return
            self.root.after(0, self._refresh_license_keys)

        threading.Thread(target=_task, daemon=True).start()

    def _copy_selected_license_key(self) -> None:
        if not hasattr(self, "keys_tree"):
            return
        sel = self.keys_tree.selection()
        if not sel:
            return
        key = (self.keys_tree.set(sel[0], "key") or "").strip()
        if not key:
            return
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(key)
            self._log(f"[LICENSE] Copied key: {key}")
        except Exception:
            pass

    # === Sync helpers ===
    def _sync_dir(self) -> str:
        base = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else _THIS_DIR
        path = os.path.join(base, "ScoopzSync")
        os.makedirs(path, exist_ok=True)
        return path

    def _center_toplevel(self, win: tk.Toplevel, width: int, height: int) -> None:
        try:
            win.update_idletasks()
            screen_w = win.winfo_screenwidth()
            screen_h = win.winfo_screenheight()
            x = int((screen_w - width) / 2)
            y = int((screen_h - height) / 2)
            win.geometry(f"{width}x{height}+{x}+{y}")
        except Exception:
            pass

    def _open_sync_progress(
        self, title: str, message: str
    ) -> tuple[tk.Toplevel, tk.StringVar, tk.IntVar, tk.StringVar]:
        win = tk.Toplevel(self.root)
        win.title(title)
        win.transient(self.root)
        win.grab_set()
        win.resizable(False, False)
        width, height = 420, 140
        self._center_toplevel(win, width, height)
        msg_var = tk.StringVar(value=message)
        pct_var = tk.IntVar(value=0)
        pct_text_var = tk.StringVar(value="0%")
        body = ttk.Frame(win)
        body.pack(fill="both", expand=True, padx=16, pady=16)
        ttk.Label(body, textvariable=msg_var, style="Subtle.TLabel").pack(anchor="w")
        ttk.Progressbar(
            body,
            orient="horizontal",
            mode="determinate",
            maximum=100,
            variable=pct_var,
            style="Accent.Horizontal.TProgressbar",
        ).pack(fill="x", pady=(12, 4))
        ttk.Label(body, textvariable=pct_text_var, style="Subtle.TLabel").pack(anchor="e")
        return win, msg_var, pct_var, pct_text_var

    def _update_sync_progress(
        self,
        win: tk.Toplevel,
        msg_var: tk.StringVar,
        pct_var: tk.IntVar,
        pct_text_var: tk.StringVar,
        percent: int | None,
        message: str | None,
    ) -> None:
        def _apply() -> None:
            if not win.winfo_exists():
                return
            if message is not None:
                msg_var.set(message)
            if percent is not None:
                safe = max(0, min(100, int(percent)))
                pct_var.set(safe)
                pct_text_var.set(f"{safe}%")
        self.root.after(0, _apply)

    def _collect_video_files(self, src_dir: str, dst_dir: str) -> list[tuple[str, str]]:
        files: list[tuple[str, str]] = []
        if not os.path.isdir(src_dir):
            return files
        for root, _dirs, filenames in os.walk(src_dir):
            for name in filenames:
                src = os.path.join(root, name)
                rel = os.path.relpath(src, src_dir)
                dst = os.path.join(dst_dir, rel)
                files.append((src, dst))
        return files

    def _ask_directory(self, title: str, initial_dir: str, must_exist: bool = False) -> str:
        # Keep for potential future use; not used by sync popup
        try:
            return filedialog.askdirectory(
                parent=self.root,
                title=title,
                initialdir=initial_dir,
                mustexist=must_exist,
            )
        except Exception:
            return ""

    def _prompt_sync_dir(self, title: str, must_exist: bool) -> str:
        selected = {"path": ""}
        win = tk.Toplevel(self.root)
        win.title(title)
        win.transient(self.root)
        win.resizable(False, False)
        win.minsize(560, 200)
        self._center_toplevel(win, 600, 220)
        body = ttk.Frame(win)
        body.pack(fill="both", expand=True, padx=14, pady=14)
        ttk.Label(body, text=title, style="Subtle.TLabel").pack(anchor="w")
        row = ttk.Frame(body)
        row.pack(fill="x", pady=(10, 8))
        path_var = tk.StringVar(value=self._sync_dir())
        entry = ttk.Entry(row, textvariable=path_var)
        entry.pack(side="left", fill="x", expand=True)
        status_var = tk.StringVar(value="")

        def _browse() -> None:
            status_var.set("Opening folder picker...")
            browse_btn.configure(state="disabled")

            def _worker() -> None:
                picked = ""
                try:
                    import tempfile

                    desc = (title or "").replace('"', '""')
                    sel = (path_var.get() or self._sync_dir()).replace('"', '""')
                    script = (
                        'Add-Type -AssemblyName System.Windows.Forms\n'
                        '$f = New-Object System.Windows.Forms.FolderBrowserDialog\n'
                        f'$f.Description = "{desc}"\n'
                        f'$f.SelectedPath = "{sel}"\n'
                        'if ($f.ShowDialog() -eq [System.Windows.Forms.DialogResult]::OK) '
                        '{ Write-Output $f.SelectedPath }\n'
                    )
                    with tempfile.NamedTemporaryFile(
                        mode="w", delete=False, suffix=".ps1", encoding="utf-8"
                    ) as tf:
                        tf.write(script)
                        ps1_path = tf.name
                    try:
                        res = subprocess.run(
                            [
                                "powershell",
                                "-NoProfile",
                                "-ExecutionPolicy",
                                "Bypass",
                                "-File",
                                ps1_path,
                            ],
                            capture_output=True,
                            text=True,
                            timeout=120,
                        )
                        picked = (res.stdout or "").strip()
                    finally:
                        try:
                            os.remove(ps1_path)
                        except Exception:
                            pass
                except Exception:
                    picked = ""

                def _apply() -> None:
                    if picked:
                        path_var.set(picked)
                        status_var.set("")
                    else:
                        status_var.set("Picker closed.")
                    browse_btn.configure(state="normal")

                self.root.after(0, _apply)

            threading.Thread(target=_worker, daemon=True).start()

        def _paste_path() -> None:
            try:
                clip = (self.root.clipboard_get() or "").strip()
            except Exception:
                clip = ""
            if clip:
                path_var.set(clip)
                status_var.set("Pasted path from clipboard.")
            else:
                status_var.set("Clipboard is empty.")

        browse_btn = ttk.Button(row, text="Browse", command=_browse)
        browse_btn.pack(side="left", padx=(8, 0))
        ttk.Button(row, text="Paste", command=_paste_path).pack(side="left", padx=(6, 0))
        ttk.Label(body, textvariable=status_var, style="Subtle.TLabel").pack(anchor="w")
        btn_row = ttk.Frame(body)
        btn_row.pack(side="bottom", fill="x", pady=(10, 0))

        def _ok() -> None:
            path = (path_var.get() or "").strip()
            if not path:
                return
            if must_exist and not os.path.isdir(path):
                messagebox.showerror("Sync", "Folder does not exist.")
                return
            selected["path"] = path
            win.destroy()

        def _cancel() -> None:
            selected["path"] = ""
            win.destroy()

        tk.Button(btn_row, text="OK", command=_ok, width=10, bg="#0F766E", fg="white").pack(side="right")
        tk.Button(btn_row, text="Cancel", command=_cancel, width=10).pack(side="right", padx=(0, 8))
        entry.focus_set()
        self.root.wait_window(win)
        return selected["path"]

    def _export_sync_bundle(self) -> None:
        sync_dir = self._prompt_sync_dir("Select export folder", must_exist=False)
        if not sync_dir:
            return
        win, msg_var, pct_var, pct_text_var = self._open_sync_progress(
            "Export Sync",
            "Preparing export...",
        )

        def _worker() -> None:
            try:
                os.makedirs(sync_dir, exist_ok=True)
                ytb_path = os.path.join(sync_dir, "ytb.json")
                fb_path = os.path.join(sync_dir, "fb.json")
                video_src = os.path.join(DATA_DIR, "video")
                video_dst = os.path.join(sync_dir, "video")
                video_files = self._collect_video_files(video_src, video_dst)

                total = max(1, 2 + len(video_files))
                done = 0

                with open(ytb_path, "w", encoding="utf-8") as f:
                    json.dump(self.accounts or [], f, ensure_ascii=False)
                done += 1
                self._update_sync_progress(
                    win,
                    msg_var,
                    pct_var,
                    pct_text_var,
                    done * 100 / total,
                    "Exporting YTB...",
                )

                with open(fb_path, "w", encoding="utf-8") as f:
                    json.dump(self.fb_accounts or [], f, ensure_ascii=False)
                done += 1
                self._update_sync_progress(
                    win,
                    msg_var,
                    pct_var,
                    pct_text_var,
                    done * 100 / total,
                    "Exporting FB...",
                )

                if video_files:
                    for idx, (src, dst) in enumerate(video_files, start=1):
                        os.makedirs(os.path.dirname(dst), exist_ok=True)
                        shutil.copy2(src, dst)
                        done += 1
                        if idx % 20 == 0 or idx == len(video_files):
                            self._update_sync_progress(
                                win,
                                msg_var,
                                pct_var,
                                pct_text_var,
                                done * 100 / total,
                                f"Copying videos {idx}/{len(video_files)}...",
                            )
                        time.sleep(0.003)

                self._log(f"[SYNC] Exported to {sync_dir}")
                self.root.after(0, lambda: messagebox.showinfo("Sync", f"Exported to {sync_dir}"))
            except Exception as e:
                self._log(f"[SYNC] Export error: {e}")
                self.root.after(0, lambda: messagebox.showerror("Sync", f"Export error: {e}"))
            finally:
                self.root.after(0, win.destroy)

        threading.Thread(target=_worker, daemon=True).start()

    def _import_sync_bundle(self) -> None:
        sync_dir = self._prompt_sync_dir("Select import folder", must_exist=True)
        if not sync_dir:
            return
        win, msg_var, pct_var, pct_text_var = self._open_sync_progress(
            "Import Sync",
            "Preparing import...",
        )

        def _worker() -> None:
            try:
                ytb_path = os.path.join(sync_dir, "ytb.json")
                fb_path = os.path.join(sync_dir, "fb.json")
                video_src = os.path.join(sync_dir, "video")
                video_dst = os.path.join(DATA_DIR, "video")
                video_files = self._collect_video_files(video_src, video_dst)

                total = max(1, 2 + len(video_files))
                done = 0

                if os.path.exists(ytb_path):
                    with open(ytb_path, "r", encoding="utf-8") as f:
                        self.accounts = json.load(f) or []
                    self._save_accounts_cache()
                done += 1
                self._update_sync_progress(
                    win,
                    msg_var,
                    pct_var,
                    pct_text_var,
                    done * 100 / total,
                    "Importing YTB...",
                )
                time.sleep(0.05)

                if os.path.exists(fb_path):
                    with open(fb_path, "r", encoding="utf-8") as f:
                        self.fb_accounts = json.load(f) or []
                    self._save_fb_accounts_cache()
                done += 1
                self._update_sync_progress(
                    win,
                    msg_var,
                    pct_var,
                    pct_text_var,
                    done * 100 / total,
                    "Importing FB...",
                )
                time.sleep(0.05)

                if video_files:
                    for idx, (src, dst) in enumerate(video_files, start=1):
                        os.makedirs(os.path.dirname(dst), exist_ok=True)
                        shutil.copy2(src, dst)
                        done += 1
                        if idx % 20 == 0 or idx == len(video_files):
                            self._update_sync_progress(
                                win,
                                msg_var,
                                pct_var,
                                pct_text_var,
                                done * 100 / total,
                                f"Copying videos {idx}/{len(video_files)}...",
                            )
                        time.sleep(0.003)

                def _finish() -> None:
                    try:
                        self.accounts = self._load_accounts_cache() or self.accounts
                        self.fb_accounts = self._load_fb_accounts_cache() or self.fb_accounts
                        self._load_rows()
                        self._load_fb_rows()
                        self._refresh_stats()
                        self._refresh_manage_emails()
                    except Exception:
                        pass
                    self._log(f"[SYNC] Imported from {sync_dir}")
                    messagebox.showinfo("Sync", f"Imported from {sync_dir}")

                self.root.after(0, _finish)
            except Exception as e:
                self._log(f"[SYNC] Import error: {e}")
                self.root.after(0, lambda: messagebox.showerror("Sync", f"Import error: {e}"))
            finally:
                self.root.after(0, win.destroy)

        threading.Thread(target=_worker, daemon=True).start()

    def start_join_circles(self) -> None:
        if self.executor is not None:
            return
        try:
            max_threads = int(self.entry_threads.get())
            if max_threads <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Loi", "So luong phai > 0")
            return

        self.stop_event.clear()
        self.executor = ThreadPoolExecutor(max_workers=max_threads)
        self.login_semaphore = threading.BoundedSemaphore(max_threads)
        checked_emails = self._get_checked_email_set(self.tree)
        if not checked_emails:
            messagebox.showinfo("Thong bao", "Khong co profile nao duoc tick.")
            self.executor = None
            return

        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080
        gap = 6
        taskbar_h = 40
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(checked_emails)
        cols = min(5, active_count)
        rows_layout = min(2, max(1, math.ceil(active_count / cols)))
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))

        join_max = self._get_join_max()
        slot_idx = 0
        max_slots = cols * rows_layout
        email_to_iid = self._map_email_to_item_id(self.tree)
        for acc in self.accounts:
            email = (acc.get("uid") or "").strip()
            if email not in checked_emails:
                continue
            item_id = email_to_iid.get(email)
            if not item_id:
                continue
            pos = slot_idx % max_slots
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            win_pos = f"{x},{y}"
            win_size = f"{win_w},{win_h}"
            self.executor.submit(self._join_circles_worker, item_id, acc, win_pos, win_size, join_max)
            slot_idx += 1

        def _waiter():
            try:
                self.executor.shutdown(wait=True)
            except Exception:
                pass
            self.executor = None

        threading.Thread(target=_waiter, daemon=True).start()

    def _join_circles_worker(self, item_id: str, acc: dict, win_pos: str, win_size: str, join_max: int) -> None:
        if self.stop_event.is_set():
            return
        profile_id = None
        try:
            self._set_status(item_id, "CREATE...")
            ok_c = False
            data_c = {}
            msg_c = ""
            with self.create_lock:
                for attempt in range(3):
                    ok_c, data_c, msg_c = create_profile(acc["uid"], acc["proxy"], SCOOPZ_URL)
                    if ok_c:
                        break
                    wait_s = 5 + attempt * 3
                    self._set_status(item_id, f"CREATE RETRY {attempt+1}/3")
                    self._log(f"[{acc['uid']}] CREATE ERR: {msg_c} | retry in {wait_s}s")
                    time.sleep(wait_s)
            if not ok_c:
                self._set_status(item_id, f"CREATE ERR: {msg_c}")
                self._log(f"[{acc['uid']}] CREATE ERR: {msg_c}")
                self._record_failed(item_id, acc, f"CREATE ERR: {msg_c}")
                return

            if isinstance(data_c, dict):
                profile_id = (data_c.get("data") or {}).get("id") or data_c.get("id") or data_c.get("profile_id")
            if not profile_id:
                self._set_status(item_id, "NO PROFILE ID")
                self._record_failed(item_id, acc, "NO PROFILE ID")
                return
            self._remember_profile_path(profile_id, data_c)
            self.created_profiles.add(profile_id)

            self._set_status(item_id, "START...", profile_id=profile_id)
            ok_s, data_s, msg_s = start_profile(profile_id, win_pos=win_pos, win_size=win_size)
            if not ok_s:
                if self._is_proxy_error(msg_s):
                    try:
                        close_profile(profile_id, 3)
                        delete_profile(profile_id, 10)
                    except Exception:
                        pass
                    try:
                        self.created_profiles.discard(profile_id)
                    except Exception:
                        pass
                    self._delete_profile_path(profile_id)
                    new_id, new_data_s, _err = self._retry_start_profile_with_new_proxy(
                        acc,
                        item_id,
                        "upload",
                        self.tree,
                        lambda s: self._set_status(item_id, s),
                        win_pos=win_pos,
                        win_size=win_size,
                        created_set="created_profiles",
                    )
                    if new_id and new_data_s:
                        profile_id = new_id
                        data_s = new_data_s
                        ok_s = True
                    else:
                        self._set_status(item_id, f"START ERR: {msg_s}")
                        self._log(f"[{acc['uid']}] START ERR: {msg_s}")
                        self._record_failed(item_id, acc, f"START ERR: {msg_s}")
                        return
                else:
                    self._set_status(item_id, f"START ERR: {msg_s}")
                    self._log(f"[{acc['uid']}] START ERR: {msg_s}")
                    self._record_failed(item_id, acc, f"START ERR: {msg_s}")
                    return

            with self.active_lock:
                self.active_profiles[item_id] = profile_id
            driver_path, remote = extract_driver_info(data_s)
            status = "STARTED" if driver_path and remote else "STARTED (no debug)"
            self._set_status(item_id, status, profile_id=profile_id)
            self._log(f"[{acc['uid']}] START OK")

            if not driver_path or not remote:
                self._record_failed(item_id, acc, "STARTED (no debug)")
                return

            self._set_status(item_id, "LOGIN...")
            self._log(f"[{acc['uid']}] LOGIN START")
            ok_login, err_login = login_scoopz(
                driver_path,
                remote,
                acc["uid"],
                acc["pass"],
                "",
                max_retries=3,
                keep_browser=True,
            )
            if not ok_login:
                status = self._format_login_error(err_login)
                self._set_status(item_id, status)
                self._log(f"[{acc['uid']}] {status}")
                self._record_failed(item_id, acc, status)
                return

            if self.stop_event.is_set():
                return

            self._set_status(item_id, "JOIN CIRCLES...")
            self._log(f"[{acc['uid']}] JOIN START")

            try:
                from selenium import webdriver
                from selenium.webdriver.chrome.service import Service as ChromeService
                from selenium.webdriver.common.by import By
                from selenium.webdriver.support.ui import WebDriverWait
            except Exception as e:
                self._set_status(item_id, f"JOIN ERR: {e}")
                self._record_failed(item_id, acc, f"JOIN ERR: {e}")
                return

            options = webdriver.ChromeOptions()
            options.add_experimental_option("debuggerAddress", remote.strip())
            driver = webdriver.Chrome(service=ChromeService(driver_path), options=options)
            wait = WebDriverWait(driver, 15)
            base_url = "https://thescoopz.com"
            driver.get(f"{base_url}/circles")

            try:
                wait.until(lambda d: d.find_elements(By.CSS_SELECTOR, "a.card[href^='/c/']"))
            except Exception:
                self._set_status(item_id, "JOIN ERR: no circles found")
                self._log(f"[{acc['uid']}] JOIN ERR: no circles found on /circles page")
                self._record_failed(item_id, acc, "JOIN ERR: no circles found")
                driver.quit()
                return

            # Scroll to load more circles before collecting
            max_scroll = 6
            for scroll_attempt in range(max_scroll):
                if self.stop_event.is_set():
                    break
                try:
                    driver.execute_script("window.scrollBy(0, document.body.scrollHeight);")
                    time.sleep(0.6)
                except Exception as e:
                    self._log(f"[{acc['uid']}] JOIN: scroll error - {e}")
                    break

            # Collect circle links
            try:
                cards = driver.find_elements(By.CSS_SELECTOR, "a.card[href^='/c/']")
                if not cards:
                    self._set_status(item_id, "JOIN ERR: no circles to join")
                    self._log(f"[{acc['uid']}] JOIN ERR: cards empty after scroll")
                    self._record_failed(item_id, acc, "JOIN ERR: no circles to join")
                    driver.quit()
                    return
                    
                hrefs = []
                for el in cards:
                    try:
                        href = (el.get_attribute("href") or "").strip()
                        if not href:
                            continue
                        if href.startswith("/"):
                            href = f"{base_url}{href}"
                        if href not in hrefs:
                            hrefs.append(href)
                    except Exception as e:
                        self._log(f"[{acc['uid']}] JOIN: error extracting href - {e}")
                        continue
                
                if not hrefs:
                    self._set_status(item_id, "JOIN ERR: no valid hrefs")
                    self._log(f"[{acc['uid']}] JOIN ERR: extracted hrefs list is empty")
                    self._record_failed(item_id, acc, "JOIN ERR: no valid hrefs")
                    driver.quit()
                    return
                    
                # Randomize the order
                try:
                    random.shuffle(hrefs)
                    self._log(f"[{acc['uid']}] Found {len(hrefs)} circles, shuffled & starting join...")
                except Exception as e:
                    self._log(f"[{acc['uid']}] JOIN: shuffle error - {e}, using original order")
                    # Continue anyway with original order if shuffle fails
                    
            except Exception as e:
                self._set_status(item_id, f"JOIN ERR: collect circles failed")
                self._log(f"[{acc['uid']}] JOIN ERR: failed to collect circles - {e}")
                self._record_failed(item_id, acc, f"JOIN ERR: {e}")
                driver.quit()
                return

            joined = 0
            for idx, href in enumerate(hrefs, 1):
                if self.stop_event.is_set() or joined >= join_max:
                    break
                try:
                    driver.get(href)
                    time.sleep(0.4)
                    already = driver.find_elements(By.XPATH, "//button[normalize-space()='Joined' or normalize-space()='Leave']")
                    if already:
                        self._log(f"[{acc['uid']}] Circle {idx}/{len(hrefs)}: already joined")
                        continue
                    join_btn = None
                    btns = driver.find_elements(By.XPATH, "//button[normalize-space()='Join' or .//*[normalize-space()='Join']]")
                    if btns:
                        join_btn = btns[0]
                    if not join_btn:
                        self._log(f"[{acc['uid']}] Circle {idx}/{len(hrefs)}: join button not found")
                        continue
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", join_btn)
                    time.sleep(0.2)
                    driver.execute_script("arguments[0].click();", join_btn)
                    joined += 1
                    self._log(f"[{acc['uid']}] JOINED {joined}/{join_max}: circle #{idx} - {href}")
                    time.sleep(0.5)
                except Exception as e:
                    self._log(f"[{acc['uid']}] Circle {idx}/{len(hrefs)}: error - {e}")
                    continue

            self._set_status(item_id, f"JOIN OK ({joined})")
            self._log(f"[{acc['uid']}] JOIN OK: {joined}")
        finally:
            try:
                if profile_id:
                    close_profile(profile_id, 3)
                    delete_profile(profile_id, 10)
            except Exception:
                pass
            try:
                if profile_id:
                    self.created_profiles.discard(profile_id)
            except Exception:
                pass
            if profile_id:
                self._delete_profile_path(profile_id)
                self._track_profile_cleanup()
            try:
                with self.active_lock:
                    self.active_profiles.pop(item_id, None)
            except Exception:
                pass
    def _toggle_advanced(self) -> None:
        if self._advanced_var.get():
            self._advanced_frame.grid()
        else:
            self._advanced_frame.grid_remove()

    def _set_busy(self, busy: bool) -> None:
        if busy and not self._busy:
            self._busy = True
            try:
                self._busy_bar.start(10)
            except Exception:
                pass
        elif not busy and self._busy:
            self._busy = False
            try:
                self._busy_bar.stop()
            except Exception:
                pass

    def _apply_row_striping(self, tree: ttk.Treeview) -> None:
        try:
            tree.tag_configure("row_odd", background="#F8FAFC")
            tree.tag_configure("row_even", background="#FFFFFF")
            items = tree.get_children("")
            for i, item in enumerate(items):
                tags = list(tree.item(item, "tags") or [])
                tags = [t for t in tags if t not in ("row_odd", "row_even")]
                tags.append("row_even" if i % 2 == 0 else "row_odd")
                tree.item(item, tags=tags)
        except Exception:
            pass

    def _load_rows(self) -> None:
        self.accounts = self._dedupe_accounts(self.accounts)
        self._save_accounts_cache()
        self.tree.delete(*self.tree.get_children())
        for idx, row in enumerate(self.accounts, start=1):
            posts = row.get("posts", "")
            followers = row.get("followers", "")
            profile_url = row.get("profile_url", "")
            self.tree.insert(
                "",
                "end",
                iid=str(idx),
                values=(
                    "v",
                    idx,
                    row.get("uid", ""),
                    row.get("pass", ""),
                    row.get("status", "READY"),
                    "" if posts is None else str(posts),
                    "" if followers is None else str(followers),
                    row.get("proxy", ""),
                    row.get("youtube", ""),
                    profile_url,
                    row.get("profile_id", ""),
                ),
            )
            try:
                self._apply_status_tag(str(idx), row.get("status", "READY"))
            except Exception:
                pass
        self._update_counts()
        self._apply_row_striping(self.tree)
        self._load_all_rows(only_errors=self._all_filter_active)

    def _load_profile_rows(self) -> None:
        self.profile_accounts = self._dedupe_accounts(self.profile_accounts)
        self._save_profile_accounts_cache()
        self.profile_tree.delete(*self.profile_tree.get_children())
        for idx, row in enumerate(self.profile_accounts, start=1):
            self.profile_tree.insert(
                "",
                "end",
                iid=str(idx),
                values=(
                    "v",
                    idx,
                    row.get("uid", ""),
                    row.get("pass", ""),
                    row.get("proxy", ""),
                    row.get("youtube", ""),
                    row.get("status", "READY"),
                ),
            )
            try:
                self._apply_profile_status_tag(str(idx), row.get("status", "READY"))
            except Exception:
                pass
        self._update_counts()
        self._apply_row_striping(self.profile_tree)

    def _load_fb_rows(self) -> None:
        self.fb_accounts = self._dedupe_accounts(self.fb_accounts)
        self._save_fb_accounts_cache()
        self.fb_tree.delete(*self.fb_tree.get_children())
        for idx, row in enumerate(self.fb_accounts, start=1):
            self.fb_tree.insert(
                "",
                "end",
                iid=str(idx),
                values=(
                    "v",
                    idx,
                    row.get("uid", ""),
                    row.get("pass", ""),
                    row.get("status", "READY"),
                    "" if row.get("posts") is None else str(row.get("posts")),
                    "" if row.get("followers") is None else str(row.get("followers")),
                    row.get("proxy", ""),
                    row.get("facebook", ""),
                    row.get("profile_url", ""),
                    row.get("profile_id", ""),
                ),
            )
            try:
                self._apply_fb_status_tag(str(idx), row.get("status", "READY"))
            except Exception:
                pass
        self._update_counts()
        self._apply_row_striping(self.fb_tree)
        self._load_all_rows(only_errors=self._all_filter_active)

    def _load_all_rows(self, only_errors: bool = False) -> None:
        if not hasattr(self, "all_tree"):
            return
        cached_chk = {}
        try:
            for iid in self.all_tree.get_children():
                email = (self.all_tree.set(iid, "email") or "").strip()
                social = (self.all_tree.set(iid, "social") or "").strip().upper()
                if email and social:
                    cached_chk[(social, email)] = self.all_tree.set(iid, "chk")
            self.all_tree.delete(*self.all_tree.get_children())
        except Exception:
            return
        rows = []
        for acc in self.accounts or []:
            rows.append(("YTB", acc))
        for acc in (getattr(self, "fb_accounts", None) or []):
            rows.append(("FB", acc))

        def _followers_to_num(val) -> int:
            txt = str(val or "").strip().upper().replace(",", "")
            if not txt:
                return -1
            m = re.match(r"^([0-9]+(?:\.[0-9]+)?)([KMB])?$", txt)
            if m:
                num = float(m.group(1))
                suffix = m.group(2)
                if suffix == "K":
                    num *= 1_000
                elif suffix == "M":
                    num *= 1_000_000
                elif suffix == "B":
                    num *= 1_000_000_000
                return int(num)
            digits = re.sub(r"[^0-9]", "", txt)
            return int(digits) if digits else -1

        # Overview fixed priority: highest followers first
        rows.sort(key=lambda pair: _followers_to_num((pair[1] or {}).get("followers")), reverse=True)

        for idx, (social, row) in enumerate(rows, start=1):
            chk = cached_chk.get((social, row.get("uid", "")), "v")
            posts = row.get("posts", "")
            followers = row.get("followers", "")
            profile_url = row.get("profile_url", "")
            link_val = row.get("youtube", "") if social == "YTB" else row.get("facebook", "")
            status_val = row.get("status", "READY")
            if only_errors:
                status_upper = (status_val or "").upper()
                if not any(
                    key in status_upper
                    for key in ["NO VIDEO", "NO VIDEO", "START ERR", "PROXY", "CREATE ERR"]
                ):
                    continue
            self.all_tree.insert(
                "",
                "end",
                iid=str(idx),
                values=(
                    chk,
                    idx,
                    social,
                    row.get("uid", ""),
                    row.get("pass", ""),
                    status_val,
                    "" if posts is None else str(posts),
                    "" if followers is None else str(followers),
                    row.get("proxy", ""),
                    link_val,
                    profile_url,
                    row.get("profile_id", ""),
                ),
            )
            try:
                self._apply_all_status_tag(str(idx), status_val)
            except Exception:
                pass

        self._apply_row_striping(self.all_tree)

    def _load_fb_profile_rows(self) -> None:
        self.fb_profile_accounts = self._dedupe_accounts(self.fb_profile_accounts)
        self._save_fb_profile_accounts_cache()
        self.fb_profile_tree.delete(*self.fb_profile_tree.get_children())
        for idx, row in enumerate(self.fb_profile_accounts, start=1):
            self.fb_profile_tree.insert(
                "",
                "end",
                iid=str(idx),
                values=(
                    "v",
                    idx,
                    row.get("uid", ""),
                    row.get("pass", ""),
                    row.get("proxy", ""),
                    row.get("facebook", ""),
                    row.get("status", "READY"),
                ),
            )
        self._update_counts()
        self._apply_row_striping(self.fb_profile_tree)

    def _update_counts(self) -> None:
        try:
            ytb_total = self._unique_count(self.accounts)
            fb_total = self._unique_count(self.fb_accounts)
            total_all = ytb_total + fb_total
            self._count_var.set(self._format_total_with_run("Total", total_all, "upload", ytb_total, fb_total))
        except Exception:
            pass
        try:
            self._profile_count_var.set(
                self._format_total_with_run("YTB Profile", self._unique_count(self.profile_accounts), "profile")
            )
        except Exception:
            pass
        try:
            self._fb_profile_count_var.set(
                self._format_total_with_run(
                    "FB Profile", self._unique_count(self.fb_profile_accounts), "fb_profile"
                )
            )
        except Exception:
            pass
        self._update_perf_label()

    def _unique_count(self, accounts: list) -> int:
        try:
            return len({(acc.get("uid") or "").strip() for acc in accounts if acc.get("uid")})
        except Exception:
            return len(accounts)

    def _dedupe_accounts(self, accounts: list) -> list:
        seen = set()
        out = []
        for acc in accounts or []:
            uid = (acc.get("uid") or "").strip()
            if not uid:
                continue
            uid_key = uid.lower()
            if uid_key in seen:
                continue
            seen.add(uid_key)
            out.append(acc)
        return out

    def _transfer_accounts_to_upload(
        self,
        source_accounts: list,
        target_accounts: list,
        checked_emails: set,
        link_field: str,
        reload_rows: bool = True,
    ) -> int:
        if not isinstance(source_accounts, list) or not source_accounts or not checked_emails:
            return 0

        changed = 0
        existing_by_uid = {
            (acc.get("uid") or "").strip().lower(): acc
            for acc in target_accounts
            if (acc.get("uid") or "").strip()
        }

        checked_keys = {str(email or "").strip().lower() for email in checked_emails if str(email or "").strip()}
        for src in source_accounts:
            uid = (src.get("uid") or "").strip()
            if not uid:
                continue
            uid_key = uid.lower()
            if uid_key not in checked_keys:
                continue
            dst = existing_by_uid.get(uid_key)
            if dst is None:
                merged = {
                    "uid": uid,
                    "pass": src.get("pass", ""),
                    "proxy": src.get("proxy", ""),
                    link_field: src.get(link_field, ""),
                    "status": "READY",
                }
                target_accounts.append(merged)
                existing_by_uid[uid_key] = merged
                changed += 1
                continue

            updated = False
            for field in ("pass", "proxy", link_field):
                src_val = (src.get(field) or "").strip()
                dst_val = (dst.get(field) or "").strip()
                if src_val and src_val != dst_val:
                    dst[field] = src_val
                    updated = True
            if updated:
                changed += 1

        if changed:
            target_accounts[:] = self._dedupe_accounts(target_accounts)
            if reload_rows and target_accounts is self.accounts:
                self._load_rows()
            elif reload_rows and target_accounts is self.fb_accounts:
                self._load_fb_rows()
            elif target_accounts is self.accounts:
                self._save_accounts_cache()
            elif target_accounts is self.fb_accounts:
                self._save_fb_accounts_cache()
        return changed

    def _format_total_with_run(
        self,
        label: str,
        total: int,
        kind: str,
        ytb_total: int | None = None,
        fb_total: int | None = None,
    ) -> str:
        if ytb_total is not None and fb_total is not None:
            return f"{label}: {total} | YTB: {ytb_total} | FB: {fb_total}"
        if ytb_total is not None:
            return f"{label}: {total} | YTB: {ytb_total}"
        return f"{label}: {total}"

    def _set_run_total(self, kind: str, total: int) -> None:
        try:
            with self._run_counts_lock:
                rc = self._run_counts.get(kind)
                if rc is None:
                    return
                rc["done"] = 0
                rc["total"] = max(0, int(total or 0))
                rc["emails"] = set()
        except Exception:
            pass
        self._update_counts()

    def _reset_run(self, kind: str) -> None:
        try:
            with self._run_counts_lock:
                rc = self._run_counts.get(kind)
                if rc is None:
                    return
                rc["done"] = 0
                rc["total"] = 0
                rc["emails"] = set()
        except Exception:
            pass
        self._update_counts()
    def _set_cycle_label(self) -> None:
        try:
            text = f"Cycle: {self._cycle_count}"
            self._cycle_var.set(text)
        except Exception:
            pass

    def _format_hms(self, total_sec: float) -> str:
        sec = max(0, int(total_sec or 0))
        hh = sec // 3600
        mm = (sec % 3600) // 60
        ss = sec % 60
        return f"{hh:02d}:{mm:02d}:{ss:02d}"

    def _current_runtime_sec(self) -> float:
        total = float(self._runtime_accum_sec or 0.0)
        if self._runtime_running and self._runtime_started_at:
            total += max(0.0, time.time() - float(self._runtime_started_at))
        return total

    def _set_runtime_label(self) -> None:
        try:
            self._runtime_var.set(f"Runtime: {self._format_hms(self._current_runtime_sec())}")
        except Exception:
            pass

    def _collect_run_progress(self) -> tuple[int, int]:
        done = 0
        total = 0
        try:
            with self._run_counts_lock:
                for rc in self._run_counts.values():
                    done += int(rc.get("done", 0) or 0)
                    total += int(rc.get("total", 0) or 0)
        except Exception:
            pass
        return max(0, done), max(0, total)

    def _update_perf_label(self) -> None:
        try:
            done, total = self._collect_run_progress()
            elapsed_min = max(0.0, self._current_runtime_sec() / 60.0)
            speed = (done / elapsed_min) if elapsed_min > 0.0 and done > 0 else 0.0
            remaining = max(0, total - done)
            eta = "-"
            if remaining == 0 and total > 0:
                eta = "00:00:00"
            elif speed > 0:
                eta_sec = int((remaining / speed) * 60.0)
                eta = self._format_hms(eta_sec)
            self._perf_var.set(f"Speed: {speed:.1f} acc/min | ETA: {eta}")
        except Exception:
            pass

    def _runtime_tick(self) -> None:
        self._set_runtime_label()
        self._update_perf_label()
        if not self._runtime_running:
            self._runtime_after_id = None
            return
        try:
            self._runtime_after_id = self.root.after(1000, self._runtime_tick)
        except Exception:
            self._runtime_after_id = None

    def _runtime_start(self) -> None:
        if self._runtime_running:
            return
        self._runtime_running = True
        self._runtime_started_at = time.time()
        if not self._runtime_after_id:
            self._runtime_tick()

    def _runtime_pause(self) -> None:
        if not self._runtime_running:
            return
        if self._runtime_started_at:
            self._runtime_accum_sec += max(0.0, time.time() - float(self._runtime_started_at))
        self._runtime_started_at = None
        self._runtime_running = False
        if self._runtime_after_id:
            try:
                self.root.after_cancel(self._runtime_after_id)
            except Exception:
                pass
            self._runtime_after_id = None
        self._set_runtime_label()
        self._update_perf_label()

    def _runtime_reset(self) -> None:
        self._runtime_pause()
        self._runtime_accum_sec = 0.0
        self._runtime_started_at = None
        self._set_runtime_label()
        self._update_perf_label()

    def _reset_upload_cycle_stats(self) -> None:
        try:
            with self._upload_stats_lock:
                self._upload_outcomes = {}
            self._upload_stats_var.set("Run: processed 0 | ok 0 | no video 0 | err 0")
        except Exception:
            pass

    def _update_upload_stats_label(self) -> None:
        ok = 0
        no_video = 0
        err = 0
        processed = 0
        try:
            with self._upload_stats_lock:
                processed = len(self._upload_outcomes)
                for cat in self._upload_outcomes.values():
                    if cat == "ok":
                        ok += 1
                    elif cat == "no_video":
                        no_video += 1
                    elif cat == "err":
                        err += 1
            self._upload_stats_var.set(
                f"Run: processed {processed} | ok {ok} | no video {no_video} | err {err}"
            )
        except Exception:
            pass

    def _classify_upload_status(self, status: str) -> str:
        text = (status or "").strip().upper()
        if not text:
            return ""
        if "NO VIDEO" in text:
            return "no_video"
        if "UPLOAD OK" in text:
            return "ok"
        if any(k in text for k in ["ERR", "ERROR", "FAIL", "BLOCKED", "LOI"]):
            return "err"
        return ""

    def _record_upload_status(self, social: str, email: str, status: str) -> None:
        social_u = (social or "").strip().upper()
        email_s = (email or "").strip()
        if social_u not in {"YTB", "FB"} or not email_s:
            return
        cat = self._classify_upload_status(status)
        if not cat:
            return
        key = f"{social_u}:{email_s}"
        try:
            with self._upload_stats_lock:
                prev = self._upload_outcomes.get(key)
                if prev == cat:
                    return
                self._upload_outcomes[key] = cat
            self._update_upload_stats_label()
        except Exception:
            pass

    def _stop_next_cycle_countdown(self) -> None:
        if self._repeat_countdown_after_id:
            try:
                self.root.after_cancel(self._repeat_countdown_after_id)
            except Exception:
                pass
            self._repeat_countdown_after_id = None
        try:
            self._next_cycle_var.set("Next cycle in: -")
        except Exception:
            pass

    def _start_next_cycle_countdown(self, total_sec: int) -> None:
        self._stop_next_cycle_countdown()
        end_time = time.time() + max(0, int(total_sec or 0))

        def _tick():
            if self.stop_event.is_set():
                self._stop_next_cycle_countdown()
                return
            remaining = max(0, end_time - time.time())
            if remaining <= 0:
                self._next_cycle_var.set("Next cycle in: 00:00")
                self._repeat_countdown_after_id = None
                return
            mm = int(remaining) // 60
            ss = int(remaining) % 60
            self._next_cycle_var.set(f"Next cycle in: {mm:02d}:{ss:02d}")
            self._repeat_countdown_after_id = self.root.after(1000, _tick)

        _tick()

    def _reset_cycle_count(self) -> None:
        self._cycle_count = 0
        self._set_cycle_label()

    def _increment_cycle(self) -> None:
        self._cycle_count = max(0, int(self._cycle_count or 0)) + 1
        self._set_cycle_label()
        self._reset_upload_cycle_stats()

    def _mark_run_done(self, kind: str, email: str) -> None:
        if not email:
            return
        updated = False
        try:
            with self._run_counts_lock:
                rc = self._run_counts.get(kind)
                if not rc or rc.get("total", 0) <= 0:
                    return
                if email in rc.get("emails", set()):
                    return
                rc["emails"].add(email)
                rc["done"] = min(int(rc.get("total", 0)), int(rc.get("done", 0)) + 1)
                updated = True
        except Exception:
            updated = False
        if updated:
            self._update_counts()

    def _get_active_search_tree(self):
        if self._is_all_tab():
            return self.all_tree, "all"
        if self._is_profile_tab():
            return self.profile_tree, "profile"
        if self._is_fb_tab():
            return self.fb_tree, "fb"
        if self._is_fb_profile_tab():
            return self.fb_profile_tree, "fb_profile"
        return self.tree, "upload"

    def _normalize_search_query(self, q: str) -> str:
        return re.sub(r"\s+", " ", (q or "").strip().lower())

    def _build_search_groups(self, q: str) -> list[list[str]]:
        groups = []
        for part in (q or "").split("|"):
            tokens = [t for t in re.split(r"\s+", part.strip()) if t]
            if tokens:
                groups.append(tokens)
        return groups

    def _row_matches_search(self, tree: ttk.Treeview, iid: str, groups: list[list[str]]) -> bool:
        if not groups:
            return False
        try:
            cols = list(tree["columns"])
        except Exception:
            cols = []
        parts = []
        for c in cols:
            try:
                v = (tree.set(iid, c) or "").strip().lower()
            except Exception:
                v = ""
            if v:
                parts.append(v)
                parts.append(f"{c}:{v}")
        text = " | ".join(parts)
        for tokens in groups:
            if all(tok in text for tok in tokens):
                return True
        return False

    def _focus_search_match(self, tree: ttk.Treeview, matches: list, idx: int, q: str) -> None:
        if not matches:
            return
        idx = max(0, min(idx, len(matches) - 1))
        self._search_cycle_index = idx
        target = matches[idx]
        try:
            tree.selection_remove(tree.selection())
        except Exception:
            pass
        for item_id in matches:
            try:
                tree.selection_add(item_id)
            except Exception:
                pass
        try:
            tree.focus(target)
        except Exception:
            pass
        try:
            tree.see(target)
        except Exception:
            pass
        self._log(f"[SEARCH] Found {len(matches)} match(es) for: {q} | focus {idx+1}/{len(matches)}")

    def _search_email(self, _evt=None) -> None:
        raw_q = (self.entry_search_email.get() or "").strip()
        if not raw_q or raw_q == self._search_placeholder:
            return
        q = self._normalize_search_query(raw_q)
        if not q:
            return

        tree, tab_key = self._get_active_search_tree()

        # Repeated FIND with same query/tab cycles to next match.
        if (
            self._search_last_query == q
            and self._search_last_tab == tab_key
            and self._search_last_matches
        ):
            next_idx = (self._search_cycle_index + 1) % len(self._search_last_matches)
            self._focus_search_match(tree, self._search_last_matches, next_idx, q)
            return

        groups = self._build_search_groups(q)
        matches = []
        try:
            for iid in tree.get_children():
                if self._row_matches_search(tree, iid, groups):
                    matches.append(iid)
        except Exception:
            matches = []

        self._search_last_query = q
        self._search_last_tab = tab_key
        self._search_last_matches = matches
        self._search_cycle_index = -1

        if not matches:
            self._log(f"[SEARCH] No match: {q}")
            return

        self._focus_search_match(tree, matches, 0, q)

    def _search_email_prev(self, _evt=None) -> None:
        raw_q = (self.entry_search_email.get() or "").strip()
        if not raw_q or raw_q == self._search_placeholder:
            return
        q = self._normalize_search_query(raw_q)
        if not q:
            return
        tree, tab_key = self._get_active_search_tree()
        if not (
            self._search_last_query == q
            and self._search_last_tab == tab_key
            and self._search_last_matches
        ):
            self._search_email()
            return
        prev_idx = (self._search_cycle_index - 1) % len(self._search_last_matches)
        self._focus_search_match(tree, self._search_last_matches, prev_idx, q)

    def _bind_item_email(self, item_id: str, email: str) -> None:
        if not item_id or not email:
            return
        with self._job_item_email_lock:
            self._job_item_email_map[item_id] = email

    def _lookup_item_email(self, item_id: str) -> str:
        with self._job_item_email_lock:
            return self._job_item_email_map.get(item_id, "")

    def _resolve_upload_item_id(self, item_id: str) -> str:
        email = self._lookup_item_email(item_id)
        if not email:
            try:
                email = self.tree.set(item_id, "email")
            except Exception:
                email = ""
        if not email:
            return item_id
        try:
            for iid in self.tree.get_children():
                if self.tree.set(iid, "email") == email:
                    return iid
        except Exception:
            pass
        return item_id

    def _get_acc_by_email(self, email: str) -> dict | None:
        if not email:
            return None
        for acc in self.accounts:
            if acc.get("uid") == email:
                return acc
        return None

    def _collect_transient_failures(self) -> list:
        items = []
        try:
            for iid in self.tree.get_children():
                status = (self.tree.set(iid, "status") or "").strip()
                if status not in self._transient_statuses:
                    continue
                email = self.tree.set(iid, "email")
                acc = self._get_acc_by_email(email)
                if not acc:
                    continue
                self._set_status(iid, "INCOMPLETE")
                items.append((iid, acc))
        except Exception:
            pass
        return items

    def _search_focus_in(self, _evt=None) -> None:
        try:
            if self.entry_search_email.get() == self._search_placeholder:
                self.entry_search_email.delete(0, tk.END)
                self.entry_search_email.configure(foreground="black")
        except Exception:
            pass

    def _search_focus_out(self, _evt=None) -> None:
        try:
            if not self.entry_search_email.get().strip():
                self.entry_search_email.delete(0, tk.END)
                self.entry_search_email.insert(0, self._search_placeholder)
                self.entry_search_email.configure(foreground="gray")
        except Exception:
            pass

    def _set_status(self, item_id: str, status: str, profile_id: str = "") -> None:
        def _update():
            resolved_id = self._resolve_upload_item_id(item_id)
            if profile_id:
                self.tree.set(resolved_id, "profile_id", profile_id)
            self.tree.set(resolved_id, "status", status)
            self._apply_status_tag(resolved_id, status)
            self._flash_tree_row(self.tree, resolved_id)
            if self._status_needs_pulse(status):
                self._pulse_tree_row(self.tree, resolved_id)
            else:
                self._stop_pulse_tree_row(self.tree, resolved_id)
            self._auto_scroll_if_needed(self.tree, resolved_id, status)
            try:
                email = (self.tree.set(resolved_id, "email") or "").strip()
                if email:
                    self._update_all_row("YTB", email, status=status, profile_id=profile_id)
            except Exception:
                pass
            try:
                email = self.tree.set(resolved_id, "email")
                if email:
                    self._record_upload_status("YTB", email, status)
                    for acc in self.accounts:
                        if acc.get("uid") == email:
                            acc["status"] = status
                            if profile_id:
                                acc["profile_id"] = profile_id
                            break
                self._save_accounts_cache()
            except Exception:
                pass

        self.root.after(0, _update)

    def _set_profile_status(self, item_id: str, status: str) -> None:
        def _update():
            self.profile_tree.set(item_id, "status", status)
            self._apply_profile_status_tag(item_id, status)
            self._flash_tree_row(self.profile_tree, item_id)
            self._auto_scroll_if_needed(self.profile_tree, item_id, status)

        self.root.after(0, _update)

    def _set_fb_status(self, item_id: str, status: str) -> None:
        def _update():
            self.fb_tree.set(item_id, "status", status)
            self._apply_fb_status_tag(item_id, status)
            self._flash_tree_row(self.fb_tree, item_id)
            self._auto_scroll_if_needed(self.fb_tree, item_id, status)
            try:
                email = (self.fb_tree.set(item_id, "email") or "").strip()
                if email:
                    self._record_upload_status("FB", email, status)
                    self._update_all_row("FB", email, status=status)
                    for acc in self.fb_accounts:
                        if (acc.get("uid") or "").strip() == email:
                            acc["status"] = status
                            break
                    self._save_fb_accounts_cache()
            except Exception:
                pass

        self.root.after(0, _update)

    def _set_fb_profile_status(self, item_id: str, status: str) -> None:
        def _update():
            self.fb_profile_tree.set(item_id, "status", status)
            self._apply_fb_profile_status_tag(item_id, status)
            self._flash_tree_row(self.fb_profile_tree, item_id)
            self._auto_scroll_if_needed(self.fb_profile_tree, item_id, status)

        self.root.after(0, _update)

    def _mark_user_scroll(self, tree: ttk.Treeview) -> None:
        try:
            self._auto_scroll_block_until[tree] = time.time() + 5.0
            # reset catch-up timer
            after_id = self._auto_scroll_catchup_after_id.get(tree)
            if after_id:
                try:
                    self.root.after_cancel(after_id)
                except Exception:
                    pass
            def _catch_up():
                # only auto-scroll if user is idle
                if self._auto_scroll_block_until.get(tree, 0) > time.time():
                    return
                item_id = self._last_active_item.get(tree)
                if item_id:
                    try:
                        tree.see(item_id)
                    except Exception:
                        pass
            self._auto_scroll_catchup_after_id[tree] = self.root.after(5000, _catch_up)
        except Exception:
            pass

    def _auto_scroll_if_needed(self, tree: ttk.Treeview, item_id: str, status: str) -> None:
        return

    def _record_failed(self, item_id: str, acc: dict, reason: str) -> None:
        reason_upper = (reason or "").upper()
        is_blocked = "BLOCKED" in reason_upper
        if is_blocked:
            return
        with self.failed_accounts_lock:
            for iid, _acc in self.failed_accounts:
                if iid == item_id:
                    return
            self.failed_accounts.append((item_id, acc))
        try:
            log_path = os.path.join(DATA_DIR, "logs", "failed_accounts.log")
            os.makedirs(os.path.dirname(log_path), exist_ok=True)
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"{acc.get('uid','')} | {reason}\n")
        except Exception:
            pass
        try:
            self._move_account_to_bottom(acc.get("uid", ""))
        except Exception:
            pass

    def _record_profile_failed(self, item_id: str, acc: dict, reason: str) -> None:
        reason_upper = (reason or "").upper()
        is_blocked = "BLOCKED" in reason_upper
        if is_blocked:
            return
        with self.profile_failed_lock:
            for iid, _acc in self.profile_failed_accounts:
                if iid == item_id:
                    return
            self.profile_failed_accounts.append((item_id, acc))
        try:
            log_path = os.path.join(DATA_DIR, "logs", "failed_profile_accounts.log")
            os.makedirs(os.path.dirname(log_path), exist_ok=True)
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"{acc.get('uid','')} | {reason}\n")
        except Exception:
            pass

    def _clear_failed_log(self) -> None:
        try:
            log_path = os.path.join(DATA_DIR, "logs", "failed_accounts.log")
            if os.path.exists(log_path):
                os.remove(log_path)
        except Exception:
            pass

    def _clear_profile_failed_log(self) -> None:
        try:
            log_path = os.path.join(DATA_DIR, "logs", "failed_profile_accounts.log")
            if os.path.exists(log_path):
                os.remove(log_path)
        except Exception:
            pass

    def _move_account_to_bottom(self, uid: str) -> None:
        if not uid:
            return
        try:
            idx = next((i for i, a in enumerate(self.accounts) if a.get("uid") == uid), None)
        except Exception:
            idx = None
        if idx is None:
            return
        acc = self.accounts.pop(idx)
        self.accounts.append(acc)
        self._rebuild_tree_from_accounts()

    def _rebuild_tree_from_accounts(self) -> None:
        state = {}
        for iid in self.tree.get_children():
            email = self.tree.set(iid, "email")
            state[email] = {
                "chk": self.tree.set(iid, "chk"),
                "status": self.tree.set(iid, "status"),
                "posts": self.tree.set(iid, "posts"),
                "followers": self.tree.set(iid, "followers"),
                "youtube": self.tree.set(iid, "youtube"),
                "profile_url": self.tree.set(iid, "profile_url"),
                "profile_id": self.tree.set(iid, "profile_id"),
                "tags": self.tree.item(iid, "tags"),
            }
        self.tree.delete(*self.tree.get_children())
        seen = set()
        out_idx = 1
        for row in self.accounts:
            email = row.get("uid", "")
            if email in seen:
                continue
            seen.add(email)
            cached = state.get(email, {})
            posts = cached.get("posts", row.get("posts", ""))
            followers = cached.get("followers", row.get("followers", ""))
            profile_url = cached.get("profile_url", row.get("profile_url", ""))
            status = row.get("status") or cached.get("status", "READY")
            chk = cached.get("chk", "v")
            tags = cached.get("tags", ())
            self.tree.insert(
                "",
                "end",
                iid=str(out_idx),
                values=(
                    chk,
                    out_idx,
                    email,
                    row.get("pass", ""),
                    status,
                    "" if posts is None else str(posts),
                    "" if followers is None else str(followers),
                    row.get("proxy", ""),
                    cached.get("youtube", row.get("youtube", "")),
                    profile_url,
                    cached.get("profile_id", row.get("profile_id", "")),
                ),
                tags=tags,
            )
            out_idx += 1

    def _sort_followers_desc(self) -> None:
        return

    def _toggle_followers_sort_all(self) -> None:
        if self.executor is not None or self._busy:
            return
        if self._is_all_tab():
            self._sort_state["followers_all"] = "desc"
            self._sort_tree_by_column(self.all_tree, "followers", descending=True)
            return
        state = self._sort_state.get("followers_all")
        if state == "desc":
            self._sort_state["followers_all"] = "asc"
            if self._is_fb_tab():
                self._sort_tree_by_column(self.fb_tree, "followers", descending=False)
            else:
                self._sort_tree_by_column(self.tree, "followers", descending=False)
            return
        self._sort_state["followers_all"] = "desc"
        if self._is_all_tab():
            self._sort_tree_by_column(self.all_tree, "followers", descending=True)
        elif self._is_fb_tab():
            self._sort_tree_by_column(self.fb_tree, "followers", descending=True)
        else:
            self._sort_tree_by_column(self.tree, "followers", descending=True)

    def _sort_accounts_by_followers(self, descending: bool = True) -> None:
        def _to_num(val) -> int:
            if val is None or val == "":
                return -1
            text = str(val).strip()
            if not text:
                return -1
            digits = re.sub(r"[^0-9]", "", text)
            if not digits:
                return -1
            try:
                return int(digits)
            except Exception:
                return -1

        try:
            email_to_followers = {}
            for iid in self.tree.get_children():
                email = (self.tree.set(iid, "email") or "").strip()
                if not email:
                    continue
                email_to_followers[email] = _to_num(self.tree.set(iid, "followers"))

            def _followers_for(acc: dict) -> int:
                email = (acc.get("uid") or "").strip()
                if email in email_to_followers:
                    return email_to_followers[email]
                return _to_num(acc.get("followers"))

            self.accounts.sort(key=_followers_for, reverse=descending)
            self._rebuild_tree_from_accounts()
            self._save_accounts_cache()
        except Exception:
            pass

    def _enqueue_upload_turn(self) -> int:
        # Upload queue disabled: allow uploads to proceed immediately (parallel).
        return 0

    def _wait_upload_turn(self, token: int) -> bool:
        # Upload queue disabled: always allow unless stopping.
        return not self.stop_event.is_set()

    def _release_upload_turn(self, token: int) -> None:
        # Upload queue disabled: no-op.
        return

    def _prepare_upload_with_retry(
        self,
        driver_path: str,
        remote: str,
        video_path: str,
        caption: str,
        acc_email: str,
        caption_limit: int = 1000,
    ) -> tuple[bool, object, str, str, str]:
        ok_p = False
        drv = None
        up_status = ""
        up_msg = ""
        current_caption = caption
        retry_reopen = {"caption_error", "dialog_error", "timeout", "unexpected_error", "error"}

        try:
            self.operation_delayer.delay_before_upload(acc_email, self._log)
        except Exception:
            pass

        with self.upload_retry_semaphore:
            token = self._enqueue_upload_turn()
            if not self._wait_upload_turn(token):
                return False, None, "stopped", "stopped", current_caption
            try:
                for attempt in range(3):
                    if self.stop_event.is_set():
                        break
                    try:
                        ok_p, drv, up_status, up_msg = upload_prepare(
                            driver_path,
                            remote,
                            video_path,
                            current_caption,
                            lambda: self.stop_event.is_set(),
                            self._log,
                            acc_email,
                            max_total_s=360,
                            file_dialog_semaphore=None,
                        )
                    except Exception as e:
                        up_msg = f"Lock timeout: {e}"
                        self._log(f"[{acc_email}] Upload lock error: {e}")

                    if ok_p:
                        break

                    if up_status in retry_reopen and attempt < 2:
                        if up_status == "caption_error":
                            current_caption = self._next_caption_after_error(current_caption, caption_limit)
                            self._log(f"[{acc_email}] Caption error -> switched fallback caption for retry")
                        wait_time = 2 + attempt
                        self._log(f"[{acc_email}] Upload page retry {attempt+1}/2 in {wait_time}s (status={up_status})")
                        time.sleep(wait_time)
                        continue

                    if up_status in ("select_not_found", "select_click_error") and attempt < 2:
                        wait_time = min(2 ** attempt, 10)
                        self._log(f"[{acc_email}] Upload retry in {wait_time}s...")
                        time.sleep(wait_time)
                        continue

                    break
            finally:
                self._release_upload_turn(token)

        return ok_p, drv, up_status, up_msg, current_caption

    def _post_uploaded_video(self, drv, acc_email: str) -> tuple[str, str, str, str, str]:
        return upload_post_async(
            drv,
            self._log,
            acc_email=acc_email,
            max_total_s=420,
            post_button_semaphore=self.post_button_semaphore,
        )

    def _schedule_follow_sort(self) -> None:
        if self.executor is not None or self._busy:
            return
        if self._follow_sort_after_id:
            try:
                self.root.after_cancel(self._follow_sort_after_id)
            except Exception:
                pass
            self._follow_sort_after_id = None

        def _run():
            self._follow_sort_after_id = None
            self._apply_follow_sort()

        self._follow_sort_after_id = self.root.after(300, _run)

    def _apply_follow_sort(self) -> None:
        if self.executor is not None or self._busy:
            return
        def _to_num(val) -> int:
            if val is None or val == "":
                return -1
            text = str(val).strip()
            if not text:
                return -1
            digits = re.sub(r"[^0-9]", "", text)
            if not digits:
                return -1
            try:
                return int(digits)
            except Exception:
                return -1

        try:
            self.accounts.sort(
                key=lambda acc: _to_num(acc.get("followers")),
                reverse=True,
            )
            self._rebuild_tree_from_accounts()
            self._save_accounts_cache()
        except Exception:
            pass

    def _sort_tree_by_column(self, tree: ttk.Treeview, col: str, descending: bool = True) -> None:
        if self.executor is not None or self._busy:
            return
        def _to_num(val) -> int:
            if val is None or val == "":
                return -1
            text = str(val).strip()
            if not text:
                return -1
            digits = re.sub(r"[^0-9]", "", text)
            if not digits:
                return -1
            try:
                return int(digits)
            except Exception:
                return -1

        try:
            items = list(tree.get_children())
            items.sort(key=lambda iid: _to_num(tree.set(iid, col)), reverse=descending)
            for idx, iid in enumerate(items):
                tree.move(iid, "", idx)
                try:
                    tree.set(iid, "stt", str(idx + 1))
                except Exception:
                    pass
        except Exception:
            pass

    def _reset_upload_tree_order(self) -> None:
        # Rebuild upload tree in original accounts order
        self._sort_state["posts"] = None
        self._sort_state["followers"] = None
        self._rebuild_tree_from_accounts()

    def _toggle_upload_sort(self, col: str) -> None:
        if self.executor is not None or self._busy:
            return
        # Cycle: desc -> asc -> reset
        state = self._sort_state.get(col)
        if state is None:
            self._sort_state[col] = "desc"
            self._sort_tree_by_column(self.tree, col, descending=True)
            return
        if state == "desc":
            self._sort_state[col] = "asc"
            self._sort_tree_by_column(self.tree, col, descending=False)
            return
        self._sort_state[col] = None
        self._reset_upload_tree_order()

    def _reset_fb_tree_order(self) -> None:
        try:
            self._reorder_tree_by_accounts(self.fb_tree, self.fb_accounts)
        except Exception:
            pass

    def _toggle_fb_sort(self, col: str) -> None:
        if self.executor is not None or self._busy:
            return
        key = f"fb_{col}"
        state = self._sort_state.get(key)
        if state is None:
            self._sort_state[key] = "desc"
            self._sort_tree_by_column(self.fb_tree, col, descending=True)
            return
        if state == "desc":
            self._sort_state[key] = "asc"
            self._sort_tree_by_column(self.fb_tree, col, descending=False)
            return
        self._sort_state[key] = None
        self._reset_fb_tree_order()

    def _reset_all_tree_order(self) -> None:
        try:
            self._load_all_rows(only_errors=self._all_filter_active)
        except Exception:
            pass

    def _toggle_all_sort(self, col: str) -> None:
        if self.executor is not None or self._busy:
            return
        key = f"all_{col}"
        state = self._sort_state.get(key)
        if state is None:
            self._sort_state[key] = "desc"
            self._sort_tree_by_column(self.all_tree, col, descending=True)
            return
        if state == "desc":
            self._sort_state[key] = "asc"
            self._sort_tree_by_column(self.all_tree, col, descending=False)
            return
        self._sort_state[key] = None
        self._reset_all_tree_order()

    def _get_checked_email_set(self, tree: ttk.Treeview) -> set:
        emails = set()
        try:
            for iid in tree.get_children():
                if tree.set(iid, "chk") != "v":
                    continue
                email = (tree.set(iid, "email") or "").strip()
                if email:
                    emails.add(email)
        except Exception:
            pass
        return emails

    def _format_fb_title_case1(self, title: str) -> str:
        if not title:
            return ""
        # remove extra quotes + trim
        title = str(title).strip().strip('"').strip()
        # split by "|" (allow spaces)
        parts = [p.strip() for p in re.split(r"\s*\|\s*", title) if p.strip()]
        if len(parts) >= 2:
            return parts[1].strip()
        return title

    def _limit_caption(self, text: str, max_len: int = 1000) -> str:
        if not text:
            return ""
        txt = str(text).strip()
        if len(txt) <= max_len:
            return txt
        return txt[:max_len].rstrip()

    def _load_fallback_captions(self) -> None:
        try:
            if not os.path.exists(self._fallback_caption_file):
                with open(self._fallback_caption_file, "w", encoding="utf-8") as f:
                    f.write("")
            with open(self._fallback_caption_file, "r", encoding="utf-8") as f:
                lines = [ln.strip() for ln in f.read().splitlines() if ln.strip()]
            with self._fallback_caption_lock:
                self._fallback_captions = lines
                self._fallback_caption_idx = 0
        except Exception:
            pass

    def _get_fallback_caption(self) -> str:
        with self._fallback_caption_lock:
            if not self._fallback_captions:
                return ""
            cap = self._fallback_captions[self._fallback_caption_idx % len(self._fallback_captions)]
            self._fallback_caption_idx = (self._fallback_caption_idx + 1) % len(self._fallback_captions)
            return cap

    def _ensure_title(self, title: str, max_len: int = 1000) -> str:
        cleaned = (title or "").strip()
        if not cleaned:
            cleaned = self._get_fallback_caption()
        return self._limit_caption(cleaned, max_len)

    def _build_caption(self, title: str, max_len: int = 1000) -> str:
        caption = (title or "").strip()
        if not caption:
            caption = self._get_fallback_caption()
        return self._limit_caption(caption, max_len)

    def _next_caption_after_error(self, current_caption: str, max_len: int = 1000) -> str:
        # On caption_error, rotate to the next fallback caption if available.
        fb = (self._get_fallback_caption() or "").strip()
        if not fb:
            return self._limit_caption(current_caption or "", max_len)
        if fb == (current_caption or "").strip():
            fb2 = (self._get_fallback_caption() or "").strip()
            if fb2:
                fb = fb2
        return self._limit_caption(fb, max_len)

    def _get_checked_all_rows(self) -> list:
        rows = []
        try:
            for iid in self.all_tree.get_children():
                if self.all_tree.set(iid, "chk") != "v":
                    continue
                email = (self.all_tree.set(iid, "email") or "").strip()
                social = (self.all_tree.set(iid, "social") or "").strip().upper()
                if email and social:
                    rows.append((social, email))
        except Exception:
            pass
        return rows

    def _build_mixed_ordered_rows(self, base_rows: list | None = None) -> list:
        rows = []
        try:
            ytb_exists = {str(a.get("uid") or "").strip().lower() for a in self.accounts}
            fb_exists = {str(a.get("uid") or "").strip().lower() for a in self.fb_accounts}

            if base_rows is None:
                checked = self._get_checked_all_rows()
                checked_set = {(str(s or "").strip().upper(), str(e or "").strip()) for s, e in checked}
                if not checked_set:
                    return []
                source_rows = []
                for iid in self.all_tree.get_children():
                    social = (self.all_tree.set(iid, "social") or "").strip().upper()
                    email = (self.all_tree.set(iid, "email") or "").strip()
                    if (social, email) in checked_set:
                        source_rows.append((social, email))
            else:
                source_rows = [
                    (str(s or "").strip().upper(), str(e or "").strip())
                    for s, e in (base_rows or [])
                ]

            seen_rows = set()
            for social, email in source_rows:
                if not social or not email:
                    continue
                email_key = email.lower()
                row_key = (social, email_key)
                if row_key in seen_rows:
                    continue
                if social == "YTB" and email_key in ytb_exists:
                    seen_rows.add(row_key)
                    rows.append((social, email))
                elif social == "FB" and email_key in fb_exists:
                    seen_rows.add(row_key)
                    rows.append((social, email))
        except Exception:
            return []
        return rows

    def _filter_all_errors(self) -> None:
        self._all_filter_active = True
        self._load_all_rows(only_errors=True)

    def _clear_all_filter(self) -> None:
        self._all_filter_active = False
        self._load_all_rows(only_errors=False)

    def _update_all_row(
        self,
        social: str,
        email: str,
        status: str | None = None,
        posts=None,
        followers=None,
        profile_url: str | None = None,
        profile_id: str | None = None,
    ) -> None:
        try:
            social = (social or "").strip().upper()
            email = (email or "").strip()
            if not social or not email:
                return
            for iid in self.all_tree.get_children():
                if (self.all_tree.set(iid, "social") or "").strip().upper() != social:
                    continue
                if (self.all_tree.set(iid, "email") or "").strip() != email:
                    continue
                if status is not None and status != "":
                    self.all_tree.set(iid, "status", status)
                    self._apply_all_status_tag(iid, status)
                    self._auto_scroll_if_needed(self.all_tree, iid, status)
                if posts is not None:
                    self.all_tree.set(iid, "posts", str(posts))
                if followers is not None:
                    self.all_tree.set(iid, "followers", str(followers))
                if profile_url:
                    self.all_tree.set(iid, "profile_url", profile_url)
                if profile_id:
                    self.all_tree.set(iid, "profile_id", profile_id)
                break
        except Exception:
            pass

    def _map_email_to_item_id(self, tree: ttk.Treeview) -> dict:
        mapping = {}
        try:
            for iid in tree.get_children():
                email = (tree.set(iid, "email") or "").strip()
                if email:
                    mapping[email] = iid
        except Exception:
            pass
        return mapping

    def _status_is_done(self, status: str, done_keys: set) -> bool:
        text = (status or "").strip().upper()
        return any(k in text for k in done_keys)

    def _collect_pending_emails(self, tree: ttk.Treeview, done_keys: set) -> set:
        pending = set()
        try:
            for iid in tree.get_children():
                email = (tree.set(iid, "email") or "").strip()
                if not email:
                    continue
                status = tree.set(iid, "status")
                if not self._status_is_done(status, done_keys):
                    pending.add(email)
        except Exception:
            pass
        return pending

    def _prompt_resume(self, kind: str, count: int) -> bool:
        msg = (
            f"There are {count} accounts not OK.\n"
            f"Do you want to continue with the not-OK accounts?"
        )
        return messagebox.askyesno("Tiep tuc", msg)

    def _reorder_tree_by_accounts(self, tree: ttk.Treeview, accounts: list) -> None:
        try:
            email_to_iid = self._map_email_to_item_id(tree)
            idx = 0
            for acc in accounts:
                email = (acc.get("uid") or "").strip()
                iid = email_to_iid.get(email)
                if not iid:
                    continue
                tree.move(iid, "", idx)
                idx += 1
        except Exception:
            pass

    def _apply_status_tag(self, item_id: str, status: str) -> None:
        status_upper = (status or "").upper()
        if any(key in status_upper for key in ["ERR", "ERROR", "FAIL", "BLOCKED", "LOI"]):
            self.tree.item(item_id, tags=("status_err",))
        elif any(key in status_upper for key in ["OK", "SUCCESS", "DONE"]):
            self.tree.item(item_id, tags=("status_ok",))
        elif any(
            key in status_upper
            for key in [
                "LOGIN",
                "DOWNLOAD",
                "POSTING",
                "UPLOAD",
                "START",
                "SCAN",
                "RUNNING",
                "CHECK",
                "SYNC",
            ]
        ):
            self.tree.item(item_id, tags=("status_work",))
        elif any(key in status_upper for key in ["WAIT", "RETRY", "PENDING"]):
            self.tree.item(item_id, tags=("status_warn",))
        else:
            self.tree.item(item_id, tags=())

    def _status_needs_pulse(self, status: str) -> bool:
        status_upper = (status or "").upper()
        if any(key in status_upper for key in ["ERR", "ERROR", "FAIL", "BLOCKED", "LOI"]):
            return False
        if any(key in status_upper for key in ["OK", "SUCCESS", "DONE", "READY"]):
            return False
        return any(
            key in status_upper
            for key in [
                "LOGIN",
                "DOWNLOAD",
                "POSTING",
                "UPLOAD",
                "START",
                "SCAN",
                "RUNNING",
                "CHECK",
                "SYNC",
            ]
        )

    def _stop_pulse_tree_row(self, tree: ttk.Treeview, item_id: str) -> None:
        try:
            token = self._pulse_tokens.get(item_id, 0) + 1
            self._pulse_tokens[item_id] = token
            cur_tags = list(tree.item(item_id, "tags") or [])
            cur_tags = [t for t in cur_tags if t not in ("status_pulse_a", "status_pulse_b")]
            tree.item(item_id, tags=tuple(cur_tags))
        except Exception:
            pass

    def _pulse_tree_row(
        self, tree: ttk.Treeview, item_id: str, cycles: int = 6, interval: int = 220
    ) -> None:
        try:
            base_tags = list(tree.item(item_id, "tags") or [])
        except Exception:
            return
        base_tags = [t for t in base_tags if t not in ("status_pulse_a", "status_pulse_b")]
        token = self._pulse_tokens.get(item_id, 0) + 1
        self._pulse_tokens[item_id] = token

        def _tick(step: int) -> None:
            if self._pulse_tokens.get(item_id) != token:
                return
            try:
                if not tree.exists(item_id):
                    return
            except Exception:
                return
            pulse_tag = "status_pulse_a" if step % 2 == 0 else "status_pulse_b"
            try:
                tree.item(item_id, tags=tuple(base_tags + [pulse_tag]))
            except Exception:
                return
            if step < (cycles * 2 - 1):
                try:
                    self.root.after(interval, lambda: _tick(step + 1))
                except Exception:
                    return
            else:
                try:
                    tree.item(item_id, tags=tuple(base_tags))
                except Exception:
                    pass

        _tick(0)

    def _apply_profile_status_tag(self, item_id: str, status: str) -> None:
        status_upper = (status or "").upper()
        if any(key in status_upper for key in ["ERR", "ERROR", "FAIL", "BLOCKED", "LOI"]):
            self.profile_tree.item(item_id, tags=("status_err",))
        elif any(key in status_upper for key in ["OK", "SUCCESS", "DONE", "UPDATED", "POSTING", "UPLOAD"]):
            self.profile_tree.item(item_id, tags=("status_ok",))
        else:
            self.profile_tree.item(item_id, tags=())

    def _apply_fb_status_tag(self, item_id: str, status: str) -> None:
        status_upper = (status or "").upper()
        if any(key in status_upper for key in ["ERR", "ERROR", "FAIL", "BLOCKED", "LOI"]):
            self.fb_tree.item(item_id, tags=("status_err",))
        elif any(key in status_upper for key in ["OK", "SUCCESS", "DONE", "POSTING", "UPLOAD"]):
            self.fb_tree.item(item_id, tags=("status_ok",))
        else:
            self.fb_tree.item(item_id, tags=())

    def _apply_all_status_tag(self, item_id: str, status: str) -> None:
        status_upper = (status or "").upper()
        if any(key in status_upper for key in ["ERR", "ERROR", "FAIL", "BLOCKED", "LOI"]):
            self.all_tree.item(item_id, tags=("status_err",))
        elif any(key in status_upper for key in ["OK", "SUCCESS", "DONE", "POSTING", "UPLOAD"]):
            self.all_tree.item(item_id, tags=("status_ok",))
        else:
            self.all_tree.item(item_id, tags=())

    def _apply_fb_profile_status_tag(self, item_id: str, status: str) -> None:
        status_upper = (status or "").upper()
        if any(key in status_upper for key in ["ERR", "ERROR", "FAIL", "BLOCKED", "LOI"]):
            self.fb_profile_tree.item(item_id, tags=("status_err",))
        elif any(key in status_upper for key in ["OK", "SUCCESS", "DONE", "UPDATED", "POSTING", "UPLOAD"]):
            self.fb_profile_tree.item(item_id, tags=("status_ok",))
        else:
            self.fb_profile_tree.item(item_id, tags=())

    def _flash_tree_row(self, tree: ttk.Treeview, item_id: str) -> None:
        try:
            tags = list(tree.item(item_id, "tags") or [])
        except Exception:
            return
        if "status_flash" not in tags:
            tags.append("status_flash")
            try:
                tree.item(item_id, tags=tuple(tags))
            except Exception:
                return

        def _clear():
            try:
                cur_tags = list(tree.item(item_id, "tags") or [])
                if "status_flash" in cur_tags:
                    cur_tags = [t for t in cur_tags if t != "status_flash"]
                    tree.item(item_id, tags=tuple(cur_tags))
            except Exception:
                pass

        try:
            self.root.after(350, _clear)
        except Exception:
            pass

    def _next_proxy(self) -> str:
        with self._extra_proxy_lock:
            if not self._extra_proxies:
                return ""
            proxy = self._extra_proxies[self._extra_proxy_idx % len(self._extra_proxies)]
            self._extra_proxy_idx += 1
            return proxy

    def _is_proxy_error(self, msg: str) -> bool:
        text = (msg or "").lower()
        if "proxy" not in text:
            return False
        proxy_keywords = [
            "connect",
            "connection",
            "cannot",
            "can't",
            "failed",
            "timeout",
            "tunnel",
            "unable",
        ]
        return any(k in text for k in proxy_keywords)

    def _set_proxy_cell(self, tree: ttk.Treeview, item_id: str, proxy: str) -> None:
        try:
            if tree == self.tree:
                item_id = self._resolve_upload_item_id(item_id)
            tree.set(item_id, "proxy", proxy)
        except Exception:
            pass

    def _save_cache_by_kind(self, kind: str) -> None:
        if kind == "upload":
            self._save_accounts_cache()
        elif kind == "profile":
            self._save_profile_accounts_cache()
        elif kind == "fb":
            self._save_fb_accounts_cache()
        elif kind == "fb_profile":
            self._save_fb_profile_accounts_cache()

    def _load_extra_proxy_list(self) -> None:
        try:
            if not os.path.exists(self._extra_proxy_file):
                return
            with open(self._extra_proxy_file, "r", encoding="utf-8") as f:
                content = f.read()
        except Exception:
            return
        proxies = []
        for raw in content.splitlines():
            line = (raw or "").strip()
            if not line:
                continue
            for sep in (",", "\t", ";", " "):
                if sep in line:
                    line = line.split(sep, 1)[0].strip()
                    break
            if line:
                proxies.append(line)
        if not proxies:
            return
        with self._extra_proxy_lock:
            self._extra_proxies = proxies
            self._extra_proxy_idx = 0

    def _save_extra_proxy_list(self) -> None:
        try:
            with self._extra_proxy_lock:
                proxies = list(self._extra_proxies)
        except Exception:
            return
        try:
            with open(self._extra_proxy_file, "w", encoding="utf-8") as f:
                f.write("\n".join(proxies))
        except Exception:
            pass

    def _replace_proxy_for_account(self, acc: dict, item_id: str, kind: str, tree: ttk.Treeview) -> bool:
        new_proxy = self._next_proxy()
        if not new_proxy:
            self._log("[PROXY] No extra proxies available")
            return False
        acc["proxy"] = new_proxy
        self._set_proxy_cell(tree, item_id, new_proxy)
        try:
            self._save_cache_by_kind(kind)
        except Exception:
            pass
        self._log(f"[PROXY] Swapped proxy for {acc.get('uid','')} -> {new_proxy}")
        return True

    def _replace_proxy_errors(self, kind: str) -> None:
        if not self._extra_proxies:
            messagebox.showinfo("Proxy", "Danh sach proxy rong. Hay IMPORT PROXY truoc.")
            return
        if kind == "fb":
            tree = self.fb_tree
            accounts = self.fb_accounts
            set_status = self._set_fb_status
        else:
            tree = self.tree
            accounts = self.accounts
            set_status = self._set_status

        replaced = 0
        email_to_iid = self._map_email_to_item_id(tree)
        for acc in accounts:
            email = (acc.get("uid") or "").strip()
            if not email:
                continue
            iid = email_to_iid.get(email)
            if not iid:
                continue
            status = tree.set(iid, "status")
            if not self._is_proxy_error(status):
                continue
            if self._replace_proxy_for_account(acc, iid, kind, tree):
                replaced += 1
                try:
                    set_status(iid, "PROXY REPLACED")
                except Exception:
                    pass
        self._log(f"[PROXY] Replaced {replaced} proxy errors ({kind})")

    def _retry_start_profile_with_new_proxy(
        self,
        acc: dict,
        item_id: str,
        kind: str,
        tree: ttk.Treeview,
        status_setter,
        win_pos: str | None = None,
        win_size: str | None = None,
        created_set: str = "created_profiles",
    ) -> tuple[str | None, dict | None, str]:
        if not self._replace_proxy_for_account(acc, item_id, kind, tree):
            return None, None, ""
        ok_c = False
        data_c = {}
        msg_c = ""
        with self.create_lock:
            for attempt in range(2):
                ok_c, data_c, msg_c = create_profile(acc["uid"], acc["proxy"], SCOOPZ_URL)
                if ok_c:
                    break
                time.sleep(2 + attempt)
        if not ok_c:
            status_setter(f"CREATE ERR: {msg_c}")
            return None, None, msg_c
        profile_id = None
        if isinstance(data_c, dict):
            profile_id = (data_c.get("data") or {}).get("id") or data_c.get("id") or data_c.get("profile_id")
        if not profile_id:
            status_setter("NO PROFILE ID")
            return None, None, "NO PROFILE ID"
        self._remember_profile_path(profile_id, data_c)
        try:
            getattr(self, created_set).add(profile_id)
        except Exception:
            pass
        status_setter("START...")
        if win_pos is None:
            ok_s, data_s, msg_s = start_profile(profile_id)
        else:
            ok_s, data_s, msg_s = start_profile(profile_id, win_pos=win_pos, win_size=win_size)
        if not ok_s:
            status_setter(f"START ERR: {msg_s}")
            return None, None, msg_s
        return profile_id, data_s, ""

    def _clear_status_tags(self) -> None:
        try:
            for iid in self.tree.get_children():
                self.tree.item(iid, tags=())
        except Exception:
            pass

    def _force_close_all_profiles(self) -> None:
        try:
            with self.active_lock:
                ids = list(self.active_profiles.values())
                self.active_profiles.clear()
        except Exception:
            ids = []
        try:
            ids.extend(list(self.created_profiles))
            self.created_profiles.clear()
        except Exception:
            pass
        try:
            ids.extend(list(self.profile_created_profiles))
            self.profile_created_profiles.clear()
        except Exception:
            pass
        ids = list({pid for pid in ids if pid})
        for pid in ids:
            try:
                close_profile(pid, 3)
                delete_profile(pid, 10)
            except Exception:
                pass
            try:
                self._delete_profile_path(pid)
            except Exception:
                pass
        try:
            self._cleanup_gpm_root(force=True)
        except Exception:
            pass

    def _reset_statuses(self, tree: ttk.Treeview, accounts: list, ready_text: str = "READY") -> None:
        try:
            for acc in accounts:
                acc["status"] = ready_text
        except Exception:
            pass
        try:
            for iid in tree.get_children():
                tree.set(iid, "status", ready_text)
                tree.item(iid, tags=())
        except Exception:
            pass

    def _reset_all_statuses(self) -> None:
        self._reset_statuses(self.tree, self.accounts, "READY")
        self._reset_statuses(self.profile_tree, self.profile_accounts, "READY")
        self._reset_statuses(self.fb_tree, self.fb_accounts, "READY")
        self._reset_statuses(self.fb_profile_tree, self.fb_profile_accounts, "READY")
        try:
            self._save_accounts_cache()
            self._save_profile_accounts_cache()
            self._save_fb_accounts_cache()
            self._save_fb_profile_accounts_cache()
        except Exception:
            pass

    def _log(self, msg: str) -> None:
        def _is_noisy(m: str) -> bool:
            text = (m or "").strip()
            upper = text.upper()
            important = (
                "ERR",
                "ERROR",
                "FAILED",
                "TIMEOUT",
                "BLOCK",
                "NO VIDEO",
                "SKIP",
                "RETRY",
                "OK",
                "DONE",
                "SUCCESS",
            )
            noisy_prefixes = (
                "[DL]",
                "[UPLOAD]",
                "[UPLOAD-DIALOG]",
                "[UPLOAD-POST]",
                "[LOGIN]",
                "[PROFILE]",
                "[SELECT]",
                "[SEARCH]",
                "[MENU]",
                "[PROXY]",
            )
            if text.startswith(noisy_prefixes) and not any(k in upper for k in important):
                return True
            return False

        if _is_noisy(msg):
            return

        def _append():
            self.log_box.configure(state="normal")
            self.log_box.insert(tk.END, msg + "\n")
            self.log_box.see(tk.END)
            self.log_box.configure(state="disabled")
        self.root.after(0, _append)

    def _start_download_watchdog(
        self,
        email: str,
        label: str,
        interval_s: int = 30,
        max_seconds: int = 600,
        on_timeout=None,
    ) -> threading.Event:
        stop_evt = threading.Event()
        start = time.time()

        def _watch() -> None:
            while not stop_evt.wait(interval_s):
                elapsed = int(time.time() - start)
                if max_seconds > 0 and elapsed >= max_seconds:
                    self._log(f"[{email}] {label} TIMEOUT after {elapsed}s")
                    if callable(on_timeout):
                        try:
                            on_timeout(elapsed)
                        except Exception:
                            pass
                    break
                self._log(f"[{email}] {label} still running {elapsed}s")

        threading.Thread(target=_watch, daemon=True).start()
        return stop_evt

    def _clear_log_files(self) -> None:
        paths = [
            os.path.join(DATA_DIR, "logs", "app.log"),
            os.path.join(DATA_DIR, "logs", "uploads.log"),
            os.path.join(DATA_DIR, "logs", "downloads.log"),
            os.path.join(DATA_DIR, "logs", "errors.log"),
            os.path.join(DATA_DIR, "logs", "threads.log"),
            os.path.join(DATA_DIR, "logs", "failed_accounts.log"),
        ]
        for path in paths:
            try:
                os.makedirs(os.path.dirname(path), exist_ok=True)
                with open(path, "w", encoding="utf-8"):
                    pass
            except Exception:
                pass

    def _clear_log_view(self) -> None:
        try:
            self.log_box.configure(state="normal")
            self.log_box.delete("1.0", tk.END)
            self.log_box.configure(state="disabled")
        except Exception:
            pass

    def _clear_all_logs(self) -> None:
        self._clear_log_files()
        self._clear_log_view()

    def _format_login_error(self, err: str) -> str:
        text = (err or "").strip().lower()
        if "invalid credentials" in text or "invalid email" in text:
            return "SAI PASS"
        return f"LOGIN ERR: {err}"

    def clear_all_email_videos(self) -> None:
        if self.executor is not None:
            self._log("[CLEAR] Dang chay job, hay STOP truoc.")
            return
        ok = messagebox.askyesno("Confirm", "Delete all videos in the email folder?")
        if not ok:
            return
        base = os.path.join(DATA_DIR, "video")
        if not os.path.isdir(base):
            self._log("[CLEAR] Folder video khong ton tai.")
            return
        exts = {".mp4", ".mov", ".mkv", ".webm", ".avi", ".m4v"}
        deleted = 0
        for root, _dirs, files in os.walk(base):
            for name in files:
                ext = os.path.splitext(name)[1].lower()
                if ext not in exts:
                    continue
                path = os.path.join(root, name)
                try:
                    os.remove(path)
                    deleted += 1
                except Exception as e:
                    self._log(f"[CLEAR] DEL ERR: {path} | {e}")
        self._log(f"[CLEAR] Deleted {deleted} video files.")

    def _set_profile_info(self, item_id: str, profile_url: str, followers, posts=None) -> None:
        def _update():
            resolved_id = self._resolve_upload_item_id(item_id)
            if profile_url:
                self.tree.set(resolved_id, "profile_url", profile_url)
            if followers is not None:
                self.tree.set(resolved_id, "followers", str(followers))
            if posts is not None:
                self.tree.set(resolved_id, "posts", str(posts))
            try:
                email = (self.tree.set(resolved_id, "email") or "").strip()
                if email:
                    self._update_all_row(
                        "YTB",
                        email,
                        profile_url=profile_url,
                        followers=followers,
                        posts=posts,
                    )
            except Exception:
                pass
        self.root.after(0, _update)
        try:
            email = self._lookup_item_email(item_id)
            if not email:
                try:
                    email = self.tree.set(item_id, "email")
                except Exception:
                    email = ""
            if email:
                for acc in self.accounts:
                    if acc.get("uid") == email:
                        if profile_url:
                            acc["profile_url"] = profile_url
                        if followers is not None:
                            acc["followers"] = followers
                        if posts is not None:
                            acc["posts"] = posts
                        break
                self._save_accounts_cache()
                # Keep original order; no auto sort after fetching followers
                try:
                    self.root.after(0, self._refresh_stats)
                except Exception:
                    pass
        except Exception:
            pass

    def _set_fb_profile_info(self, item_id: str, profile_url: str = "", followers=None, posts=None, profile_id: str = "") -> None:
        def _update():
            try:
                if profile_url:
                    self.fb_tree.set(item_id, "profile_url", profile_url)
                if followers is not None:
                    self.fb_tree.set(item_id, "followers", str(followers))
                if posts is not None:
                    self.fb_tree.set(item_id, "posts", str(posts))
                if profile_id:
                    self.fb_tree.set(item_id, "profile_id", profile_id)
                try:
                    email = (self.fb_tree.set(item_id, "email") or "").strip()
                    if email:
                        self._update_all_row(
                            "FB",
                            email,
                            profile_url=profile_url,
                            followers=followers,
                            posts=posts,
                            profile_id=profile_id,
                        )
                except Exception:
                    pass
            except Exception:
                pass
        self.root.after(0, _update)
        try:
            email = ""
            try:
                email = (self.fb_tree.set(item_id, "email") or "").strip()
            except Exception:
                email = ""
            if email:
                for acc in self.fb_accounts:
                    if acc.get("uid") == email:
                        if profile_url:
                            acc["profile_url"] = profile_url
                        if followers is not None:
                            acc["followers"] = followers
                        if posts is not None:
                            acc["posts"] = posts
                        if profile_id:
                            acc["profile_id"] = profile_id
                        break
                self._save_fb_accounts_cache()
                try:
                    self.root.after(0, self._refresh_stats)
                except Exception:
                    pass
        except Exception:
            pass

    def _delete_uploaded_video(self, path: str, email: str) -> None:
        if not path:
            return
        try:
            if os.path.exists(path):
                os.remove(path)
                self._log(f"[{email}] DELETE OK: {os.path.basename(path)}")
        except Exception as e:
            self._log(f"[{email}] DELETE ERR: {e}")

    def _ensure_video_folder(self, email: str) -> None:
        if not email:
            return
        safe = email.strip().lower().replace("@", "_at_").replace(".", "_")
        try:
            base = os.path.join(DATA_DIR, "video", safe)
            os.makedirs(base, exist_ok=True)
        except Exception:
            pass

    def _save_profile_assets(self, email: str, name: str, username: str, avatar_path: str) -> None:
        if not email:
            return
        safe = email.strip().lower().replace("@", "_at_").replace(".", "_")
        out_dir = os.path.join(DATA_DIR, "video", safe)
        try:
            os.makedirs(out_dir, exist_ok=True)
            out_path = os.path.join(out_dir, "profile_assets.json")
            with open(out_path, "w", encoding="utf-8") as f:
                json.dump(
                    {"email": email, "name": name, "username": username, "avatar_path": avatar_path},
                    f,
                    ensure_ascii=False,
                    indent=2,
                )
        except Exception:
            pass

    def _load_profile_assets(self, email: str) -> dict:
        if not email:
            return {}
        safe = email.strip().lower().replace("@", "_at_").replace(".", "_")
        path = os.path.join(DATA_DIR, "video", safe, "profile_assets.json")
        if not os.path.exists(path):
            return {}
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def _log_progress(self, msg: str) -> None:
        def _append():
            self.log_box.configure(state="normal")
            # Replace last line for progress logs
            if msg.startswith("[DL]"):
                try:
                    content = self.log_box.get("1.0", tk.END)
                    lines = content.rstrip("\n").split("\n")
                    if lines and lines[-1].startswith("[DL]"):
                        self.log_box.delete("1.0", tk.END)
                        self.log_box.insert(tk.END, "\n".join(lines[:-1]) + ("\n" if lines[:-1] else ""))
                except Exception:
                    pass
            self.log_box.insert(tk.END, msg + "\n")
            self.log_box.see(tk.END)
            self.log_box.configure(state="disabled")
        with self._log_lock:
            self.root.after(0, _append)

    def _toggle_checked(self, item_id: str) -> None:
        cur = self.tree.set(item_id, "chk")
        self.tree.set(item_id, "chk", "" if cur == "v" else "v")

    def _toggle_checked_all(self, item_id: str) -> None:
        cur = self.all_tree.set(item_id, "chk")
        self.all_tree.set(item_id, "chk", "" if cur == "v" else "v")

    def _set_checked_by_email(self, tree: ttk.Treeview, emails: set) -> None:
        try:
            for iid in tree.get_children():
                email = (tree.set(iid, "email") or "").strip()
                tree.set(iid, "chk", "v" if email in emails else "")
        except Exception:
            pass

    def _select_all_accounts(self) -> None:
        """Select all accounts (mark all as checked)"""
        count = 0
        for item_id in self.tree.get_children():
            self.tree.set(item_id, "chk", "v")
            count += 1
        self._log(f"[SELECT] UPLOAD select all ({count})")

    def _deselect_all_accounts(self) -> None:
        """Deselect all accounts (unmark all)"""
        count = 0
        for item_id in self.tree.get_children():
            self.tree.set(item_id, "chk", "")
            count += 1
        self._log(f"[SELECT] UPLOAD deselect all ({count})")

    def _select_all_all_accounts(self) -> None:
        count = 0
        for item_id in self.all_tree.get_children():
            self.all_tree.set(item_id, "chk", "v")
            count += 1
        self._log(f"[SELECT] ALL select all ({count})")

    def _deselect_all_all_accounts(self) -> None:
        count = 0
        for item_id in self.all_tree.get_children():
            self.all_tree.set(item_id, "chk", "")
            count += 1
        self._log(f"[SELECT] ALL deselect all ({count})")

    def _set_checked_selected_all(self, checked: bool) -> None:
        mark = "v" if checked else ""
        for item_id in self.all_tree.selection():
            self.all_tree.set(item_id, "chk", mark)

    def _set_checked_selected(self, checked: bool) -> None:
        mark = "v" if checked else ""
        for item_id in self.tree.selection():
            self.tree.set(item_id, "chk", mark)

    def _close_cell_editor(self, save: bool) -> None:
        editor = self._cell_editor
        if not editor:
            return
        entry = editor.get("entry")
        if save and entry:
            try:
                self._apply_cell_edit(editor, entry.get())
            except Exception:
                pass
        try:
            if entry:
                entry.destroy()
        except Exception:
            pass
        self._cell_editor = None

    def _apply_cell_edit(self, editor: dict, new_value: str) -> None:
        tree = editor.get("tree")
        item_id = editor.get("item_id")
        col_name = editor.get("col_name")
        if not tree or not item_id or not col_name:
            return
        old_value = tree.set(item_id, col_name)
        tree.set(item_id, col_name, new_value)
        try:
            idx = int(item_id) - 1
        except Exception:
            idx = None
        if tree == self.tree:
            if idx is None or idx >= len(self.accounts):
                return
            if col_name == "email":
                self.accounts[idx]["uid"] = new_value
            elif col_name in ("pass", "proxy", "youtube"):
                self.accounts[idx][col_name] = new_value
            self._save_accounts_cache()
        elif tree == self.profile_tree:
            if idx is None or idx >= len(self.profile_accounts):
                return
            if col_name == "email":
                self.profile_accounts[idx]["uid"] = new_value
            elif col_name in ("pass", "proxy", "youtube"):
                self.profile_accounts[idx][col_name] = new_value
            self._save_profile_accounts_cache()
        elif tree == self.fb_tree:
            if idx is None or idx >= len(self.fb_accounts):
                return
            if col_name == "email":
                self.fb_accounts[idx]["uid"] = new_value
            elif col_name in ("pass", "proxy", "facebook"):
                self.fb_accounts[idx][col_name] = new_value
            self._save_fb_accounts_cache()
        elif tree == self.fb_profile_tree:
            if idx is None or idx >= len(self.fb_profile_accounts):
                return
            if col_name == "email":
                self.fb_profile_accounts[idx]["uid"] = new_value
            elif col_name in ("pass", "proxy", "facebook"):
                self.fb_profile_accounts[idx][col_name] = new_value
            self._save_fb_profile_accounts_cache()
        elif tree == self.all_tree:
            social = (self.all_tree.set(item_id, "social") or "").strip().upper()
            email = (self.all_tree.set(item_id, "email") or "").strip()
            old_email = (old_value if col_name == "email" else email).strip()
            if social and old_email:
                if social == "YTB":
                    updated = False
                    for acc in self.accounts:
                        if (acc.get("uid") or "").strip() == old_email:
                            if col_name == "email":
                                acc["uid"] = new_value
                            elif col_name in {"pass", "proxy"}:
                                acc[col_name] = new_value
                            elif col_name == "link":
                                acc["youtube"] = new_value
                            updated = True
                            break
                    if updated:
                        self._save_accounts_cache()
                        try:
                            iid = self._map_email_to_item_id(self.tree).get(old_email)
                            if iid:
                                target_col = "youtube" if col_name == "link" else col_name
                                self.tree.set(iid, target_col, new_value)
                        except Exception:
                            pass
                elif social == "FB":
                    updated = False
                    for acc in self.fb_accounts:
                        if (acc.get("uid") or "").strip() == old_email:
                            if col_name == "email":
                                acc["uid"] = new_value
                            elif col_name in {"pass", "proxy"}:
                                acc[col_name] = new_value
                            elif col_name == "link":
                                acc["facebook"] = new_value
                            updated = True
                            break
                    if updated:
                        self._save_fb_accounts_cache()
                        try:
                            iid = self._map_email_to_item_id(self.fb_tree).get(old_email)
                            if iid:
                                target_col = "facebook" if col_name == "link" else col_name
                                self.fb_tree.set(iid, target_col, new_value)
                        except Exception:
                            pass
        elif tree == self.manage_tree:
            try:
                idx = int(item_id) - 1
            except Exception:
                return
            if idx < 0 or idx >= len(self._manage_rows):
                return
            self._manage_rows[idx][col_name] = new_value
            self._update_manage_counts()

    def _begin_cell_edit(self, tree: ttk.Treeview, item_id: str, col_name: str) -> None:
        self._close_cell_editor(save=True)
        bbox = tree.bbox(item_id, col_name)
        if not bbox:
            return
        x, y, w, h = bbox
        entry = tk.Entry(tree)
        entry.insert(0, tree.set(item_id, col_name))
        entry.place(x=x, y=y, width=w, height=h)
        entry.focus_set()
        entry.selection_range(0, tk.END)
        self._cell_editor = {"tree": tree, "item_id": item_id, "col_name": col_name, "entry": entry}

        def _save(_evt=None):
            self._close_cell_editor(save=True)

        def _cancel(_evt=None):
            self._close_cell_editor(save=False)

        entry.bind("<Return>", _save)
        entry.bind("<Escape>", _cancel)
        entry.bind("<FocusOut>", _save)

    def _on_tree_click(self, event) -> None:
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        column = self.tree.identify_column(event.x)
        row = self.tree.identify_row(event.y)
        if row:
            try:
                col_idx = int(column[1:]) - 1
                col_name = self.tree["columns"][col_idx]
            except Exception:
                col_name = ""
            if col_name in {"email", "pass", "proxy", "youtube"}:
                self._begin_cell_edit(self.tree, row, col_name)
                return "break"
        if column == "#1":
            if row:
                self._toggle_checked(row)
                return "break"
        if row:
            self._dragging = True
            self._drag_start = row

    def _on_tree_drag(self, event) -> None:
        if not self._dragging or not self._drag_start:
            return
        row = self.tree.identify_row(event.y)
        if not row:
            return
        children = list(self.tree.get_children())
        try:
            start_idx = children.index(self._drag_start)
            cur_idx = children.index(row)
        except ValueError:
            return
        lo = min(start_idx, cur_idx)
        hi = max(start_idx, cur_idx)
        self.tree.selection_set(children[lo : hi + 1])

    def _on_tree_release(self, event) -> None:
        self._dragging = False
        self._drag_start = None

    def _on_tree_right_click(self, event) -> None:
        row = self.tree.identify_row(event.y)
        if row:
            if row not in self.tree.selection():
                self.tree.selection_set(row)
            self._context_item = row
            self.menu.tk_popup(event.x_root, event.y_root)

    def _on_all_tree_click(self, event) -> None:
        region = self.all_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        column = self.all_tree.identify_column(event.x)
        row = self.all_tree.identify_row(event.y)
        if row:
            try:
                col_idx = int(column[1:]) - 1
                col_name = self.all_tree["columns"][col_idx]
            except Exception:
                col_name = ""
            if col_name in {"email", "pass", "proxy", "link"}:
                self._begin_cell_edit(self.all_tree, row, col_name)
                return "break"
        if column == "#1":
            if row:
                self._toggle_checked_all(row)
                return "break"
        if row:
            self._all_dragging = True
            self._all_drag_start = row

    def _on_all_tree_drag(self, event) -> None:
        if not self._all_dragging or not self._all_drag_start:
            return
        row = self.all_tree.identify_row(event.y)
        if not row:
            return
        children = list(self.all_tree.get_children())
        try:
            start_idx = children.index(self._all_drag_start)
            cur_idx = children.index(row)
        except ValueError:
            return
        lo = min(start_idx, cur_idx)
        hi = max(start_idx, cur_idx)
        self.all_tree.selection_set(children[lo : hi + 1])

    def _on_all_tree_release(self, event) -> None:
        self._all_dragging = False
        self._all_drag_start = None

    def _on_all_tree_right_click(self, event) -> None:
        row = self.all_tree.identify_row(event.y)
        if row:
            if row not in self.all_tree.selection():
                self.all_tree.selection_set(row)
            self._all_context_item = row
            self.all_menu.tk_popup(event.x_root, event.y_root)

    def _toggle_checked_profile(self, item_id: str) -> None:
        cur = self.profile_tree.set(item_id, "chk")
        self.profile_tree.set(item_id, "chk", "" if cur == "v" else "v")

    def _select_all_profile_accounts(self) -> None:
        count = 0
        for item_id in self.profile_tree.get_children():
            self.profile_tree.set(item_id, "chk", "v")
            count += 1
        self._log(f"[SELECT] PROFILE select all ({count})")

    def _deselect_all_profile_accounts(self) -> None:
        count = 0
        for item_id in self.profile_tree.get_children():
            self.profile_tree.set(item_id, "chk", "")
            count += 1
        self._log(f"[SELECT] PROFILE deselect all ({count})")

    def _set_checked_selected_profile(self, checked: bool) -> None:
        mark = "v" if checked else ""
        for item_id in self.profile_tree.selection():
            self.profile_tree.set(item_id, "chk", mark)

    def _on_profile_tree_click(self, event) -> None:
        region = self.profile_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        column = self.profile_tree.identify_column(event.x)
        row = self.profile_tree.identify_row(event.y)
        if row:
            try:
                col_idx = int(column[1:]) - 1
                col_name = self.profile_tree["columns"][col_idx]
            except Exception:
                col_name = ""
            if col_name in {"email", "pass", "proxy", "youtube"}:
                self._begin_cell_edit(self.profile_tree, row, col_name)
                return "break"
        if column == "#1":
            if row:
                self._toggle_checked_profile(row)
                return "break"
        if row:
            self._profile_dragging = True
            self._profile_drag_start = row

    def _on_profile_tree_drag(self, event) -> None:
        if not self._profile_dragging or not self._profile_drag_start:
            return
        row = self.profile_tree.identify_row(event.y)
        if not row:
            return
        children = list(self.profile_tree.get_children())
        try:
            start_idx = children.index(self._profile_drag_start)
            cur_idx = children.index(row)
        except ValueError:
            return
        lo = min(start_idx, cur_idx)
        hi = max(start_idx, cur_idx)
        self.profile_tree.selection_set(children[lo : hi + 1])

    def _on_profile_tree_release(self, event) -> None:
        self._profile_dragging = False
        self._profile_drag_start = None

    def _on_profile_tree_right_click(self, event) -> None:
        row = self.profile_tree.identify_row(event.y)
        if row:
            if row not in self.profile_tree.selection():
                self.profile_tree.selection_set(row)
            self._profile_context_item = row
            self.profile_menu.tk_popup(event.x_root, event.y_root)

    def _toggle_checked_fb(self, item_id: str) -> None:
        cur = self.fb_tree.set(item_id, "chk")
        self.fb_tree.set(item_id, "chk", "" if cur == "v" else "v")

    def _set_checked_selected_fb(self, checked: bool) -> None:
        mark = "v" if checked else ""
        for item_id in self.fb_tree.selection():
            self.fb_tree.set(item_id, "chk", mark)

    def _select_all_fb_accounts(self) -> None:
        count = 0
        for item_id in self.fb_tree.get_children():
            self.fb_tree.set(item_id, "chk", "v")
            count += 1
        self._log(f"[SELECT] FB select all ({count})")

    def _deselect_all_fb_accounts(self) -> None:
        count = 0
        for item_id in self.fb_tree.get_children():
            self.fb_tree.set(item_id, "chk", "")
            count += 1
        self._log(f"[SELECT] FB deselect all ({count})")

    def _on_fb_tree_click(self, event) -> None:
        region = self.fb_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        column = self.fb_tree.identify_column(event.x)
        row = self.fb_tree.identify_row(event.y)
        if row:
            try:
                col_idx = int(column[1:]) - 1
                col_name = self.fb_tree["columns"][col_idx]
            except Exception:
                col_name = ""
            if col_name in {"email", "pass", "proxy", "facebook"}:
                self._begin_cell_edit(self.fb_tree, row, col_name)
                return "break"
        if column == "#1" and row:
            self._toggle_checked_fb(row)
            return "break"
        if row:
            self._fb_dragging = True
            self._fb_drag_start = row

    def _on_fb_tree_right_click(self, event) -> None:
        row = self.fb_tree.identify_row(event.y)
        if row:
            if row not in self.fb_tree.selection():
                self.fb_tree.selection_set(row)
            self._fb_context_item = row
            self.fb_menu.tk_popup(event.x_root, event.y_root)

    def _on_fb_tree_drag(self, event) -> None:
        if not self._fb_dragging or not self._fb_drag_start:
            return
        row = self.fb_tree.identify_row(event.y)
        if not row:
            return
        children = list(self.fb_tree.get_children())
        try:
            start_idx = children.index(self._fb_drag_start)
            cur_idx = children.index(row)
        except ValueError:
            return
        lo = min(start_idx, cur_idx)
        hi = max(start_idx, cur_idx)
        self.fb_tree.selection_set(children[lo : hi + 1])

    def _on_fb_tree_release(self, event) -> None:
        self._fb_dragging = False
        self._fb_drag_start = None

    def _toggle_checked_fb_profile(self, item_id: str) -> None:
        cur = self.fb_profile_tree.set(item_id, "chk")
        self.fb_profile_tree.set(item_id, "chk", "" if cur == "v" else "v")

    def _set_checked_selected_fb_profile(self, checked: bool) -> None:
        mark = "v" if checked else ""
        for item_id in self.fb_profile_tree.selection():
            self.fb_profile_tree.set(item_id, "chk", mark)

    def _select_all_fb_profile_accounts(self) -> None:
        count = 0
        for item_id in self.fb_profile_tree.get_children():
            self.fb_profile_tree.set(item_id, "chk", "v")
            count += 1
        self._log(f"[SELECT] FB PROFILE select all ({count})")

    def _deselect_all_fb_profile_accounts(self) -> None:
        count = 0
        for item_id in self.fb_profile_tree.get_children():
            self.fb_profile_tree.set(item_id, "chk", "")
            count += 1
        self._log(f"[SELECT] FB PROFILE deselect all ({count})")

    def _on_fb_profile_tree_click(self, event) -> None:
        region = self.fb_profile_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        column = self.fb_profile_tree.identify_column(event.x)
        row = self.fb_profile_tree.identify_row(event.y)
        if row:
            try:
                col_idx = int(column[1:]) - 1
                col_name = self.fb_profile_tree["columns"][col_idx]
            except Exception:
                col_name = ""
            if col_name in {"email", "pass", "proxy", "facebook"}:
                self._begin_cell_edit(self.fb_profile_tree, row, col_name)
                return "break"
        if column == "#1" and row:
            self._toggle_checked_fb_profile(row)
            return "break"
        if row:
            self._fb_profile_dragging = True
            self._fb_profile_drag_start = row

    def _on_fb_profile_tree_right_click(self, event) -> None:
        row = self.fb_profile_tree.identify_row(event.y)
        if row:
            if row not in self.fb_profile_tree.selection():
                self.fb_profile_tree.selection_set(row)
            self.fb_profile_menu.tk_popup(event.x_root, event.y_root)

    def _on_fb_profile_tree_drag(self, event) -> None:
        if not self._fb_profile_dragging or not self._fb_profile_drag_start:
            return
        row = self.fb_profile_tree.identify_row(event.y)
        if not row:
            return
        children = list(self.fb_profile_tree.get_children())
        try:
            start_idx = children.index(self._fb_profile_drag_start)
            cur_idx = children.index(row)
        except ValueError:
            return
        lo = min(start_idx, cur_idx)
        hi = max(start_idx, cur_idx)
        self.fb_profile_tree.selection_set(children[lo : hi + 1])

    def _on_fb_profile_tree_release(self, event) -> None:
        self._fb_profile_dragging = False
        self._fb_profile_drag_start = None

    def _get_selected_accounts(self):
        items = []
        acc_by_email = {str(a.get("uid") or "").strip(): a for a in self.accounts}
        for iid in self.tree.selection():
            email = (self.tree.set(iid, "email") or "").strip()
            acc = acc_by_email.get(email)
            if acc:
                items.append((iid, acc))
        return items

    def _get_selected_profile_accounts(self):
        items = []
        acc_by_email = {str(a.get("uid") or "").strip(): a for a in self.profile_accounts}
        for iid in self.profile_tree.selection():
            email = (self.profile_tree.set(iid, "email") or "").strip()
            acc = acc_by_email.get(email)
            if acc:
                items.append((iid, acc))
        return items

    def _get_checked_accounts(self):
        items = []
        acc_by_email = {str(a.get("uid") or "").strip(): a for a in self.accounts}
        for iid in self.tree.get_children():
            if self.tree.set(iid, "chk") != "v":
                continue
            email = (self.tree.set(iid, "email") or "").strip()
            acc = acc_by_email.get(email)
            if acc:
                items.append((iid, acc))
        return items

    def _get_selected_fb_accounts(self):
        items = []
        acc_by_email = {str(a.get("uid") or "").strip(): a for a in self.fb_accounts}
        for iid in self.fb_tree.selection():
            email = (self.fb_tree.set(iid, "email") or "").strip()
            acc = acc_by_email.get(email)
            if acc:
                items.append((iid, acc))
        return items

    def _get_checked_fb_accounts(self):
        items = []
        acc_by_email = {str(a.get("uid") or "").strip(): a for a in self.fb_accounts}
        for iid in self.fb_tree.get_children():
            if self.fb_tree.set(iid, "chk") != "v":
                continue
            email = (self.fb_tree.set(iid, "email") or "").strip()
            acc = acc_by_email.get(email)
            if acc:
                items.append((iid, acc))
        return items

    def _get_context_fb_accounts(self):
        if self._fb_context_item:
            email = (self.fb_tree.set(self._fb_context_item, "email") or "").strip()
            acc_by_email = {str(a.get("uid") or "").strip(): a for a in self.fb_accounts}
            acc = acc_by_email.get(email)
            if acc:
                return [(self._fb_context_item, acc)]
        return []

    def _get_selected_all_rows(self):
        items = []
        for iid in self.all_tree.selection():
            email = (self.all_tree.set(iid, "email") or "").strip()
            social = (self.all_tree.set(iid, "social") or "").strip().upper()
            if email and social:
                items.append((social, email))
        return items

    def _get_context_all_rows(self):
        if self._all_context_item:
            email = (self.all_tree.set(self._all_context_item, "email") or "").strip()
            social = (self.all_tree.set(self._all_context_item, "social") or "").strip().upper()
            if email and social:
                return [(social, email)]
        return []

    def _get_checked_profile_accounts(self):
        items = []
        acc_by_email = {str(a.get("uid") or "").strip(): a for a in self.profile_accounts}
        for iid in self.profile_tree.get_children():
            if self.profile_tree.set(iid, "chk") != "v":
                continue
            email = (self.profile_tree.set(iid, "email") or "").strip()
            acc = acc_by_email.get(email)
            if acc:
                items.append((iid, acc))
        return items

    def _get_context_accounts(self):
        if self._context_item:
            email = (self.tree.set(self._context_item, "email") or "").strip()
            acc_by_email = {str(a.get("uid") or "").strip(): a for a in self.accounts}
            acc = acc_by_email.get(email)
            if acc:
                return [(self._context_item, acc)]
        return []

    def _get_context_profile_accounts(self):
        if self._profile_context_item:
            email = (self.profile_tree.set(self._profile_context_item, "email") or "").strip()
            acc_by_email = {str(a.get("uid") or "").strip(): a for a in self.profile_accounts}
            acc = acc_by_email.get(email)
            if acc:
                return [(self._profile_context_item, acc)]
        return []

    def menu_login_selected(self) -> None:
        if self.executor is not None:
            self._log("[MENU] Dang chay job, hay STOP truoc.")
            return
        self.stop_event.clear()
        selected = self._get_context_accounts() or self._get_selected_accounts() or self._get_checked_accounts()
        if not selected:
            return
        max_threads = max(1, int(self.entry_threads.get() or 1))
        pool = ThreadPoolExecutor(max_workers=max_threads)
        for item_id, acc in selected:
            self._bind_item_email(item_id, acc.get("uid", ""))
            pool.submit(self._login_only_worker, item_id, acc)

    def menu_upload_selected(self) -> None:
        if self.executor is not None:
            self._log("[MENU] Dang chay job, hay STOP truoc.")
            return
        self.stop_event.clear()
        selected = self._get_context_accounts() or self._get_selected_accounts() or self._get_checked_accounts()
        if not selected:
            return
        max_threads = max(1, int(self.entry_threads.get() or 1))
        pool = ThreadPoolExecutor(max_workers=max_threads)
        for item_id, acc in selected:
            self._bind_item_email(item_id, acc.get("uid", ""))
            pool.submit(self._upload_only_worker, item_id, acc)

    def menu_follow_selected(self) -> None:
        if self.executor is not None:
            self._log("[MENU] Dang chay job, hay STOP truoc.")
            return
        self.stop_event.clear()
        selected = self._get_context_accounts() or self._get_selected_accounts() or self._get_checked_accounts()
        if not selected:
            return
        max_threads = max(1, int(self.entry_threads.get() or 1))
        pool = ThreadPoolExecutor(max_workers=max_threads)
        for item_id, acc in selected:
            self._bind_item_email(item_id, acc.get("uid", ""))
            pool.submit(self._follow_only_worker, item_id, acc)

    def menu_fb_login_selected(self) -> None:
        if self.executor is not None:
            self._log("[MENU] Dang chay job, hay STOP truoc.")
            return
        self.stop_event.clear()
        selected = self._get_context_fb_accounts() or self._get_selected_fb_accounts() or self._get_checked_fb_accounts()
        if not selected:
            return
        max_threads = max(1, int(self.entry_threads.get() or 1))
        pool = ThreadPoolExecutor(max_workers=max_threads)
        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080
        gap = 6
        taskbar_h = 40
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(selected)
        cols = min(5, active_count)
        rows_layout = min(2, max(1, math.ceil(active_count / cols)))
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))
        for idx, (item_id, acc) in enumerate(selected):
            pos = idx % (cols * rows_layout)
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            win_pos = f"{x},{y}"
            win_size = f"{win_w},{win_h}"
            self._bind_item_email(item_id, acc.get("uid", ""))
            pool.submit(self._fb_login_only_worker, item_id, acc, win_pos, win_size)

    def menu_fb_upload_selected(self) -> None:
        if self.executor is not None:
            self._log("[MENU] Dang chay job, hay STOP truoc.")
            return
        self.stop_event.clear()
        selected = self._get_context_fb_accounts() or self._get_selected_fb_accounts() or self._get_checked_fb_accounts()
        if not selected:
            return
        try:
            max_videos = int(self.entry_videos.get())
            if max_videos <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Loi", "Videos phai > 0")
            return
        max_threads = max(1, int(self.entry_threads.get() or 1))
        pool = ThreadPoolExecutor(max_workers=max_threads)
        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080
        gap = 6
        taskbar_h = 40
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(selected)
        cols = min(5, active_count)
        rows_layout = min(2, max(1, math.ceil(active_count / cols)))
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))
        for idx, (item_id, acc) in enumerate(selected):
            pos = idx % (cols * rows_layout)
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            win_pos = f"{x},{y}"
            win_size = f"{win_w},{win_h}"
            self._bind_item_email(item_id, acc.get("uid", ""))
            pool.submit(self._fb_worker_one, item_id, acc, win_pos, win_size, max_videos)

    def menu_fb_follow_selected(self) -> None:
        if self.executor is not None:
            self._log("[MENU] Dang chay job, hay STOP truoc.")
            return
        self.stop_event.clear()
        selected = self._get_context_fb_accounts() or self._get_selected_fb_accounts() or self._get_checked_fb_accounts()
        if not selected:
            return
        max_threads = max(1, int(self.entry_threads.get() or 1))
        pool = ThreadPoolExecutor(max_workers=max_threads)
        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080
        gap = 6
        taskbar_h = 40
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(selected)
        cols = min(5, active_count)
        rows_layout = min(2, max(1, math.ceil(active_count / cols)))
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))
        for idx, (item_id, acc) in enumerate(selected):
            pos = idx % (cols * rows_layout)
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            win_pos = f"{x},{y}"
            win_size = f"{win_w},{win_h}"
            self._bind_item_email(item_id, acc.get("uid", ""))
            pool.submit(self._fb_follow_only_worker, item_id, acc, win_pos, win_size)

    def menu_all_login_selected(self) -> None:
        if self.executor is not None:
            self._log("[MENU] Dang chay job, hay STOP truoc.")
            return
        self.stop_event.clear()
        selected = self._get_context_all_rows() or self._get_selected_all_rows() or self._get_checked_all_rows()
        if not selected:
            return
        ytb_emails = [email for social, email in selected if social == "YTB"]
        fb_emails = [email for social, email in selected if social == "FB"]
        max_threads = max(1, int(self.entry_threads.get() or 1))
        pool = ThreadPoolExecutor(max_workers=max_threads)
        acc_by_email = {str(a.get("uid") or "").strip(): a for a in self.accounts}
        email_to_iid = self._map_email_to_item_id(self.tree)
        for email in ytb_emails:
            acc = acc_by_email.get(email)
            item_id = email_to_iid.get(email)
            if acc and item_id:
                self._bind_item_email(item_id, acc.get("uid", ""))
                pool.submit(self._login_only_worker, item_id, acc)
        if fb_emails:
            try:
                screen_w = self.root.winfo_screenwidth()
                screen_h = self.root.winfo_screenheight()
            except Exception:
                screen_w, screen_h = 1920, 1080
            gap = 6
            taskbar_h = 40
            usable_w = screen_w - (gap * 2)
            usable_h = (screen_h - taskbar_h) - (gap * 2)
            active_count = len(fb_emails)
            cols = min(5, active_count)
            rows_layout = min(2, max(1, math.ceil(active_count / cols)))
            win_w = int((usable_w - gap * (cols - 1)) / cols)
            win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
            win_w = max(150, min(280, win_w))
            win_h = max(420, min(600, win_h))
            fb_by_email = {str(a.get("uid") or "").strip(): a for a in self.fb_accounts}
            fb_email_to_iid = self._map_email_to_item_id(self.fb_tree)
            for idx, email in enumerate(fb_emails):
                acc = fb_by_email.get(email)
                item_id = fb_email_to_iid.get(email)
                if not acc or not item_id:
                    continue
                pos = idx % (cols * rows_layout)
                col = pos % cols
                row = pos // cols
                x = gap + col * (win_w + gap)
                y = gap + row * (win_h + gap)
                win_pos = f"{x},{y}"
                win_size = f"{win_w},{win_h}"
                self._bind_item_email(item_id, acc.get("uid", ""))
                pool.submit(self._fb_login_only_worker, item_id, acc, win_pos, win_size)

    def menu_all_upload_selected(self) -> None:
        if self.executor is not None:
            self._log("[MENU] Dang chay job, hay STOP truoc.")
            return
        self.stop_event.clear()
        selected = self._get_context_all_rows() or self._get_selected_all_rows() or self._get_checked_all_rows()
        if not selected:
            return
        ytb_emails = [email for social, email in selected if social == "YTB"]
        fb_emails = [email for social, email in selected if social == "FB"]
        if ytb_emails:
            max_threads = max(1, int(self.entry_threads.get() or 1))
            pool = ThreadPoolExecutor(max_workers=max_threads)
            acc_by_email = {str(a.get("uid") or "").strip(): a for a in self.accounts}
            email_to_iid = self._map_email_to_item_id(self.tree)
            for email in ytb_emails:
                acc = acc_by_email.get(email)
                item_id = email_to_iid.get(email)
                if acc and item_id:
                    self._bind_item_email(item_id, acc.get("uid", ""))
                    pool.submit(self._upload_only_worker, item_id, acc)
        if fb_emails:
            self._set_checked_by_email(self.fb_tree, set(fb_emails))
            self.start_fb_jobs()

    def menu_all_follow_selected(self) -> None:
        if self.executor is not None:
            self._log("[MENU] Dang chay job, hay STOP truoc.")
            return
        self.stop_event.clear()
        selected = self._get_context_all_rows() or self._get_selected_all_rows() or self._get_checked_all_rows()
        if not selected:
            return
        ytb_emails = [email for social, email in selected if social == "YTB"]
        fb_emails = [email for social, email in selected if social == "FB"]
        max_threads = max(1, int(self.entry_threads.get() or 1))
        pool = ThreadPoolExecutor(max_workers=max_threads)
        acc_by_email = {str(a.get("uid") or "").strip(): a for a in self.accounts}
        email_to_iid = self._map_email_to_item_id(self.tree)
        for email in ytb_emails:
            acc = acc_by_email.get(email)
            item_id = email_to_iid.get(email)
            if acc and item_id:
                self._bind_item_email(item_id, acc.get("uid", ""))
                pool.submit(self._follow_only_worker, item_id, acc)
        if fb_emails:
            try:
                screen_w = self.root.winfo_screenwidth()
                screen_h = self.root.winfo_screenheight()
            except Exception:
                screen_w, screen_h = 1920, 1080
            gap = 6
            taskbar_h = 40
            usable_w = screen_w - (gap * 2)
            usable_h = (screen_h - taskbar_h) - (gap * 2)
            active_count = len(fb_emails)
            cols = min(5, active_count)
            rows_layout = min(2, max(1, math.ceil(active_count / cols)))
            win_w = int((usable_w - gap * (cols - 1)) / cols)
            win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
            win_w = max(150, min(280, win_w))
            win_h = max(420, min(600, win_h))
            fb_by_email = {str(a.get("uid") or "").strip(): a for a in self.fb_accounts}
            fb_email_to_iid = self._map_email_to_item_id(self.fb_tree)
            for idx, email in enumerate(fb_emails):
                acc = fb_by_email.get(email)
                item_id = fb_email_to_iid.get(email)
                if not acc or not item_id:
                    continue
                pos = idx % (cols * rows_layout)
                col = pos % cols
                row = pos // cols
                x = gap + col * (win_w + gap)
                y = gap + row * (win_h + gap)
                win_pos = f"{x},{y}"
                win_size = f"{win_w},{win_h}"
                self._bind_item_email(item_id, acc.get("uid", ""))
                pool.submit(self._fb_follow_only_worker, item_id, acc, win_pos, win_size)

    def menu_all_replace_proxy_errors(self) -> None:
        self._replace_proxy_errors("upload")
        self._replace_proxy_errors("fb")

    def menu_profile_selected(self) -> None:
        if self.executor is not None:
            self._log("[MENU] Dang chay job, hay STOP truoc.")
            return
        if not self._is_profile_tab():
            self._log("[MENU] Vao tab PROFILE de chay.")
            return
        selected = (
            self._get_context_profile_accounts()
            or self._get_selected_profile_accounts()
            or self._get_checked_profile_accounts()
        )
        if not selected:
            return
        max_threads = max(1, int(self.entry_threads.get() or 1))
        pool = ThreadPoolExecutor(max_workers=max_threads)
        self.profile_semaphore = threading.BoundedSemaphore(max_threads)
        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080
        gap = 6  # Tight spacing between windows
        taskbar_h = 40  # Reserve space for taskbar
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(selected)
        cols = min(5, active_count)
        rows_layout = min(2, max(1, math.ceil(active_count / cols)))
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))
        slot_idx = 0
        max_slots = cols * rows_layout
        for idx_item, (item_id, acc) in enumerate(selected):
            pos = slot_idx % max_slots
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            win_pos = f"{x},{y}"
            win_size = f"{win_w},{win_h}"
            pool.submit(self._profile_open_worker, item_id, acc, win_pos, win_size)
            slot_idx += 1

    def _load_accounts_cache(self) -> list:
        data = self._load_cache_list("accounts")
        if isinstance(data, list):
            # Normalize specific YouTube URLs that should always point to Shorts.
            try:
                for acc in data:
                    acc["payment_status"] = self._normalize_payment_status(acc.get("payment_status"))
                    if acc.get("uid") == "opendauria@hotmail.com":
                        yt = (acc.get("youtube") or "").strip()
                        if yt == "https://www.youtube.com/@Vice_Verses":
                            acc["youtube"] = "https://www.youtube.com/@Vice_Verses/shorts"
            except Exception:
                pass
        return data

    def _save_accounts_cache(self) -> None:
        self._save_cache_list("accounts", self.accounts)

    def _load_profile_accounts_cache(self) -> list:
        return self._load_cache_list("profile_accounts")

    def _save_profile_accounts_cache(self) -> None:
        self._save_cache_list("profile_accounts", self.profile_accounts)

    def _load_fb_accounts_cache(self) -> list:
        data = self._load_cache_list("fb_accounts")
        if isinstance(data, list):
            try:
                for acc in data:
                    acc["payment_status"] = self._normalize_payment_status(acc.get("payment_status"))
            except Exception:
                pass
        return data

    def _save_fb_accounts_cache(self) -> None:
        self._save_cache_list("fb_accounts", self.fb_accounts)

    def _load_fb_profile_accounts_cache(self) -> list:
        return self._load_cache_list("fb_profile_accounts")

    def _save_fb_profile_accounts_cache(self) -> None:
        self._save_cache_list("fb_profile_accounts", self.fb_profile_accounts)

    def _clear_creator_fund_status_cache(self) -> None:
        try:
            for key in ("accounts", "fb_accounts"):
                data = self._load_cache_list(key)
                if not isinstance(data, list):
                    continue
                updated = False
                for acc in data:
                    if "creator_fund_status" in acc:
                        acc.pop("creator_fund_status", None)
                        updated = True
                if updated:
                    self._save_cache_list(key, data)
        except Exception:
            pass

    def _clear_not_applied_status_cache(self) -> None:
        """Remove only NOT_APPLIED creator fund statuses from both YTB and FB cache."""
        try:
            for key in ("accounts", "fb_accounts"):
                data = self._load_cache_list(key)
                if not isinstance(data, list):
                    continue
                updated = False
                for acc in data:
                    status = (acc.get("creator_fund_status") or "").strip().upper()
                    if status == "NOT_APPLIED":
                        acc.pop("creator_fund_status", None)
                        updated = True
                if updated:
                    self._save_cache_list(key, data)
        except Exception:
            pass

    def _normalize_payment_status(self, status: str) -> str:
        return normalize_payment_status(status)

    def _normalize_payment_status_cache(self) -> None:
        try:
            for key in ("accounts", "fb_accounts"):
                data = self._load_cache_list(key)
                if not isinstance(data, list):
                    continue
                updated = False
                for acc in data:
                    cur = (acc.get("payment_status") or "").strip().upper()
                    nxt = self._normalize_payment_status(cur)
                    if cur != nxt:
                        acc["payment_status"] = nxt
                        updated = True
                if updated:
                    self._save_cache_list(key, data)
        except Exception:
            pass

    def _reset_operational_status_cache(self) -> None:
        """Reset runtime statuses for upload/profile tabs when app starts."""
        try:
            for key in ("accounts", "profile_accounts", "fb_accounts", "fb_profile_accounts"):
                data = self._load_cache_list(key)
                if not isinstance(data, list):
                    continue
                updated = False
                for acc in data:
                    cur = (acc.get("status") or "").strip().upper()
                    if cur != "READY":
                        acc["status"] = "READY"
                        updated = True
                if updated:
                    self._save_cache_list(key, data)
        except Exception:
            pass

    def _open_cache_db(self) -> sqlite3.Connection | None:
        try:
            conn = sqlite3.connect(self._cache_db, timeout=5)
            conn.execute("PRAGMA journal_mode=WAL;")
            conn.execute("PRAGMA synchronous=NORMAL;")
            conn.execute(
                "CREATE TABLE IF NOT EXISTS cache (key TEXT PRIMARY KEY, data TEXT, updated_at REAL)"
            )
            return conn
        except Exception:
            return None

    def _cache_db_has_data(self) -> bool:
        try:
            with self._cache_db_lock:
                conn = self._open_cache_db()
                if not conn:
                    return False
                try:
                    cur = conn.execute("SELECT data FROM cache")
                    rows = cur.fetchall()
                    if not rows:
                        return False
                    for (payload,) in rows:
                        try:
                            loaded = json.loads(payload or "[]")
                            if isinstance(loaded, list) and len(loaded) > 0:
                                return True
                        except Exception:
                            continue
                    return False
                finally:
                    conn.close()
        except Exception:
            return False

    def _load_cache_list(self, key: str) -> list:
        data = []
        try:
            with self._cache_db_lock:
                conn = self._open_cache_db()
                if conn:
                    try:
                        cur = conn.execute("SELECT data FROM cache WHERE key = ?", (key,))
                        row = cur.fetchone()
                        if row and row[0]:
                            loaded = json.loads(row[0])
                            if isinstance(loaded, list):
                                data = loaded
                                return data
                    finally:
                        conn.close()
        except Exception:
            pass

    def _should_check_creator_fund(self, acc: dict, followers) -> bool:
        if followers is None or str(followers).strip() == "":
            followers = acc.get("followers")
        fnum = self._followers_to_int(followers)
        if fnum < 1000:
            return False
        status = (acc.get("creator_fund_status") or "").strip().upper()
        if status in {"JOINED", "PENDING"}:
            return False
        return True

    def _creator_fund_status_from_page(self, page_text: str) -> str:
        text = (page_text or "").lower()
        if "joined the creator fund" in text or "you\u2019ve joined the creator fund" in text:
            return "JOINED"
        if "application under review" in text:
            return "PENDING"
        return ""

    def _creator_fund_bars_full(self, page_text: str) -> bool:
        text = (page_text or "").lower()

        def _percent_for(label: str) -> float:
            m = re.search(label + r"[^%]*?([0-9]+(?:\\.[0-9]+)?)%", text, re.DOTALL)
            if not m:
                return 0.0
            try:
                return float(m.group(1))
            except Exception:
                return 0.0

        return (
            _percent_for("followers") >= 100.0
            and _percent_for("views") >= 100.0
            and _percent_for("videos") >= 100.0
        )

    def _set_creator_fund_status(self, acc: dict, source: str, status: str) -> None:
        try:
            status = (status or "").strip().upper()
            if not status:
                return
            acc["creator_fund_status"] = status
            # Persist directly to DB by uid to avoid stale RAM data
            uid = (acc.get("uid") or "").strip().lower()
            if uid:
                key = "fb_accounts" if source == "FB" else "accounts"
                data = self._load_cache_list(key)
                if isinstance(data, list):
                    for row in data:
                        if (row.get("uid") or "").strip().lower() == uid:
                            row["creator_fund_status"] = status
                            break
                    self._save_cache_list(key, data)
            self._refresh_stats()
        except Exception:
            pass

    def _set_payment_status(self, acc: dict, source: str, status: str) -> None:
        try:
            status = self._normalize_payment_status(status)
            if not status:
                return
            acc["payment_status"] = status
            uid = (acc.get("uid") or "").strip().lower()
            if uid:
                key = "fb_accounts" if source == "FB" else "accounts"
                data = self._load_cache_list(key)
                if isinstance(data, list):
                    for row in data:
                        if (row.get("uid") or "").strip().lower() == uid:
                            row["payment_status"] = status
                            break
                    self._save_cache_list(key, data)
            self._refresh_stats()
        except Exception:
            pass

    def _close_gpm_profile_for_acc(self, acc: dict) -> None:
        try:
            profile_id = (acc.get("profile_id") or "").strip()
            if not profile_id:
                return
            try:
                close_profile(profile_id, 3)
            except Exception:
                pass
            try:
                delete_profile(profile_id, 10)
            except Exception:
                pass
            try:
                self.created_profiles.discard(profile_id)
            except Exception:
                pass
            try:
                self._delete_profile_path(profile_id)
            except Exception:
                pass
        except Exception:
            pass

    def _check_creator_fund(self, driver_path: str, remote: str, acc: dict, source: str) -> None:
        email = (acc.get("uid") or "").strip()
        if not email:
            return
        if self.stop_event.is_set():
            return
        with self._creator_fund_lock:
            if email in self._creator_fund_checked:
                return
            self._creator_fund_checked.add(email)
        try:
            from selenium import webdriver
            from selenium.webdriver.chrome.service import Service as ChromeService
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
        except Exception as e:
            self._log(f"[{email}] CREATOR FUND ERR: {e}")
            return

        status = ""
        options = webdriver.ChromeOptions()
        options.add_experimental_option("debuggerAddress", remote.strip())
        driver = webdriver.Chrome(service=ChromeService(driver_path), options=options)
        wait = WebDriverWait(driver, 12)
        try:
            if self.stop_event.is_set():
                return
            def _close_tutorial_if_any() -> None:
                # Close "How to apply" or other onboarding modal if it appears
                try:
                    close_btns = driver.find_elements(By.XPATH, "//button[contains(., 'Close') or contains(., 'Skip') or contains(., 'Done')]")
                    for b in close_btns:
                        if b.is_displayed():
                            driver.execute_script("arguments[0].click();", b)
                            time.sleep(0.4)
                            return
                except Exception:
                    pass
                try:
                    # Some modals have an X button
                    x_btns = driver.find_elements(By.XPATH, "//button[@aria-label='Close'] | //button[contains(., '×')]")
                    for b in x_btns:
                        if b.is_displayed():
                            driver.execute_script("arguments[0].click();", b)
                            time.sleep(0.4)
                            return
                except Exception:
                    pass
                try:
                    driver.execute_script(
                        "document.body.dispatchEvent(new KeyboardEvent('keydown', {key:'Escape'}));"
                    )
                except Exception:
                    pass

            driver.get("https://thescoopz.com/creator-fund")
            time.sleep(2)
            if self.stop_event.is_set():
                return
            _close_tutorial_if_any()
            page = driver.page_source or ""
            status = self._creator_fund_status_from_page(page)
            if not status:
                try:
                    body_text = driver.execute_script("return document.body && document.body.innerText ? document.body.innerText : ''") or ""
                    status = self._creator_fund_status_from_page(body_text)
                except Exception:
                    pass
            if status:
                self._log(f"[{email}] CREATOR FUND: status={status}")
                try:
                    self._set_creator_fund_status(acc, source, status)
                    self._log(f"[{email}] CREATOR FUND: saved status={status} to DB")
                except Exception as e:
                    self._log(f"[{email}] CREATOR FUND: save status failed: {e}")
                return

            # Try to apply
            try:
                # Ensure on apply section when no status
                # Prefer Apply button in Creator Fund card
                apply_btns = driver.find_elements(
                    By.XPATH,
                    "//section//*[contains(., 'Creator Fund')]/ancestor::*[self::section or self::div][1]//button[normalize-space()='Apply' or contains(., 'Apply')]",
                )
                if not apply_btns:
                    apply_btns = driver.find_elements(By.XPATH, "//button[normalize-space()='Apply' or contains(., 'Apply')]")
                apply_clicked = False
                for btn in apply_btns:
                    try:
                        if not btn.is_displayed():
                            continue
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
                        time.sleep(0.2)
                        driver.execute_script("arguments[0].click();", btn)
                        apply_clicked = True
                        break
                    except Exception:
                        continue
                if not apply_clicked:
                    return
                time.sleep(1.5)
            except Exception:
                return

            self._log(f"[{email}] CREATOR FUND: apply clicked, opening modal")

            # Handle tutorial modal: click Next then Done
            try:
                for _ in range(2):
                    next_btns = driver.find_elements(By.XPATH, "//button[.//span[normalize-space()='Next'] or normalize-space()='Next']")
                    clicked = False
                    for b in next_btns:
                        if b.is_displayed():
                            driver.execute_script("arguments[0].click();", b)
                            clicked = True
                            time.sleep(1.0)
                            break
                    if not clicked:
                        break
                done_btns = driver.find_elements(By.XPATH, "//button[.//span[normalize-space()='Done'] or normalize-space()='Done']")
                for b in done_btns:
                    if b.is_displayed():
                        driver.execute_script("arguments[0].click();", b)
                        time.sleep(1.0)
                        break
            except Exception:
                pass
            # After Done, scroll down inside the modal content
            try:
                modal_scroll = driver.find_element(
                    By.CSS_SELECTOR,
                    "div.fixed.inset-0.top-\\[48px\\].overflow-y-auto",
                )
                driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", modal_scroll)
                time.sleep(0.8)
            except Exception:
                pass

            # Check three % blocks inside the modal; log if all 100%
            try:
                modal = driver.find_element(
                    By.CSS_SELECTOR,
                    "div.fixed.inset-0.top-\\[48px\\].overflow-y-auto",
                )
                section = None
                try:
                    section = modal.find_element(
                        By.XPATH,
                        ".//section[.//h2//span[contains(., 'Am I qualified to apply?')]]",
                    )
                except Exception:
                    section = modal

                text = ""
                try:
                    text = driver.execute_script("return arguments[0].innerText || '';", section) or ""
                except Exception:
                    text = section.text or ""

                def _get_stats(label: str) -> tuple[str, str]:
                    idx = text.lower().find(label.lower())
                    seg = text[idx: idx + 120] if idx >= 0 else text
                    pct = ""
                    ratio = ""
                    m_pct = re.search(r"([0-9]{1,3})\s*%", seg)
                    if m_pct:
                        pct = m_pct.group(1) + "%"
                    m_ratio = re.search(r"([0-9][0-9,\.]*)\s*/\s*([0-9][0-9,\.]*)", seg)
                    if m_ratio:
                        ratio = f"{m_ratio.group(1)}/{m_ratio.group(2)}"
                    return pct, ratio

                f_pct, f_ratio = _get_stats("Followers")
                v_pct, v_ratio = _get_stats("Views")
                vd_pct, vd_ratio = _get_stats("Videos")
                self._log(f"[{email}] CREATOR FUND: Followers {f_pct} {f_ratio}")
                self._log(f"[{email}] CREATOR FUND: Views {v_pct} {v_ratio}")
                self._log(f"[{email}] CREATOR FUND: Videos {vd_pct} {vd_ratio}")

                if f_pct == "100%" and v_pct == "100%" and vd_pct == "100%":
                    self._log(f"[{email}] CREATOR FUND: QUALIFIED (100% on all 3)")
                    try:
                        # Tick 2 checkboxes in the same modal section
                        boxes = modal.find_elements(By.CSS_SELECTOR, "input[type='checkbox']")
                        for cb in boxes:
                            if not cb.is_selected():
                                try:
                                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cb)
                                except Exception:
                                    pass
                                try:
                                    driver.execute_script("arguments[0].click();", cb)
                                except Exception:
                                    try:
                                        cb.click()
                                    except Exception:
                                        pass
                                time.sleep(0.3)
                        try:
                            # Verify checkboxes
                            checked = [cb for cb in boxes if cb.is_selected()]
                            self._log(f"[{email}] CREATOR FUND: checkboxes checked={len(checked)}/{len(boxes)}")
                        except Exception:
                            pass
                        try:
                            apply_btn = modal.find_element(
                                By.XPATH,
                                ".//button[normalize-space()='Apply' or contains(., 'Apply')]",
                            )
                            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", apply_btn)
                            time.sleep(0.2)
                            driver.execute_script("arguments[0].click();", apply_btn)
                            time.sleep(0.8)
                            # Default to PENDING after apply click
                            try:
                                self._set_creator_fund_status(acc, source, "PENDING")
                                self._log(f"[{email}] CREATOR FUND: saved status=PENDING to DB")
                            except Exception as e:
                                self._log(f"[{email}] CREATOR FUND: save status failed: {e}")
                        except Exception:
                            pass
                    except Exception:
                        pass
                else:
                    self._log(f"[{email}] CREATOR FUND: NOT QUALIFIED (percent check)")
            except Exception:
                pass

            
        finally:
            if not status:
                with self._creator_fund_lock:
                    self._creator_fund_checked.discard(email)
            # Session close is handled by caller cleanup to avoid tab flicker and duplicate close actions.

    def _maybe_check_creator_fund(self, driver_path: str, remote: str, acc: dict, followers, source: str, force_check: bool = False) -> None:
        try:
            if (not force_check) and (not self._should_check_creator_fund(acc, followers)):
                try:
                    email = (acc.get("uid") or "").strip()
                    self._log(f"[{email}] CREATOR FUND: skip (not eligible or already set)")
                except Exception:
                    pass
                return
            try:
                email = (acc.get("uid") or "").strip()
                self._log(f"[{email}] CREATOR FUND: start check")
            except Exception:
                pass
            self._check_creator_fund(driver_path, remote, acc, source)
        except Exception as e:
            email = (acc.get("uid") or "").strip()
            self._log(f"[{email}] CREATOR FUND ERR: {e}")

    def _payment_check_worker(self, item_id: str, acc: dict, source: str, win_pos: str = "", win_size: str = "") -> None:
        if self.stop_event.is_set():
            return
        profile_id = None
        email = (acc.get("uid") or "").strip()
        try:
            driver_path, remote, profile_id = self._ensure_logged_in(
                item_id,
                acc,
                win_pos=win_pos if win_pos else None,
                win_size=win_size if win_size else None,
            )
            if not driver_path or not remote:
                return
            if self.stop_event.is_set():
                return
            try:
                from selenium import webdriver
                from selenium.webdriver.chrome.service import Service as ChromeService
                from selenium.webdriver.support.ui import WebDriverWait
            except Exception as e:
                self._log(f"[{email}] PAYMENT ERR: {e}")
                return

            options = webdriver.ChromeOptions()
            options.add_experimental_option("debuggerAddress", remote.strip())
            driver = webdriver.Chrome(service=ChromeService(driver_path), options=options)
            wait = WebDriverWait(driver, 12)
            status = "NOT_SETUP"
            try:
                if self.stop_event.is_set():
                    return
                driver.get("https://thescoopz.com/wallet")
                try:
                    wait.until(lambda d: d.execute_script("return document.readyState") == "complete")
                except Exception:
                    pass
                time.sleep(1.2)
                try:
                    text = driver.execute_script(
                        "return document.body && document.body.innerText ? document.body.innerText : ''"
                    ) or ""
                except Exception:
                    text = ""
                page = (text or "") + " " + (driver.page_source or "")
                needle = "set up completed: your payment has been successfully set up through stripe."
                if needle in page.lower():
                    status = "SETUP"
                else:
                    status = "NOT_SETUP"
                self._set_payment_status(acc, source, status)
                self._log(f"[{email}] PAYMENT: {status}")
            except Exception as e:
                self._log(f"[{email}] PAYMENT ERR: {e}")
                self._set_payment_status(acc, source, "CHECK_ERR")
            finally:
                try:
                    driver.quit()
                except Exception:
                    pass
        finally:
            if profile_id:
                self._cleanup_profile_session(item_id, profile_id)

    def _save_cache_list(self, key: str, data: list) -> None:
        try:
            payload = json.dumps(data or [], ensure_ascii=False)
            with self._cache_db_lock:
                conn = self._open_cache_db()
                if conn:
                    try:
                        conn.execute(
                            "INSERT OR REPLACE INTO cache (key, data, updated_at) VALUES (?, ?, ?)",
                            (key, payload, time.time()),
                        )
                        conn.commit()
                    finally:
                        conn.close()
        except Exception:
            pass

    def _on_close(self) -> None:
        self._save_accounts_cache()
        self._save_profile_accounts_cache()
        self._save_fb_accounts_cache()
        self._save_fb_profile_accounts_cache()
        # Print error summary before closing
        self.error_logger.print_error_summary()
        try:
            self.root.destroy()
        except Exception:
            pass

    def import_accounts(self) -> None:
        if self._is_profile_tab():
            self.import_profile_accounts()
            return
        if self._is_fb_tab():
            self.import_fb_accounts()
            return
        if self._is_fb_profile_tab():
            self.import_fb_profile_accounts()
            return
        path = filedialog.askopenfilename(
            title="Import accounts",
            filetypes=[
                ("Excel", "*.xlsx"),
                ("Text/CSV", "*.txt;*.csv;*.tsv"),
                ("All files", "*.*"),
            ],
        )
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        rows = []
        try:
            if ext == ".xlsx":
                try:
                    import openpyxl  # type: ignore
                except Exception:
                    messagebox.showerror("Import", "Can phai cai dat openpyxl de doc file .xlsx")
                    return
                wb = openpyxl.load_workbook(path, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if not row:
                        continue
                    rows.append([str(c).strip() if c is not None else "" for c in row])
            else:
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()
                delimiter = "\t" if "\t" in content else ","
                reader = csv.reader(content.splitlines(), delimiter=delimiter)
                rows = [r for r in reader if r]
        except Exception as e:
            messagebox.showerror("Import", f"Loi doc file: {e}")
            return

        existing_by_uid = {str(a.get("uid") or "").strip(): a for a in self.accounts}
        new_accounts = []
        for row in rows:
            if len(row) < 4:
                continue
            uid = (row[0] or "").strip()
            pwd = (row[1] or "").strip()
            proxy = (row[2] or "").strip()
            yt = (row[3] or "").strip()
            if uid.lower() in ("email", "uid") and pwd.lower() in ("pass", "password") and proxy.lower() in ("proxy", "raw_proxy"):
                continue
            if not uid:
                continue
            acc = {"uid": uid, "pass": pwd, "proxy": proxy, "youtube": yt}
            old = existing_by_uid.get(uid)
            if old:
                acc["followers"] = old.get("followers")
                acc["posts"] = old.get("posts")
                acc["profile_url"] = old.get("profile_url", "")
                acc["profile_id"] = old.get("profile_id", "")
                acc["status"] = old.get("status", acc.get("status"))
            new_accounts.append(acc)

        if not new_accounts:
            messagebox.showinfo("Import", "Khong tim thay dong du lieu hop le.")
            return

        self.accounts = new_accounts
        self._load_rows()
        self._save_accounts_cache()
        self._log(f"[IMPORT] Loaded {len(new_accounts)} accounts")

    def export_accounts_excel(self) -> None:
        try:
            import openpyxl  # type: ignore
        except Exception:
            messagebox.showerror("Export", "Can phai cai dat openpyxl de xuat file .xlsx")
            return

        path = filedialog.asksaveasfilename(
            title="Export accounts to Excel",
            defaultextension=".xlsx",
            initialfile="accounts_export.xlsx",
            filetypes=[("Excel", "*.xlsx")],
        )
        if not path:
            return

        def _to_text(v):
            if v is None:
                return ""
            return str(v)

        yt_headers = ["STT", "EMAIL", "PASS", "POSTS", "FOLLOWERS", "PROXY", "YOUTUBE", "PROFILE_URL", "PROFILE_ID"]
        fb_headers = ["STT", "EMAIL", "PASS", "POSTS", "FOLLOWERS", "PROXY", "FACEBOOK", "PROFILE_URL", "PROFILE_ID"]

        try:
            wb = openpyxl.Workbook()
            ws_yt = wb.active
            ws_yt.title = "YOUTUBE"
            ws_yt.append(yt_headers)
            yt_items = list(self.tree.get_children()) if hasattr(self, "tree") else []
            for idx, iid in enumerate(yt_items, start=1):
                ws_yt.append(
                    [
                        idx,
                        _to_text(self.tree.set(iid, "email")),
                        _to_text(self.tree.set(iid, "pass")),
                        _to_text(self.tree.set(iid, "posts")),
                        _to_text(self.tree.set(iid, "followers")),
                        _to_text(self.tree.set(iid, "proxy")),
                        _to_text(self.tree.set(iid, "youtube")),
                        _to_text(self.tree.set(iid, "profile_url")),
                        _to_text(self.tree.set(iid, "profile_id")),
                    ]
                )

            ws_fb = wb.create_sheet("FACEBOOK")
            ws_fb.append(fb_headers)
            fb_items = list(self.fb_tree.get_children()) if hasattr(self, "fb_tree") else []
            for idx, iid in enumerate(fb_items, start=1):
                ws_fb.append(
                    [
                        idx,
                        _to_text(self.fb_tree.set(iid, "email")),
                        _to_text(self.fb_tree.set(iid, "pass")),
                        _to_text(self.fb_tree.set(iid, "posts")),
                        _to_text(self.fb_tree.set(iid, "followers")),
                        _to_text(self.fb_tree.set(iid, "proxy")),
                        _to_text(self.fb_tree.set(iid, "facebook")),
                        _to_text(self.fb_tree.set(iid, "profile_url")),
                        _to_text(self.fb_tree.set(iid, "profile_id")),
                    ]
                )

            wb.save(path)
            self._log(
                f"[EXPORT] Saved {len(yt_items)} YTB + {len(fb_items)} FB accounts -> {path}"
            )
            messagebox.showinfo("Export", f"Da xuat xong file:\n{path}")
        except Exception as e:
            messagebox.showerror("Export", f"Loi xuat file: {e}")

    def import_proxy_list(self) -> None:
        path = self._extra_proxy_file
        try:
            if not os.path.exists(path):
                os.makedirs(os.path.dirname(path), exist_ok=True)
                with open(path, "w", encoding="utf-8") as f:
                    f.write("")
        except Exception as e:
            messagebox.showerror("Import Proxy", f"Loi tao file proxy: {e}")
            return
        try:
            os.startfile(path)
        except Exception:
            pass
        try:
            with open(path, "r", encoding="utf-8") as f:
                content = f.read()
        except Exception as e:
            messagebox.showerror("Import Proxy", f"Loi doc file: {e}")
            return
        proxies = []
        for raw in content.splitlines():
            line = (raw or "").strip()
            if not line:
                continue
            # Accept first field if CSV/TSV
            for sep in (",", "\t", ";", " "):
                if sep in line:
                    line = line.split(sep, 1)[0].strip()
                    break
            if line:
                proxies.append(line)
        if not proxies:
            messagebox.showinfo(
                "Import Proxy",
                "Da tao/mo file proxy. Hay dan proxy vao file roi bam IMPORT PROXY lai.",
            )
            return
        with self._extra_proxy_lock:
            self._extra_proxies = proxies
            self._extra_proxy_idx = 0
        self._save_extra_proxy_list()
        self._log(f"[PROXY] Loaded {len(proxies)} proxies")

    def import_profile_accounts(self) -> None:
        path = filedialog.askopenfilename(
            title="Import profile accounts",
            filetypes=[
                ("Excel", "*.xlsx"),
                ("Text/CSV", "*.txt;*.csv;*.tsv"),
                ("All files", "*.*"),
            ],
        )
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        rows = []
        try:
            if ext == ".xlsx":
                try:
                    import openpyxl  # type: ignore
                except Exception:
                    messagebox.showerror("Import", "Can phai cai dat openpyxl de doc file .xlsx")
                    return
                wb = openpyxl.load_workbook(path, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if not row:
                        continue
                    rows.append([str(c).strip() if c is not None else "" for c in row])
            else:
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()
                delimiter = "\t" if "\t" in content else ","
                reader = csv.reader(content.splitlines(), delimiter=delimiter)
                rows = [r for r in reader if r]
        except Exception as e:
            messagebox.showerror("Import", f"Loi doc file: {e}")
            return

        new_accounts = []
        for row in rows:
            if len(row) < 4:
                continue
            uid = (row[0] or "").strip()
            pwd = (row[1] or "").strip()
            proxy = (row[2] or "").strip()
            yt = (row[3] or "").strip()
            if uid.lower() in ("email", "uid") and pwd.lower() in ("pass", "password") and proxy.lower() in ("proxy", "raw_proxy"):
                continue
            if not uid:
                continue
            new_accounts.append({"uid": uid, "pass": pwd, "proxy": proxy, "youtube": yt})

        if not new_accounts:
            messagebox.showinfo("Import", "Khong tim thay dong du lieu hop le.")
            return

        self.profile_accounts = new_accounts
        self._load_profile_rows()
        self._save_profile_accounts_cache()
        self._log(f"[IMPORT PROFILE] Loaded {len(new_accounts)} accounts")

    def import_fb_accounts(self) -> None:
        path = filedialog.askopenfilename(
            title="Import FB accounts",
            filetypes=[
                ("Excel", "*.xlsx"),
                ("Text/CSV", "*.txt;*.csv;*.tsv"),
                ("All files", "*.*"),
            ],
        )
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        rows = []
        try:
            if ext == ".xlsx":
                try:
                    import openpyxl  # type: ignore
                except Exception:
                    messagebox.showerror("Import", "Can phai cai dat openpyxl de doc file .xlsx")
                    return
                wb = openpyxl.load_workbook(path, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if not row:
                        continue
                    rows.append([str(c).strip() if c is not None else "" for c in row])
            else:
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()
                delimiter = "\t" if "\t" in content else ","
                reader = csv.reader(content.splitlines(), delimiter=delimiter)
                rows = [r for r in reader if r]
        except Exception as e:
            messagebox.showerror("Import", f"Loi doc file: {e}")
            return

        existing_by_uid = {str(a.get("uid") or "").strip(): a for a in self.fb_accounts}
        new_accounts = []
        for row in rows:
            if len(row) < 4:
                continue
            uid = (row[0] or "").strip()
            pwd = (row[1] or "").strip()
            proxy = (row[2] or "").strip()
            fb_link = (row[3] or "").strip()
            if uid.lower() in ("email", "uid") and pwd.lower() in ("pass", "password") and proxy.lower() in ("proxy", "raw_proxy"):
                continue
            if not uid:
                continue
            acc = {"uid": uid, "pass": pwd, "proxy": proxy, "facebook": fb_link}
            old = existing_by_uid.get(uid)
            if old:
                acc["followers"] = old.get("followers")
                acc["posts"] = old.get("posts")
                acc["profile_url"] = old.get("profile_url", "")
                acc["profile_id"] = old.get("profile_id", "")
                acc["status"] = old.get("status", acc.get("status"))
            new_accounts.append(acc)

        if not new_accounts:
            messagebox.showinfo("Import", "Khong tim thay dong du lieu hop le.")
            return

        self.fb_accounts = new_accounts
        self._load_fb_rows()
        self._save_fb_accounts_cache()
        self._log(f"[IMPORT FB] Loaded {len(new_accounts)} accounts")

    def import_fb_profile_accounts(self) -> None:
        path = filedialog.askopenfilename(
            title="Import FB profile accounts",
            filetypes=[
                ("Excel", "*.xlsx"),
                ("Text/CSV", "*.txt;*.csv;*.tsv"),
                ("All files", "*.*"),
            ],
        )
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        rows = []
        try:
            if ext == ".xlsx":
                try:
                    import openpyxl  # type: ignore
                except Exception:
                    messagebox.showerror("Import", "Can phai cai dat openpyxl de doc file .xlsx")
                    return
                wb = openpyxl.load_workbook(path, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    if not row:
                        continue
                    rows.append([str(c).strip() if c is not None else "" for c in row])
            else:
                with open(path, "r", encoding="utf-8") as f:
                    content = f.read()
                delimiter = "\t" if "\t" in content else ","
                reader = csv.reader(content.splitlines(), delimiter=delimiter)
                rows = [r for r in reader if r]
        except Exception as e:
            messagebox.showerror("Import", f"Loi doc file: {e}")
            return

        new_accounts = []
        for row in rows:
            if len(row) < 4:
                continue
            uid = (row[0] or "").strip()
            pwd = (row[1] or "").strip()
            proxy = (row[2] or "").strip()
            fb_link = (row[3] or "").strip()
            if uid.lower() in ("email", "uid") and pwd.lower() in ("pass", "password") and proxy.lower() in ("proxy", "raw_proxy"):
                continue
            if not uid:
                continue
            new_accounts.append({"uid": uid, "pass": pwd, "proxy": proxy, "facebook": fb_link})

        if not new_accounts:
            messagebox.showinfo("Import", "Khong tim thay dong du lieu hop le.")
            return

        self.fb_profile_accounts = new_accounts
        self._load_fb_profile_rows()
        self._save_fb_profile_accounts_cache()
        self._log(f"[IMPORT FB PROFILE] Loaded {len(new_accounts)} accounts")

    def _is_profile_tab(self) -> bool:
        try:
            return self.notebook.nametowidget(self.notebook.select()) == self.tab_profile
        except Exception:
            return False

    def _is_all_tab(self) -> bool:
        try:
            return self.notebook.nametowidget(self.notebook.select()) == self.tab_all
        except Exception:
            return False

    def _is_fb_tab(self) -> bool:
        try:
            return self.notebook.nametowidget(self.notebook.select()) == self.tab_fb
        except Exception:
            return False

    def _is_fb_profile_tab(self) -> bool:
        try:
            return self.notebook.nametowidget(self.notebook.select()) == self.tab_fb_profile
        except Exception:
            return False

    def _transfer_active_profile_accounts(self) -> None:
        if self._is_fb_profile_tab():
            tree = self.fb_profile_tree
            source_accounts = self.fb_profile_accounts
            target_accounts = self.fb_accounts
            link_field = "facebook"
            target_tab = self.tab_fb
            label = "FB"
        else:
            tree = self.profile_tree
            source_accounts = self.profile_accounts
            target_accounts = self.accounts
            link_field = "youtube"
            target_tab = self.tab_upload
            label = "YTB"

        checked_emails = self._get_checked_email_set(tree)
        if not checked_emails:
            messagebox.showinfo("Chuyen du lieu", "Hay tick cac dong profile can chuyen.")
            return

        moved = self._transfer_accounts_to_upload(
            source_accounts=source_accounts,
            target_accounts=target_accounts,
            checked_emails=checked_emails,
            link_field=link_field,
            reload_rows=True,
        )
        if moved <= 0:
            messagebox.showinfo("Chuyen du lieu", f"Khong co du lieu {label} nao can cap nhat sang Upload.")
            return

        self._set_sidebar_active("upvideo")
        self._select_tab(target_tab)
        self._log(f"[TRANSFER {label}] moved {moved} account(s) from Profile -> Upload")
        messagebox.showinfo("Chuyen du lieu", f"Da chuyen {moved} tai khoan {label} sang man Upload.")

    def start_profile_jobs(self) -> None:
        if self._profile_batch_running or self.executor is not None:
            return
        self._force_close_all_profiles()
        self._reset_all_statuses()
        if self._repeat_after_id:
            try:
                self.root.after_cancel(self._repeat_after_id)
            except Exception:
                pass
            self._repeat_after_id = None
        try:
            max_threads = int(self.entry_threads.get())
            if max_threads <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Error", "Threads must be > 0")
            return

        self._fixed_threads = max_threads
        self.stop_event.clear()
        self._profile_retry_round = 0
        pending = self._resume_pending.get("profile") or set()
        if pending:
            if self._prompt_resume("profile", len(pending)):
                checked_emails = pending
            else:
                self._resume_pending["profile"] = set()
                checked_emails = self._get_checked_email_set(self.profile_tree)
        else:
            checked_emails = self._get_checked_email_set(self.profile_tree)
        if not checked_emails:
            messagebox.showinfo("Thong bao", "Khong co profile nao duoc tick.")
            return
        self._set_run_total("profile", len(checked_emails))
        self._profile_batch_running = True

        def _run_batch(batch_items: list) -> None:
            if not batch_items or self.stop_event.is_set():
                return

            try:
                screen_w = self.root.winfo_screenwidth()
                screen_h = self.root.winfo_screenheight()
            except Exception:
                screen_w, screen_h = 1920, 1080

            gap = 6
            taskbar_h = 40
            usable_w = screen_w - (gap * 2)
            usable_h = (screen_h - taskbar_h) - (gap * 2)

            retry_round = 0
            current_items = list(batch_items)
            while current_items and not self.stop_event.is_set():
                active_count = len(current_items)
                cols = min(5, active_count)
                rows_layout = min(2, max(1, math.ceil(active_count / cols)))
                win_w = int((usable_w - gap * (cols - 1)) / cols)
                win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
                win_w = max(150, min(280, win_w))
                win_h = max(420, min(600, win_h))

                with self.profile_failed_lock:
                    self.profile_failed_accounts = []

                self.executor = ThreadPoolExecutor(max_workers=max_threads)
                self.profile_semaphore = threading.BoundedSemaphore(max_threads)
                futures = []
                max_slots = cols * rows_layout
                for idx, (item_id, acc) in enumerate(current_items):
                    if self.stop_event.is_set():
                        break
                    pos = idx % max_slots
                    col = pos % cols
                    row = pos // cols
                    x = gap + col * (win_w + gap)
                    y = gap + row * (win_h + gap)
                    win_pos = f"{x},{y}"
                    win_size = f"{win_w},{win_h}"
                    futures.append(self.executor.submit(self._profile_open_worker, item_id, acc, win_pos, win_size))
                    if PROFILE_BATCH_STAGGER_SEC > 0:
                        time.sleep(PROFILE_BATCH_STAGGER_SEC)

                for f in as_completed(futures):
                    try:
                        f.result()
                    except Exception:
                        pass

                try:
                    self.executor.shutdown(wait=True)
                except Exception:
                    pass
                self.executor = None

                with self.profile_failed_lock:
                    failed_list = self.profile_failed_accounts.copy()
                    self.profile_failed_accounts = []

                if not failed_list or self.stop_event.is_set():
                    break

                retry_round += 1
                if retry_round > self._max_retry_rounds:
                    self._log(
                        f"[PROFILE RETRY] Stop retry after {self._max_retry_rounds} rounds (remaining: {len(failed_list)})"
                    )
                    break
                self._log(
                    f"[PROFILE RETRY] Retrying {len(failed_list)} failed accounts (round {retry_round}/{self._max_retry_rounds})..."
                )
                current_items = failed_list

            # keep profile failed log

        def _batch_runner():
            try:
                items = []
                email_to_iid = self._map_email_to_item_id(self.profile_tree)
                for acc in self.profile_accounts:
                    email = (acc.get("uid") or "").strip()
                    if email in checked_emails:
                        item_id = email_to_iid.get(email)
                        if item_id:
                            items.append((item_id, acc))

                if not items:
                    return

                total = len(items)
                batch_size = max(1, int(PROFILE_BATCH_SIZE))
                for start_idx in range(0, total, batch_size):
                    if self.stop_event.is_set():
                        break
                    batch = items[start_idx : start_idx + batch_size]
                    batch_no = (start_idx // batch_size) + 1
                    total_batches = math.ceil(total / batch_size)
                    self._log(f"[PROFILE BATCH] Start batch {batch_no}/{total_batches} ({len(batch)} profiles)")
                    _run_batch(batch)
                    if self.stop_event.is_set():
                        break
                self._log("[PROFILE BATCH] Done")
            finally:
                self._profile_batch_running = False
                self._reset_run("profile")

        threading.Thread(target=_batch_runner, daemon=True).start()

    def start_all_jobs_mixed(self, snapshot: dict | None = None) -> None:
        if self.executor is not None:
            return
        ordered_rows = self._build_mixed_ordered_rows()
        if not ordered_rows and snapshot is not None:
            ordered_rows = self._build_mixed_ordered_rows(list(snapshot.get("ordered_rows") or []))
        if not ordered_rows:
            if snapshot is None:
                messagebox.showinfo("Thong bao", "Khong co profile nao duoc tick.")
            else:
                self._log("[ALL] Skip repeat cycle: khong co du lieu map moi tu YTB/FB.")
            self._repeat_cycle_pending = False
            self._from_all_tab = False
            return

        ytb_emails = [email for social, email in ordered_rows if social == "YTB"]
        fb_emails = [email for social, email in ordered_rows if social == "FB"]
        if not ytb_emails and not fb_emails:
            self._repeat_cycle_pending = False
            self._from_all_tab = False
            return

        if self._repeat_cycle_pending:
            self._repeat_cycle_pending = False
            self._increment_cycle()
        else:
            self._cycle_count = 1
            self._set_cycle_label()
            self._reset_upload_cycle_stats()
        self._runtime_reset()
        self._runtime_start()

        self._from_all_tab = True
        self._set_checked_by_email(self.tree, set(ytb_emails))
        self._set_checked_by_email(self.fb_tree, set(fb_emails))

        self._force_close_all_profiles()
        self._reset_all_statuses()
        # Keep logs (do not clear on start)

        if snapshot is not None and snapshot.get("max_threads"):
            max_threads = int(snapshot.get("max_threads"))
        else:
            try:
                max_threads = int(self.entry_threads.get())
                if max_threads <= 0:
                    raise ValueError
            except Exception:
                messagebox.showerror("Loi", "So luong phai > 0")
                return

        if snapshot is not None and snapshot.get("max_videos"):
            max_videos = int(snapshot.get("max_videos"))
        else:
            try:
                max_videos = int(self.entry_videos.get())
                if max_videos <= 0:
                    raise ValueError
            except Exception:
                messagebox.showerror("Loi", "Videos phai > 0")
                return

        self._fixed_threads = max_threads
        self.stop_event.clear()
        self._retry_round = 0
        self._all_retry_round = 0
        with self.failed_accounts_lock:
            self.failed_accounts = []

        self.executor = ThreadPoolExecutor(max_workers=max_threads)
        self.login_semaphore = threading.BoundedSemaphore(max_threads)
        self.upload_retry_semaphore = threading.BoundedSemaphore(max_threads)

        self._set_run_total("upload", len(ytb_emails))
        self._set_run_total("fb", len(fb_emails))

        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080

        gap = 6
        taskbar_h = 40
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(ytb_emails) + len(fb_emails)
        cols = min(5, max(1, active_count))
        rows_layout = min(2, max(1, math.ceil(active_count / cols)))
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))

        futures = []
        slot_idx = 0
        max_slots = cols * rows_layout

        ytb_by_email = {str(a.get("uid") or "").strip(): a for a in self.accounts}
        fb_by_email = {str(a.get("uid") or "").strip(): a for a in self.fb_accounts}
        ytb_email_to_iid = self._map_email_to_item_id(self.tree)
        fb_email_to_iid = self._map_email_to_item_id(self.fb_tree)

        for social, email in ordered_rows:
            if social == "YTB":
                acc = ytb_by_email.get(email)
                item_id = ytb_email_to_iid.get(email)
                if not acc or not item_id:
                    continue
                self._bind_item_email(item_id, acc.get("uid", ""))
                pos = slot_idx % max_slots
                col = pos % cols
                row = pos // cols
                x = gap + col * (win_w + gap)
                y = gap + row * (win_h + gap)
                win_pos = f"{x},{y}"
                win_size = f"{win_w},{win_h}"
                futures.append(self.executor.submit(self._worker_one, item_id, acc, win_pos, win_size, max_videos))
                slot_idx += 1
            elif social == "FB":
                acc = fb_by_email.get(email)
                item_id = fb_email_to_iid.get(email)
                if not acc or not item_id:
                    continue
                self._bind_item_email(item_id, acc.get("uid", ""))
                pos = slot_idx % max_slots
                col = pos % cols
                row = pos // cols
                x = gap + col * (win_w + gap)
                y = gap + row * (win_h + gap)
                win_pos = f"{x},{y}"
                win_size = f"{win_w},{win_h}"
                futures.append(self.executor.submit(self._fb_worker_one, item_id, acc, win_pos, win_size, max_videos))
                slot_idx += 1

        self._repeat_enabled = bool(self.repeat_var.get())
        try:
            delay_min = float(self.entry_repeat_delay.get())
            if delay_min < 0:
                delay_min = 0
        except Exception:
            delay_min = 0
        self._repeat_delay_sec = delay_min * 60.0
        self._all_repeat_snapshot = {
            "ordered_rows": list(ordered_rows),
            "max_threads": max_threads,
            "max_videos": max_videos,
        }

        def _waiter():
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception:
                    pass
            self.executor = None
            with self.failed_accounts_lock:
                failed_list = self.failed_accounts.copy()
                self.failed_accounts = []
            if failed_list and not self.stop_event.is_set():
                if self._all_retry_round < self._upload_retry_rounds:
                    self._all_retry_round += 1
                    self._log(
                        f"[ALL RETRY] Retrying {len(failed_list)} failed accounts (round {self._all_retry_round}/{self._upload_retry_rounds})..."
                    )
                    self.root.after(
                        1000,
                        lambda fl=failed_list: self._retry_failed_all_mixed(fl, max_threads, max_videos),
                    )
                    return
                self._log(f"[ALL RETRY] Stop retry after {self._upload_retry_rounds} rounds (remaining: {len(failed_list)})")
            # keep failed log
            self._reset_run("upload")
            self._reset_run("fb")

            if self._repeat_enabled and not self.stop_event.is_set():
                delay_sec = max(0.0, float(self._repeat_delay_sec or 0.0))
                delay_ms = int(delay_sec * 1000)
                self._start_next_cycle_countdown(int(delay_sec))

                def _repeat_start():
                    if self.stop_event.is_set():
                        return
                    self._repeat_after_id = None
                    self._repeat_cycle_pending = True
                    self.start_all_jobs_mixed(self._all_repeat_snapshot)

                self._repeat_after_id = self.root.after(max(0, delay_ms), _repeat_start)
            self._from_all_tab = False

        threading.Thread(target=_waiter, daemon=True).start()

    def _retry_failed_all_mixed(self, failed_accounts: list, max_threads: int, max_videos: int) -> None:
        if self.stop_event.is_set():
            return
        self.executor = ThreadPoolExecutor(max_workers=max_threads)
        futures = []
        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080

        gap = 6
        taskbar_h = 40
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(failed_accounts)
        cols = min(5, max(1, active_count))
        rows_layout = min(2, max(1, math.ceil(active_count / cols)))
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))

        ytb_emails = {str(a.get("uid") or "").strip() for a in self.accounts}
        fb_emails = {str(a.get("uid") or "").strip() for a in self.fb_accounts}
        order_map = {}
        try:
            for idx, iid in enumerate(self.all_tree.get_children()):
                social = (self.all_tree.set(iid, "social") or "").strip().upper()
                email = (self.all_tree.set(iid, "email") or "").strip()
                order_map[(social, email)] = idx
        except Exception:
            pass
        slot_idx = 0
        max_slots = cols * rows_layout
        ordered_failed = []
        for item_id, acc in failed_accounts:
            email = (acc.get("uid") or "").strip()
            social = "FB" if email in fb_emails else "YTB"
            ordered_failed.append((order_map.get((social, email), 10**9), item_id, acc))
        ordered_failed.sort(key=lambda x: x[0])
        for _ord, item_id, acc in ordered_failed:
            if self.stop_event.is_set():
                break
            email = (acc.get("uid") or "").strip()
            pos = slot_idx % max_slots
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            win_pos = f"{x},{y}"
            win_size = f"{win_w},{win_h}"
            if email in fb_emails:
                futures.append(self.executor.submit(self._fb_worker_one, item_id, acc, win_pos, win_size, max_videos))
            else:
                futures.append(self.executor.submit(self._worker_one, item_id, acc, win_pos, win_size, max_videos))
            slot_idx += 1

        def _retry_waiter():
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception:
                    pass
            self.executor = None
            with self.failed_accounts_lock:
                failed_list = self.failed_accounts.copy()
                self.failed_accounts = []
            if failed_list and not self.stop_event.is_set():
                if self._all_retry_round < self._upload_retry_rounds:
                    self._all_retry_round += 1
                    self._log(
                        f"[ALL RETRY] Retrying {len(failed_list)} failed accounts (round {self._all_retry_round}/{self._upload_retry_rounds})..."
                    )
                    self.root.after(
                        1000,
                        lambda fl=failed_list: self._retry_failed_all_mixed(fl, max_threads, max_videos),
                    )
                    return
                self._log(f"[ALL RETRY] Stop retry after {self._upload_retry_rounds} rounds (remaining: {len(failed_list)})")
            # keep failed log
            self._reset_run("upload")
            self._reset_run("fb")

            if self._repeat_enabled and not self.stop_event.is_set():
                delay_ms = 5 * 60 * 1000
                self._start_next_cycle_countdown(300)
                try:
                    self._log("[REPEAT] Running GPM cleanup during wait...")
                    self.root.after(0, self._clear_all_gpm_profiles)
                except Exception:
                    pass

                def _repeat_start():
                    if self.stop_event.is_set():
                        return
                    self._repeat_after_id = None
                    self._repeat_cycle_pending = True
                    self.start_all_jobs_mixed(self._all_repeat_snapshot)

                self._repeat_after_id = self.root.after(delay_ms, _repeat_start)
            self._from_all_tab = False

        threading.Thread(target=_retry_waiter, daemon=True).start()


    def start_jobs(self) -> None:
        if not self._require_license_or_warn():
            return
        self._set_busy(True)
        if self._is_all_tab():
            self.start_all_jobs_mixed()
            return
        else:
            self._from_all_tab = False
            self._all_pending_fb_emails = set()
        if self._is_profile_tab():
            self.start_profile_jobs()
            return
        if self._is_fb_tab():
            self.start_fb_jobs()
            return
        if self._is_fb_profile_tab():
            self.start_fb_profile_jobs()
            return
        if self.executor is not None:
            return
        if not self._force_upload_only:
            upload_checked = self._get_checked_email_set(self.tree)
            fb_checked = self._get_checked_email_set(self.fb_tree)
            if upload_checked:
                self._run_upload_after_fb = False
            elif fb_checked:
                # no upload checked, run FB directly
                self._run_upload_after_fb = False
                self.start_fb_jobs()
                return
        self._force_upload_only = False
        self._force_close_all_profiles()
        self._reset_all_statuses()
        if self._repeat_after_id:
            try:
                self.root.after_cancel(self._repeat_after_id)
            except Exception:
                pass
            self._repeat_after_id = None
        try:
            max_threads = int(self.entry_threads.get())
            if max_threads <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Error", "Threads must be > 0")
            return

        self._fixed_threads = max_threads
        self.stop_event.clear()
        self._clear_status_tags()
        if self._repeat_cycle_pending:
            self._repeat_cycle_pending = False
            self._increment_cycle()
        else:
            self._cycle_count = 1
            self._set_cycle_label()
            self._reset_upload_cycle_stats()
        self._runtime_reset()
        self._runtime_start()
        self._retry_round = 0
        # Use exact number of threads (no extra retry threads)
        self.executor = ThreadPoolExecutor(max_workers=max_threads)
        self.login_semaphore = threading.BoundedSemaphore(max_threads)
        self.upload_retry_semaphore = threading.BoundedSemaphore(max_threads)

        self._reset_upload_tree_order()
        pending = self._resume_pending.get("upload") or set()
        if pending:
            if self._prompt_resume("upload", len(pending)):
                checked_emails = pending
            else:
                self._resume_pending["upload"] = set()
                checked_emails = self._get_checked_email_set(self.tree)
        else:
            checked_emails = self._get_checked_email_set(self.tree)
        if not checked_emails:
            messagebox.showinfo("Thong bao", "Khong co profile nao duoc tick.")
            return
        self._set_run_total("upload", len(checked_emails))

        # Optimized layout: 5 profiles per row, full screen width - COMPACT MODE
        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080
        
        gap = 6  # Tight spacing between windows
        taskbar_h = 40  # Reserve space for taskbar
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(checked_emails)
        
        # Fixed layout: 5 columns, max 2 rows (overflow cycles back to row 1)
        cols = min(5, active_count)  # Max 5 per row
        rows_layout = min(2, max(1, math.ceil(active_count / cols)))
        
        # Calculate window sizes
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        
        # Compact windows: narrow width, optimal height
        win_w = max(150, min(280, win_w))  # Even more compact: 150-280px
        win_h = max(420, min(600, win_h))  # Height range: 420-600px

        futures = []

        try:
            max_videos = int(self.entry_videos.get())
            if max_videos <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Loi", "Videos phai > 0")
            return

        self._repeat_enabled = bool(self.repeat_var.get())
        try:
            delay_min = float(self.entry_repeat_delay.get())
            if delay_min < 0:
                delay_min = 0
        except Exception:
            delay_min = 0
        self._repeat_delay_sec = delay_min * 60.0

        slot_idx = 0
        max_slots = cols * rows_layout
        email_to_iid = self._map_email_to_item_id(self.tree)
        for acc in self.accounts:
            email = (acc.get("uid") or "").strip()
            if email not in checked_emails:
                continue
            item_id = email_to_iid.get(email)
            if not item_id:
                continue
            self._bind_item_email(item_id, acc.get("uid", ""))
            pos = slot_idx % max_slots
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            win_pos = f"{x},{y}"
            win_size = f"{win_w},{win_h}"
            futures.append(self.executor.submit(self._worker_one, item_id, acc, win_pos, win_size, max_videos))
            slot_idx += 1

        def _waiter():
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception:
                    pass
            
            # Check if there are failed accounts to retry
            with self.failed_accounts_lock:
                failed_list = self.failed_accounts.copy()
                self.failed_accounts = []
            try:
                failed_list.extend(self._collect_transient_failures())
            except Exception:
                pass
            
            self.executor = None
            
            # If failed accounts exist, retry them immediately (only during run)
            if failed_list and not self.stop_event.is_set():
                if self._retry_round < self._upload_retry_rounds:
                    self._retry_round += 1
                    self._log(
                        f"[RETRY] Retrying {len(failed_list)} failed accounts (round {self._retry_round}/{self._upload_retry_rounds})..."
                    )
                    # Don't use the last loop's win_pos/win_size; let _retry_failed_accounts calculate its own layout
                    self.root.after(1000, lambda fl=failed_list: self._retry_failed_accounts(fl, max_threads, max_videos))
                    return
                self._log(f"[RETRY] Stop retry after {self._upload_retry_rounds} rounds (remaining: {len(failed_list)})")
            # keep failed log

            if self._from_all_tab and self._all_pending_fb_emails and not self.stop_event.is_set():
                pending_fb = set(self._all_pending_fb_emails)
                self._all_pending_fb_emails = set()
                self._set_checked_by_email(self.fb_tree, pending_fb)
                self.start_fb_jobs()
                return
            # Do not auto-switch to FB from YTB
            if self._from_all_tab:
                self._from_all_tab = False

            if self._repeat_enabled and not self.stop_event.is_set():
                delay_ms = 5 * 60 * 1000
                self._start_next_cycle_countdown(300)
                try:
                    self._log("[REPEAT] Running GPM cleanup during wait...")
                    self.root.after(0, self._clear_all_gpm_profiles)
                except Exception:
                    pass

                def _repeat_start():
                    if self.stop_event.is_set():
                        return
                    self._repeat_after_id = None
                    self._repeat_cycle_pending = True
                    self.start_jobs()

                self._repeat_after_id = self.root.after(delay_ms, _repeat_start)
            self._reset_run("upload")

        threading.Thread(target=_waiter, daemon=True).start()

    def _do_repeat_cycle(self) -> None:
        """Repeat cycle for smooth continuous upload without recalculating layout"""
        if self.executor is not None:
            return
        self._runtime_start()
        self._increment_cycle()
        self._reset_all_statuses()
        
        try:
            max_threads = int(self._fixed_threads or int(self.entry_threads.get()))
            if max_threads <= 0:
                raise ValueError
        except Exception:
            return

        # Get checked items (should still be checked from previous run)
        checked_emails = self._get_checked_email_set(self.tree)
        if not checked_emails:
            return
        self._set_run_total("upload", len(checked_emails))

        # Clear executor state
        self.stop_event.clear()
        self._clear_status_tags()
        self._retry_round = 0
        with self.failed_accounts_lock:
            self.failed_accounts = []

        # Use same thread count
        self.executor = ThreadPoolExecutor(max_workers=max_threads)
        self.login_semaphore = threading.BoundedSemaphore(max_threads)
        self.upload_retry_semaphore = threading.BoundedSemaphore(max_threads)

        # Get max videos
        try:
            max_videos = int(self.entry_videos.get())
            if max_videos <= 0:
                raise ValueError
        except Exception:
            max_videos = 1

        # Reuse layout from previous run (or recalculate if needed)
        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080

        gap = 6
        taskbar_h = 40
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(checked_emails)

        cols = min(5, active_count)
        rows_layout = min(2, max(1, math.ceil(active_count / cols)))

        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)

        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))

        futures = []
        slot_idx = 0
        max_slots = cols * rows_layout

        self._log(f"\n[REPEAT] Starting repeat cycle... (checked: {len(checked_emails)} accounts)")

        email_to_iid = self._map_email_to_item_id(self.tree)
        for acc in self.accounts:
            email = (acc.get("uid") or "").strip()
            if email not in checked_emails:
                continue
            item_id = email_to_iid.get(email)
            if not item_id:
                continue
            self._bind_item_email(item_id, acc.get("uid", ""))
            
            # Reset status to prepare for new cycle
            self.tree.set(item_id, "status", "WAIT")
            
            pos = slot_idx % max_slots
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            win_pos = f"{x},{y}"
            win_size = f"{win_w},{win_h}"
            futures.append(self.executor.submit(self._worker_one, item_id, acc, win_pos, win_size, max_videos))
            slot_idx += 1

        def _repeat_waiter():
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception:
                    pass

            with self.failed_accounts_lock:
                failed_list = self.failed_accounts.copy()
                self.failed_accounts = []
            try:
                failed_list.extend(self._collect_transient_failures())
            except Exception:
                pass

            self.executor = None

            # Retry failed accounts
            if failed_list and not self.stop_event.is_set():
                if self._retry_round < self._upload_retry_rounds:
                    self._retry_round += 1
                    self._log(
                        f"[RETRY] Retrying {len(failed_list)} failed accounts (round {self._retry_round}/{self._upload_retry_rounds})..."
                    )
                    self.root.after(1000, lambda fl=failed_list: self._retry_failed_accounts(fl, max_threads, max_videos))
                    return
                self._log(f"[RETRY] Stop retry after {self._upload_retry_rounds} rounds (remaining: {len(failed_list)})")

            # keep failed log

            # Schedule next repeat cycle
            if self._repeat_enabled and not self.stop_event.is_set():
                delay_ms = 5 * 60 * 1000
                self._start_next_cycle_countdown(300)

                def _repeat_again():
                    if self.stop_event.is_set():
                        return
                    self._repeat_after_id = None
                    self._repeat_cycle_pending = True
                    self.start_jobs()

                self._repeat_after_id = self.root.after(delay_ms, _repeat_again)
                self._log("[REPEAT] Next cycle in 300 seconds...")
                try:
                    self._log("[REPEAT] Running GPM cleanup during wait...")
                    self.root.after(0, self._clear_all_gpm_profiles)
                except Exception:
                    pass
            self._reset_run("upload")

        threading.Thread(target=_repeat_waiter, daemon=True).start()

    def start_scan(self) -> None:
        if not self._require_license_or_warn():
            return
        if self._is_fb_profile_tab():
            self._log("[FB PROFILE] Khong co scan o tab nay.")
            return
        if self._is_all_tab():
            self.start_all_scan()
            return
        if self._is_fb_tab():
            self.start_fb_scan()
            return
        if self.executor is not None:
            return

        try:
            max_threads = int(self.entry_threads.get())
            if max_threads <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Loi", "So luong phai > 0")
            return

        checked_emails = self._get_checked_email_set(self.tree)
        if not checked_emails:
            messagebox.showinfo("Thong bao", "Khong co profile nao duoc tick.")
            return

        self.stop_event.clear()
        self.executor = ThreadPoolExecutor(max_workers=max_threads)

        futures = []
        email_to_iid = self._map_email_to_item_id(self.tree)
        for acc in self.accounts:
            email = (acc.get("uid") or "").strip()
            if email not in checked_emails:
                continue
            item_id = email_to_iid.get(email)
            if not item_id:
                continue
            futures.append(self.executor.submit(self._scan_worker, item_id, acc))

        def _waiter():
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception:
                    pass
            self.executor = None

        threading.Thread(target=_waiter, daemon=True).start()

    def _scan_worker(self, item_id: str, acc: dict) -> None:
        if self.stop_event.is_set():
            return
        shorts_url = (acc.get("youtube") or "").strip()
        if not shorts_url:
            self._log(f"[{acc['uid']}] SCAN SKIP: No shorts URL")
            return
        self._set_status(item_id, "SCAN...")
        self._log(f"[{acc['uid']}] SCAN START")
        total, added = scan_shorts_for_email(
            acc["uid"],
            shorts_url,
            lambda: self.stop_event.is_set(),
            self._log,
        )
        self._set_status(item_id, f"SCAN OK ({added})")
        self._log(f"[{acc['uid']}] SCAN OK: added {added}, total {total}")

    def start_fb_scan(self) -> None:
        if self.executor is not None:
            return
        try:
            max_threads = int(self.entry_threads.get())
            if max_threads <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Loi", "So luong phai > 0")
            return

        pending = self._resume_pending.get("fb") or set()
        if pending:
            if self._prompt_resume("fb", len(pending)):
                checked_emails = pending
            else:
                self._resume_pending["fb"] = set()
                checked_emails = self._get_checked_email_set(self.fb_tree)
        else:
            checked_emails = self._get_checked_email_set(self.fb_tree)
        if not checked_emails:
            messagebox.showinfo("Thong bao", "Khong co profile nao duoc tick.")
            return

        self.stop_event.clear()
        try:
            fb_sort_state = self._sort_state.get("fb_followers") or self._sort_state.get("followers_all")
            if not fb_sort_state:
                self._reorder_tree_by_accounts(self.fb_tree, self.fb_accounts)
        except Exception:
            pass
        selected = []
        email_to_item = {}
        email_to_iid = self._map_email_to_item_id(self.fb_tree)
        for acc in self.fb_accounts:
            email = (acc.get("uid") or "").strip()
            if email not in checked_emails:
                continue
            item_id = email_to_iid.get(email)
            if not item_id:
                continue
            selected.append(acc)
            if email:
                email_to_item[email] = item_id
                self._set_fb_status(item_id, "QUEUED")

        def _on_status(email: str, status: str) -> None:
            item_id = email_to_item.get((email or "").strip())
            if not item_id:
                return
            self._set_fb_status(item_id, status)

        def _runner():
            scan_facebook_reels_multi(
                selected,
                stop_check=lambda: self.stop_event.is_set(),
                logger=self._log,
                cookie_file=os.path.join(DATA_DIR, "cookiefb.txt"),
                max_workers=2,
                on_status=_on_status,
            )

        self.executor = ThreadPoolExecutor(max_workers=1)
        fut = self.executor.submit(_runner)

        def _waiter():
            try:
                fut.result()
            except Exception as e:
                self._log(f"[FB SCAN] ERR: {e}")
            finally:
                try:
                    self.executor.shutdown(wait=False)
                except Exception:
                    pass
                self.executor = None

        threading.Thread(target=_waiter, daemon=True).start()

    def _fb_scan_worker(self, item_id: str, acc: dict) -> None:
        if self.stop_event.is_set():
            return
        reels_url = (acc.get("facebook") or "").strip()
        if not reels_url:
            self._log(f"[{acc['uid']}] FB SCAN SKIP: No reels URL")
            return
        self._set_fb_status(item_id, "SCAN...")
        self._log(f"[{acc['uid']}] FB SCAN START")
        total, added = scan_facebook_reels_for_email(
            acc["uid"],
            reels_url,
            lambda: self.stop_event.is_set(),
            self._log,
            cookie_file=os.path.join(DATA_DIR, "cookiefb.txt"),
        )
        self._set_fb_status(item_id, f"SCAN OK ({added})")
        self._log(f"[{acc['uid']}] FB SCAN OK: added {added}, total {total}")

    def start_fb_profile_jobs(self) -> None:
        if self.executor is not None:
            return
        self._force_close_all_profiles()
        self._reset_all_statuses()
        try:
            max_threads = int(self.entry_threads.get())
            if max_threads <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Loi", "So luong phai > 0")
            return

        pending = self._resume_pending.get("fb_profile") or set()
        if pending:
            if self._prompt_resume("fb_profile", len(pending)):
                checked_emails = pending
            else:
                self._resume_pending["fb_profile"] = set()
                checked_emails = self._get_checked_email_set(self.fb_profile_tree)
        else:
            checked_emails = self._get_checked_email_set(self.fb_profile_tree)
        if not checked_emails:
            messagebox.showinfo("Thong bao", "Khong co profile nao duoc tick.")
            return
        self._set_run_total("fb_profile", len(checked_emails))

        self._fixed_threads = max_threads
        self.stop_event.clear()
        self.executor = ThreadPoolExecutor(max_workers=max_threads)
        self.profile_semaphore = threading.BoundedSemaphore(max_threads)

        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080
        gap = 6
        taskbar_h = 40
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(checked_emails)
        cols = min(5, active_count)
        rows_layout = min(2, max(1, math.ceil(active_count / cols)))
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))

        futures = []
        slot_idx = 0
        max_slots = cols * rows_layout
        email_to_iid = self._map_email_to_item_id(self.fb_profile_tree)
        for acc in self.fb_profile_accounts:
            email = (acc.get("uid") or "").strip()
            if email not in checked_emails:
                continue
            item_id = email_to_iid.get(email)
            if not item_id:
                continue
            pos = slot_idx % max_slots
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            win_pos = f"{x},{y}"
            win_size = f"{win_w},{win_h}"
            futures.append(self.executor.submit(self._fb_profile_worker, item_id, acc, win_pos, win_size))
            slot_idx += 1

        def _waiter():
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception:
                    pass
            self.executor = None
            self._reset_run("fb_profile")

        threading.Thread(target=_waiter, daemon=True).start()

    def _fb_profile_worker(self, item_id: str, acc: dict, win_pos: str, win_size: str) -> None:
        sem = self.profile_semaphore
        if sem:
            sem.acquire()
        if self.stop_event.is_set():
            if sem:
                sem.release()
            return
        profile_id = None
        started = False
        email = (acc.get("uid") or "").strip()
        try:
            started = True
            fb_url = (acc.get("facebook") or "").strip()
            if not fb_url:
                self._set_fb_profile_status(item_id, "FB LINK ERR")
                return

            self._set_fb_profile_status(item_id, "CREATE...")
            ok_c = False
            data_c = {}
            msg_c = ""
            with self.create_lock:
                for attempt in range(3):
                    ok_c, data_c, msg_c = create_profile(acc["uid"], acc["proxy"], SCOOPZ_URL)
                    if ok_c:
                        break
                    self._set_fb_profile_status(item_id, f"CREATE RETRY {attempt+1}/3")
                    time.sleep(3 + attempt)
            if not ok_c:
                self._set_fb_profile_status(item_id, f"CREATE ERR: {msg_c}")
                return

            if isinstance(data_c, dict):
                profile_id = (data_c.get("data") or {}).get("id") or data_c.get("id") or data_c.get("profile_id")
            if not profile_id:
                self._set_fb_profile_status(item_id, "NO PROFILE ID")
                return
            self._remember_profile_path(profile_id, data_c)
            self.created_profiles.add(profile_id)

            self._set_fb_profile_status(item_id, "START...")
            ok_s, data_s, msg_s = start_profile(profile_id, win_pos=win_pos, win_size=win_size)
            if not ok_s:
                if self._is_proxy_error(msg_s):
                    try:
                        close_profile(profile_id, 3)
                        delete_profile(profile_id, 10)
                    except Exception:
                        pass
                    try:
                        self.created_profiles.discard(profile_id)
                    except Exception:
                        pass
                    self._delete_profile_path(profile_id)
                    new_id, new_data_s, _err = self._retry_start_profile_with_new_proxy(
                        acc,
                        item_id,
                        "fb_profile",
                        self.fb_profile_tree,
                        lambda s: self._set_fb_profile_status(item_id, s),
                        win_pos=win_pos,
                        win_size=win_size,
                        created_set="created_profiles",
                    )
                    if new_id and new_data_s:
                        profile_id = new_id
                        data_s = new_data_s
                        ok_s = True
                    else:
                        self._set_fb_profile_status(item_id, f"START ERR: {msg_s}")
                        return
                else:
                    self._set_fb_profile_status(item_id, f"START ERR: {msg_s}")
                    return
            driver_path, remote = extract_driver_info(data_s)
            if not (driver_path and remote):
                self._set_fb_profile_status(item_id, "STARTED (no debug)")
                return

            self._set_fb_profile_status(item_id, "LOGIN...")
            ok_login, err_login = login_scoopz(
                driver_path,
                remote,
                acc["uid"],
                acc["pass"],
                "",
                max_retries=3,
                keep_browser=True,
            )
            if not ok_login:
                self._set_fb_profile_status(item_id, self._format_login_error(err_login))
                return
            if self.stop_event.is_set():
                return

            self._set_fb_profile_status(item_id, "LOGIN OK")
            self._set_fb_profile_status(item_id, "FB FETCH...")
            fb_name, fb_username, avatar_path = fetch_facebook_profile_assets_local(fb_url, self._log)
            if not (fb_name and fb_username and avatar_path and os.path.exists(avatar_path)):
                self._set_fb_profile_status(item_id, "FB FETCH ERR")
                return
            self._save_profile_assets(acc["uid"], fb_name, fb_username, avatar_path)

            self._set_fb_profile_status(item_id, "OPEN PROFILE...")
            ok_pf = False
            err_pf = ""
            for attempt in range(1, 4):
                if self.stop_event.is_set():
                    return
                self._set_fb_profile_status(item_id, f"OPEN PROFILE... ({attempt}/3)")
                ok_pf, err_pf = open_profile_in_scoopz(
                    driver_path,
                    remote,
                    avatar_path,
                    fb_name,
                    fb_username,
                    logger=self._log,
                    max_retries=3,
                )
                if ok_pf:
                    break
                retryable = (
                    "cannot connect to chrome" in (err_pf or "").lower()
                    or "profile link not found" in (err_pf or "").lower()
                    or "profile page load timeout" in (err_pf or "").lower()
                )
                if not retryable:
                    break
                wait_s = 2 + attempt * 2
                self._log(f"[{acc['uid']}] FB PROFILE RETRY {attempt}/3 in {wait_s}s: {err_pf}")
                time.sleep(wait_s)
            if not ok_pf:
                self._set_fb_profile_status(item_id, f"PROFILE ERR: {err_pf}")
                return
            self._set_fb_profile_status(item_id, "PROFILE OPENED")
            self._set_fb_profile_status(item_id, "DONE")
        finally:
            if started:
                self._mark_run_done("fb_profile", email)
            if sem:
                try:
                    sem.release()
                except Exception:
                    pass
            try:
                if profile_id:
                    close_profile(profile_id, 3)
                    delete_profile(profile_id, 10)
            except Exception:
                pass
            try:
                if profile_id:
                    self.created_profiles.discard(profile_id)
            except Exception:
                pass
            if profile_id:
                self._delete_profile_path(profile_id)
                self._track_profile_cleanup()

    def start_fb_jobs(self) -> None:
        if self.executor is not None:
            return
        if not self._from_all_tab:
            if self._repeat_cycle_pending:
                self._repeat_cycle_pending = False
                self._increment_cycle()
            else:
                self._cycle_count = 1
                self._set_cycle_label()
                self._reset_upload_cycle_stats()
            self._runtime_reset()
            self._runtime_start()
        else:
            self._runtime_start()
        self._reset_upload_cycle_stats()
        self._force_close_all_profiles()
        self._reset_all_statuses()
        try:
            max_threads = int(self.entry_threads.get())
            if max_threads <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Loi", "So luong phai > 0")
            return
        try:
            max_videos = int(self.entry_videos.get())
            if max_videos <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Loi", "Videos phai > 0")
            return

        self._fixed_threads = max_threads
        pending = self._resume_pending.get("fb") or set()
        if pending:
            if self._prompt_resume("fb", len(pending)):
                checked_emails = pending
            else:
                self._resume_pending["fb"] = set()
                checked_emails = self._get_checked_email_set(self.fb_tree)
        else:
            checked_emails = self._get_checked_email_set(self.fb_tree)
        if not checked_emails:
            messagebox.showinfo("Thong bao", "Khong co profile nao duoc tick.")
            return
        self._set_run_total("fb", len(checked_emails))

        self.stop_event.clear()
        self._retry_round = 0
        with self.failed_accounts_lock:
            self.failed_accounts = []
        self._repeat_enabled = bool(self.repeat_var.get())
        try:
            delay_min = float(self.entry_repeat_delay.get())
            if delay_min < 0:
                delay_min = 0
        except Exception:
            delay_min = 0
        self._repeat_delay_sec = delay_min * 60.0
        self.executor = ThreadPoolExecutor(max_workers=max_threads)
        self.login_semaphore = threading.BoundedSemaphore(max_threads)
        self.upload_retry_semaphore = threading.BoundedSemaphore(max_threads)

        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080

        gap = 6
        taskbar_h = 40
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(checked_emails)
        cols = min(5, active_count)
        rows_layout = min(2, max(1, math.ceil(active_count / cols)))
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))

        futures = []
        slot_idx = 0
        max_slots = cols * rows_layout
        email_to_iid = self._map_email_to_item_id(self.fb_tree)
        acc_by_email = {str(a.get("uid") or "").strip().lower(): a for a in self.fb_accounts}
        ordered_emails = []
        seen_email_keys = set()
        try:
            for iid in self.fb_tree.get_children():
                email = (self.fb_tree.set(iid, "email") or "").strip()
                email_key = email.lower()
                if email and email in checked_emails and email_key not in seen_email_keys:
                    seen_email_keys.add(email_key)
                    ordered_emails.append(email)
        except Exception:
            for e in checked_emails:
                ek = (e or "").strip().lower()
                if ek and ek not in seen_email_keys:
                    seen_email_keys.add(ek)
                    ordered_emails.append(e)

        for email in ordered_emails:
            acc = acc_by_email.get((email or "").strip().lower())
            item_id = email_to_iid.get(email)
            if not acc or not item_id:
                continue
            self._bind_item_email(item_id, acc.get("uid", ""))
            pos = slot_idx % max_slots
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            win_pos = f"{x},{y}"
            win_size = f"{win_w},{win_h}"
            futures.append(self.executor.submit(self._fb_worker_one, item_id, acc, win_pos, win_size, max_videos))
            slot_idx += 1

        def _waiter():
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception:
                    pass
            self.executor = None
            with self.failed_accounts_lock:
                failed_list = self.failed_accounts.copy()
                self.failed_accounts = []
            if failed_list and not self.stop_event.is_set():
                if self._retry_round < self._upload_retry_rounds:
                    self._retry_round += 1
                    self._log(
                        f"[FB RETRY] Retrying {len(failed_list)} failed accounts (round {self._retry_round}/{self._upload_retry_rounds})..."
                    )
                    self.root.after(
                        1000,
                        lambda fl=failed_list: self._retry_failed_fb_accounts(fl, max_threads, max_videos),
                    )
                    return
                self._log(f"[FB RETRY] Stop retry after {self._upload_retry_rounds} rounds (remaining: {len(failed_list)})")
            # keep failed log
            self._reset_run("fb")
            if self._from_all_tab:
                self._from_all_tab = False
                if self._repeat_enabled and not self.stop_event.is_set():
                    delay_ms = 5 * 60 * 1000
                    self._start_next_cycle_countdown(300)
                    try:
                        self._log("[REPEAT] Running GPM cleanup during wait...")
                        self.root.after(0, self._clear_all_gpm_profiles)
                    except Exception:
                        pass

                    def _repeat_start():
                        if self.stop_event.is_set():
                            return
                        self._repeat_after_id = None
                        self._repeat_cycle_pending = True
                        self.start_jobs()

                    self._repeat_after_id = self.root.after(delay_ms, _repeat_start)
                return
            if self._repeat_enabled and not self.stop_event.is_set():
                delay_ms = 5 * 60 * 1000
                self._start_next_cycle_countdown(300)
                try:
                    self._log("[REPEAT] Running GPM cleanup during wait...")
                    self.root.after(0, self._clear_all_gpm_profiles)
                except Exception:
                    pass

                def _repeat_start():
                    if self.stop_event.is_set():
                        return
                    self._repeat_after_id = None
                    self._repeat_cycle_pending = True
                    self.start_fb_jobs()

                self._repeat_after_id = self.root.after(delay_ms, _repeat_start)

        threading.Thread(target=_waiter, daemon=True).start()

    def start_all_scan(self) -> None:
        if self.executor is not None:
            return
        checked = self._get_checked_all_rows()
        if not checked:
            messagebox.showinfo("Thong bao", "Khong co profile nao duoc tick.")
            return
        ytb_emails = [email for social, email in checked if social == "YTB"]
        fb_emails = [email for social, email in checked if social == "FB"]

        try:
            max_threads = int(self.entry_threads.get())
            if max_threads <= 0:
                raise ValueError
        except Exception:
            messagebox.showerror("Loi", "So luong phai > 0")
            return

        self.stop_event.clear()
        self.executor = ThreadPoolExecutor(max_workers=max_threads)

        futures = []
        ytb_by_email = {str(a.get("uid") or "").strip(): a for a in self.accounts}
        fb_by_email = {str(a.get("uid") or "").strip(): a for a in self.fb_accounts}
        ytb_email_to_iid = self._map_email_to_item_id(self.tree)
        fb_email_to_iid = self._map_email_to_item_id(self.fb_tree)

        for email in ytb_emails:
            acc = ytb_by_email.get(email)
            item_id = ytb_email_to_iid.get(email)
            if not acc or not item_id:
                continue
            futures.append(self.executor.submit(self._scan_worker, item_id, acc))

        for email in fb_emails:
            acc = fb_by_email.get(email)
            item_id = fb_email_to_iid.get(email)
            if not acc or not item_id:
                continue
            futures.append(self.executor.submit(self._fb_scan_worker, item_id, acc))

        def _waiter():
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception:
                    pass
            self.executor = None

        threading.Thread(target=_waiter, daemon=True).start()

    def _retry_failed_fb_accounts(self, failed_accounts: list, max_threads: int, max_videos: int) -> None:
        if self.stop_event.is_set():
            return

        self._log(f"[FB RETRY] Retrying {len(failed_accounts)} failed accounts...")
        try:
            for item_id, _acc in failed_accounts:
                self._set_fb_status(item_id, f"RETRY {self._retry_round}/{self._upload_retry_rounds}")
        except Exception:
            pass

        self.executor = ThreadPoolExecutor(max_workers=max_threads)
        futures = []

        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080
        gap = 6
        taskbar_h = 40
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(failed_accounts)
        cols = min(5, active_count)
        rows_layout = max(1, (active_count + cols - 1) // cols)
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))

        for idx, (item_id, acc) in enumerate(failed_accounts):
            if self.stop_event.is_set():
                break
            self._bind_item_email(item_id, acc.get("uid", ""))
            pos = idx % (cols * rows_layout)
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            retry_win_pos = f"{x},{y}"
            retry_win_size = f"{win_w},{win_h}"
            futures.append(self.executor.submit(self._fb_worker_one, item_id, acc, retry_win_pos, retry_win_size, max_videos))

        def _retry_waiter():
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception:
                    pass
            self.executor = None
            with self.failed_accounts_lock:
                failed_list = self.failed_accounts.copy()
                self.failed_accounts = []
            if failed_list and not self.stop_event.is_set():
                if self._retry_round < self._upload_retry_rounds:
                    self._retry_round += 1
                    self._log(
                        f"[FB RETRY] Retrying {len(failed_list)} failed accounts (round {self._retry_round}/{self._upload_retry_rounds})..."
                    )
                    self.root.after(
                        1000,
                        lambda fl=failed_list: self._retry_failed_fb_accounts(fl, max_threads, max_videos),
                    )
                    return
                self._log(f"[FB RETRY] Stop retry after {self._upload_retry_rounds} rounds (remaining: {len(failed_list)})")
            # keep failed log
            self._reset_run("fb")
            if self._from_all_tab:
                self._from_all_tab = False
                if self._repeat_enabled and not self.stop_event.is_set():
                    delay_ms = 5 * 60 * 1000
                    self._start_next_cycle_countdown(300)
                    try:
                        self._log("[REPEAT] Running GPM cleanup during wait...")
                        self.root.after(0, self._clear_all_gpm_profiles)
                    except Exception:
                        pass

                    def _repeat_start():
                        if self.stop_event.is_set():
                            return
                        self._repeat_after_id = None
                        self._repeat_cycle_pending = True
                        self.start_jobs()

                    self._repeat_after_id = self.root.after(delay_ms, _repeat_start)
                return

        threading.Thread(target=_retry_waiter, daemon=True).start()

    def _fb_worker_one(self, item_id: str, acc: dict, win_pos: str, win_size: str, max_videos: int) -> None:
        email = (acc.get("uid") or "").strip()
        started = False
        try:
            if self.stop_event.is_set():
                return
            started = True
            profile_id = None
            max_file_size_bytes = 100 * 1024 * 1024
    
            def _extract_fb_video_id(text: str) -> str:
                val = (text or "").strip()
                if not val:
                    return ""
                m = re.search(r"/reel/(\d+)", val)
                if m:
                    return m.group(1)
                return ""
    
            self._log(f"[{acc['uid']}] FB START")
            self._set_fb_status(item_id, "CREATE...")
            ok_c = False
            data_c = {}
            msg_c = ""
            with self.create_lock:
                for attempt in range(3):
                    ok_c, data_c, msg_c = create_profile(acc["uid"], acc["proxy"], SCOOPZ_URL)
                    if ok_c:
                        break
                    wait_s = 5 + attempt * 3
                    self._set_fb_status(item_id, f"CREATE RETRY {attempt+1}/3")
                    self._log(f"[{acc['uid']}] CREATE ERR: {msg_c} | retry in {wait_s}s")
                    time.sleep(wait_s)
            if not ok_c:
                self._set_fb_status(item_id, f"CREATE ERR: {msg_c}")
                self._record_failed(item_id, acc, f"CREATE ERR: {msg_c}")
                return
    
            profile_id = None
            if isinstance(data_c, dict):
                profile_id = (data_c.get("data") or {}).get("id") or data_c.get("id") or data_c.get("profile_id")
            if not profile_id:
                self._set_fb_status(item_id, "NO PROFILE ID")
                self._record_failed(item_id, acc, "NO PROFILE ID")
                return
            self._set_fb_profile_info(item_id, profile_id=profile_id)
            self._remember_profile_path(profile_id, data_c)
            self.created_profiles.add(profile_id)
    
            self._set_fb_status(item_id, "START...")
            ok_s, data_s, msg_s = start_profile(profile_id, win_pos=win_pos, win_size=win_size)
            if not ok_s:
                if self._is_proxy_error(msg_s):
                    try:
                        close_profile(profile_id, 3)
                        delete_profile(profile_id, 10)
                    except Exception:
                        pass
                    try:
                        self.created_profiles.discard(profile_id)
                    except Exception:
                        pass
                    self._delete_profile_path(profile_id)
                    new_id, new_data_s, _err = self._retry_start_profile_with_new_proxy(
                        acc,
                        item_id,
                        "fb",
                        self.fb_tree,
                        lambda s: self._set_fb_status(item_id, s),
                        win_pos=win_pos,
                        win_size=win_size,
                        created_set="created_profiles",
                    )
                    if new_id and new_data_s:
                        profile_id = new_id
                        data_s = new_data_s
                        ok_s = True
                    else:
                        self._set_fb_status(item_id, f"START ERR: {msg_s}")
                        self._record_failed(item_id, acc, f"START ERR: {msg_s}")
                        return
                else:
                    self._set_fb_status(item_id, f"START ERR: {msg_s}")
                    self._record_failed(item_id, acc, f"START ERR: {msg_s}")
                    return
    
            with self.active_lock:
                self.active_profiles[item_id] = profile_id
            driver_path, remote = extract_driver_info(data_s)
            status = "STARTED" if driver_path and remote else "STARTED (no debug)"
            self._set_fb_status(item_id, status)
    
            if not (driver_path and remote):
                self._record_failed(item_id, acc, "STARTED (no debug)")
                return
    
            self._set_fb_status(item_id, "LOGIN...")
            ok_login, err_login = login_scoopz(
                driver_path,
                remote,
                acc["uid"],
                acc["pass"],
                "",
                max_retries=3,
                keep_browser=True,
            )
            if not ok_login:
                status = self._format_login_error(err_login)
                self._set_fb_status(item_id, status)
                self._record_failed(item_id, acc, status)
                return
    
            self._set_fb_status(item_id, "LOGIN OK")
            if SKIP_DOWNLOAD_UPLOAD:
                try:
                    followers = None
                    profile_url = ""
                    posts = None
                    for attempt in range(3):
                        followers, profile_url, posts = fetch_followers(driver_path, remote, self._log, acc.get("uid", ""))
                        if followers is not None or posts is not None:
                            break
                        time.sleep(2 + attempt)
                    if followers is not None or posts is not None:
                        if followers is not None:
                            self._log(f"[{acc['uid']}] FOLLOWERS: {followers}")
                        if posts is not None:
                            self._log(f"[{acc['uid']}] POSTS: {posts}")
                        self._set_fb_profile_info(item_id, profile_url, followers, posts)
                        self._set_fb_status(item_id, "FOLLOW OK")
                    else:
                        self._set_fb_status(item_id, "FOLLOW ERR")
                        self._log(f"[{acc['uid']}] FOLLOW ERR")
                except Exception as e:
                    self._log(f"[{acc['uid']}] FOLLOW ERR: {e}")
                return
            success_count = 0
            safety_guard = 0
            while success_count < max_videos:
                if self.stop_event.is_set():
                    break
                safety_guard += 1
                if safety_guard > max_videos * 5:
                    break
    
                self.operation_delayer.delay_before_download(acc["uid"], self._log_progress)
                ok_next, row = get_next_unuploaded(acc["uid"])
                if not ok_next:
                    self._set_fb_status(item_id, "NO VIDEO")
                    break
                row_url = (row.get("url") or "").strip()
                row_id = (row.get("video_id") or "").strip()
                if not row_url:
                    if row_id.startswith("http"):
                        row_url = row_id
                    elif row_id:
                        row_url = f"https://www.facebook.com/reel/{row_id}"
                if not row_url:
                    break
    
                self._set_fb_status(item_id, f"DOWNLOAD {success_count+1}/{max_videos}...")
                retry_dl = 0
                skip_current = False
                skip_account = False
                ok_dl = False
                path_or_err = ""
                vid_id = ""
                title = ""
                download_start_ts = time.time()
                download_timed_out = threading.Event()
                watchdog = self._start_download_watchdog(
                    acc["uid"],
                    "FB DOWNLOAD",
                    max_seconds=600,
                    on_timeout=lambda _e: download_timed_out.set(),
                )
                while True:
                    ok_dl, path_or_err, vid_id, title = download_one_facebook(
                        acc["uid"],
                        row_url,
                        self._log_progress,
                        cookie_path=os.path.join(DATA_DIR, "cookiefb.txt"),
                        timeout_s=30,
                    )
                    if ok_dl:
                        break
                    err_text = str(path_or_err)
                    lower = err_text.lower()
                    is_timeout = "timeout" in lower or "timed out" in lower
                    if is_timeout:
                        if retry_dl < 1:
                            retry_dl += 1
                            self._log(f"[{acc['uid']}] FB DOWNLOAD TIMEOUT - RETRY 1/1")
                            continue
                        self._log(f"[{acc['uid']}] FB DOWNLOAD TIMEOUT - SKIP ACCOUNT")
                        self._set_fb_status(item_id, "DOWNLOAD TIMEOUT")
                        self._record_failed(item_id, acc, "DOWNLOAD TIMEOUT")
                        skip_account = True
                        break
                    break
                try:
                    watchdog.set()
                except Exception:
                    pass
                elapsed = int(time.time() - download_start_ts)
                if ok_dl:
                    self._log(f"[{acc['uid']}] FB DOWNLOAD END OK after {elapsed}s")
                else:
                    self._log(f"[{acc['uid']}] FB DOWNLOAD END ERR after {elapsed}s: {path_or_err}")
                if download_timed_out.is_set():
                    self._set_fb_status(item_id, "DOWNLOAD TIMEOUT")
                    try:
                        self.error_logger.log_download_error(acc["uid"], row_url, "DOWNLOAD TIMEOUT (watchdog 600s)")
                    except Exception:
                        pass
                    self._record_failed(item_id, acc, "DOWNLOAD TIMEOUT")
                    skip_account = True
                mark_id = vid_id or row_id or _extract_fb_video_id(row_url)
                if not ok_dl:
                    err_text = str(path_or_err)
                    lower = err_text.lower()
                    if (
                        "video skipped" in lower
                        or "private" in lower
                        or "isn't available" in lower
                        or "video unavailable" in lower
                        or "video is unavailable" in lower
                        or "watch video on youtube" in lower
                    ):
                        try:
                            mark_uploaded(acc["uid"], mark_id)
                        except Exception:
                            pass
                        continue
                    if skip_account:
                        break
                    if skip_current:
                        continue
                    self._set_fb_status(item_id, f"DOWNLOAD ERR: {err_text}")
                    self._record_failed(item_id, acc, f"DOWNLOAD ERR: {err_text}")
                    break
    
                title = self._ensure_title(title, 1000)
                if title:
                    try:
                        title = self._format_fb_title_case1(title)
                        title = self._limit_caption(title, 1000)
                    except Exception:
                        pass
                if vid_id and title:
                    try:
                        update_title_if_empty(acc["uid"], vid_id, title)
                    except Exception:
                        pass
                try:
                    if os.path.exists(path_or_err):
                        file_size = os.path.getsize(path_or_err)
                        if file_size > max_file_size_bytes:
                            self._log(
                                f"[{acc['uid']}] SKIP BIG FILE: {file_size / (1024 * 1024):.1f}MB > 100MB"
                            )
                            try:
                                mark_uploaded(acc["uid"], mark_id)
                            except Exception:
                                pass
                            self._delete_uploaded_video(path_or_err, acc["uid"])
                            continue
                except Exception:
                    pass
                title = self._ensure_title(title, 1000)
                if vid_id and title:
                    try:
                        update_title_if_empty(acc["uid"], vid_id, title)
                    except Exception:
                        pass
                caption = self._build_caption(title, 1000)
                ok_p, drv, up_status, up_msg, caption = self._prepare_upload_with_retry(
                    driver_path,
                    remote,
                    path_or_err,
                    caption,
                    acc.get("uid", ""),
                    caption_limit=1000,
                )
    
                if not ok_p:
                    if up_status in ("select_not_found", "select_click_error"):
                        self._set_fb_status(item_id, f"UPLOAD LOI: {up_status}")
                        self._record_failed(item_id, acc, f"UPLOAD {up_status}")
                    else:
                        self._set_fb_status(item_id, f"UPLOAD ERR: {up_msg or up_status}")
                        self._record_failed(item_id, acc, f"UPLOAD ERR: {up_msg or up_status}")
                    break
    
                self._set_fb_status(item_id, f"POSTING {success_count+1}/{max_videos}...")
                st, msg, _purl, _foll, _posts = self._post_uploaded_video(drv, acc.get("uid", ""))
                if st == "success":
                    try:
                        mark_uploaded(acc["uid"], mark_id)
                    except Exception:
                        pass
                    self._set_fb_profile_info(item_id, _purl, _foll, _posts)
                    self._set_fb_status(item_id, "UPLOAD OK")
                    self._delete_uploaded_video(path_or_err, acc["uid"])
                    success_count += 1
                else:
                    err_text = msg or st
                    self._set_fb_status(item_id, f"UPLOAD ERR: {err_text}")
                    self._record_failed(item_id, acc, f"UPLOAD ERR: {err_text}")
                    break
    
            try:
                if profile_id:
                    close_profile(profile_id, 3)
                    delete_profile(profile_id, 10)
            except Exception:
                pass
            try:
                if profile_id:
                    self.created_profiles.discard(profile_id)
            except Exception:
                pass
            if profile_id:
                self._delete_profile_path(profile_id)
                self._track_profile_cleanup()
            try:
                with self.active_lock:
                    self.active_profiles.pop(item_id, None)
            except Exception:
                pass
        finally:
            if started:
                self._mark_run_done("fb", email)
    def _retry_failed_accounts(self, failed_accounts: list, max_threads: int, max_videos: int) -> None:
        """Retry failed accounts with new threads"""
        if self.stop_event.is_set():
            return
        
        self._log(f"[RETRY] Retrying {len(failed_accounts)} failed accounts (max {max_threads} threads)...")
        self._clear_status_tags()
        try:
            for item_id, _acc in failed_accounts:
                self._set_status(item_id, f"RETRY {self._retry_round}/{self._upload_retry_rounds}")
        except Exception:
            pass
        
        self.executor = ThreadPoolExecutor(max_workers=max_threads)
        futures = []
        
        # Re-calculate layout for failed accounts
        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080
        
        gap = 6  # Tight spacing between windows
        taskbar_h = 40  # Reserve space for taskbar
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(failed_accounts)
        
        # Fixed: 5 columns per row (5 profiles = 1 row, 10 profiles = 2 rows)
        cols = min(5, active_count)  # Max 5 per row
        rows_layout = max(1, (active_count + cols - 1) // cols)
        
        # Calculate window sizes
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        
        # Compact windows: narrow width, optimal height
        win_w = max(150, min(280, win_w))  # Compact: 150-280px
        win_h = max(420, min(600, win_h))  # Height range: 420-600px
        
        for idx, (item_id, acc) in enumerate(failed_accounts):
            if self.stop_event.is_set():
                break
            self._bind_item_email(item_id, acc.get("uid", ""))
            pos = idx % (cols * rows_layout)
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            retry_win_pos = f"{x},{y}"
            retry_win_size = f"{win_w},{win_h}"
            self._log(f"[RETRY] Submitting {acc.get('uid', '')} (item_id={item_id}) at position {retry_win_pos}")
            futures.append(self.executor.submit(self._worker_one, item_id, acc, retry_win_pos, retry_win_size, max_videos))
        
        def _retry_waiter():
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception:
                    pass
            self.executor = None
            with self.failed_accounts_lock:
                failed_list = self.failed_accounts.copy()
                self.failed_accounts = []
            try:
                failed_list.extend(self._collect_transient_failures())
            except Exception:
                pass
            if failed_list and not self.stop_event.is_set():
                if self._retry_round < self._upload_retry_rounds:
                    self._retry_round += 1
                    self._log(
                        f"[RETRY] Retrying {len(failed_list)} failed accounts (round {self._retry_round}/{self._upload_retry_rounds})..."
                    )
                    self.root.after(1000, lambda fl=failed_list: self._retry_failed_accounts(fl, max_threads, max_videos))
                    return
                self._log(f"[RETRY] Stop retry after {self._upload_retry_rounds} rounds (remaining: {len(failed_list)})")
            # keep failed log
            if self._repeat_enabled and not self.stop_event.is_set():
                delay_ms = int(self._repeat_delay_sec * 1000)
                if delay_ms < 0:
                    delay_ms = 0

                def _repeat_start():
                    if self.stop_event.is_set():
                        return
                    self._repeat_after_id = None
                    self._repeat_cycle_pending = True
                    self.start_jobs()

                self._repeat_after_id = self.root.after(delay_ms, _repeat_start)
        
        threading.Thread(target=_retry_waiter, daemon=True).start()

    def _retry_failed_profile_accounts(self, failed_accounts: list, max_threads: int) -> None:
        if self.stop_event.is_set():
            return

        self._log(f"[PROFILE RETRY] Retrying {len(failed_accounts)} failed accounts (max {max_threads} threads)...")
        self.executor = ThreadPoolExecutor(max_workers=max_threads)
        futures = []

        try:
            screen_w = self.root.winfo_screenwidth()
            screen_h = self.root.winfo_screenheight()
        except Exception:
            screen_w, screen_h = 1920, 1080

        gap = 6
        taskbar_h = 40
        usable_w = screen_w - (gap * 2)
        usable_h = (screen_h - taskbar_h) - (gap * 2)
        active_count = len(failed_accounts)
        cols = min(5, active_count)
        rows_layout = max(1, (active_count + cols - 1) // cols)
        win_w = int((usable_w - gap * (cols - 1)) / cols)
        win_h = int((usable_h - gap * (rows_layout - 1)) / rows_layout)
        win_w = max(150, min(280, win_w))
        win_h = max(420, min(600, win_h))

        for idx, (item_id, acc) in enumerate(failed_accounts):
            if self.stop_event.is_set():
                break
            pos = idx % (cols * rows_layout)
            col = pos % cols
            row = pos // cols
            x = gap + col * (win_w + gap)
            y = gap + row * (win_h + gap)
            retry_win_pos = f"{x},{y}"
            retry_win_size = f"{win_w},{win_h}"
            self._log(f"[PROFILE RETRY] Submitting {acc.get('uid', '')} (item_id={item_id}) at position {retry_win_pos}")
            futures.append(self.executor.submit(self._profile_open_worker, item_id, acc, retry_win_pos, retry_win_size))

        def _retry_waiter():
            for f in as_completed(futures):
                try:
                    f.result()
                except Exception:
                    pass
            self.executor = None
            with self.profile_failed_lock:
                failed_list = self.profile_failed_accounts.copy()
                self.profile_failed_accounts = []
            if failed_list and not self.stop_event.is_set():
                if self._profile_retry_round < self._max_retry_rounds:
                    self._profile_retry_round += 1
                    self._log(
                        f"[PROFILE RETRY] Retrying {len(failed_list)} failed accounts (round {self._profile_retry_round}/{self._max_retry_rounds})..."
                    )
                    self.root.after(1000, lambda fl=failed_list: self._retry_failed_profile_accounts(fl, max_threads))
                    return
            # keep profile failed log

        threading.Thread(target=_retry_waiter, daemon=True).start()


    def _ensure_logged_in(self, item_id: str, acc: dict, win_pos: str | None = None, win_size: str | None = None):
        with self.active_drivers_lock:
            info = self.active_drivers.get(item_id)
        if info and info.get("driver_path") and info.get("remote"):
            return info.get("driver_path"), info.get("remote"), info.get("profile_id")

        self._set_status(item_id, "CREATE...")
        ok_c = False
        data_c = {}
        msg_c = ""
        with self.create_lock:
            for attempt in range(3):
                ok_c, data_c, msg_c = create_profile(acc["uid"], acc["proxy"], SCOOPZ_URL)
                if ok_c:
                    break
                wait_s = 5 + attempt * 3
                self._set_status(item_id, f"CREATE RETRY {attempt+1}/3")
                self._log(f"[{acc['uid']}] CREATE ERR: {msg_c} | retry in {wait_s}s")
                time.sleep(wait_s)
        if not ok_c:
            self._set_status(item_id, f"CREATE ERR: {msg_c}")
            self._log(f"[{acc['uid']}] CREATE ERR: {msg_c}")
            self._record_failed(item_id, acc, f"CREATE ERR: {msg_c}")
            return None, None, None

        profile_id = None
        if isinstance(data_c, dict):
            profile_id = (data_c.get("data") or {}).get("id") or data_c.get("id") or data_c.get("profile_id")
        if not profile_id:
            self._set_status(item_id, "NO PROFILE ID")
            return None, None, None
        self._remember_profile_path(profile_id, data_c)
        self.created_profiles.add(profile_id)

        self._set_status(item_id, "START...", profile_id=profile_id)
        if win_pos is None:
            ok_s, data_s, msg_s = start_profile(profile_id)
        else:
            ok_s, data_s, msg_s = start_profile(profile_id, win_pos=win_pos, win_size=win_size)
        if not ok_s:
            if self._is_proxy_error(msg_s):
                try:
                    close_profile(profile_id, 3)
                    delete_profile(profile_id, 10)
                except Exception:
                    pass
                try:
                    self.created_profiles.discard(profile_id)
                except Exception:
                    pass
                self._delete_profile_path(profile_id)
                new_id, new_data_s, _err = self._retry_start_profile_with_new_proxy(
                    acc,
                    item_id,
                    "upload",
                    self.tree,
                    lambda s: self._set_status(item_id, s),
                    win_pos=win_pos,
                    win_size=win_size,
                    created_set="created_profiles",
                )
                if new_id and new_data_s:
                    profile_id = new_id
                    data_s = new_data_s
                    ok_s = True
                else:
                    self._set_status(item_id, f"START ERR: {msg_s}")
                    self._log(f"[{acc['uid']}] START ERR: {msg_s}")
                    self._record_failed(item_id, acc, f"START ERR: {msg_s}")
                    return None, None, None
            else:
                self._set_status(item_id, f"START ERR: {msg_s}")
                self._log(f"[{acc['uid']}] START ERR: {msg_s}")
                self._record_failed(item_id, acc, f"START ERR: {msg_s}")
                return None, None, None

        driver_path, remote = extract_driver_info(data_s)
        if not driver_path or not remote:
            self._set_status(item_id, "STARTED (no debug)")
            try:
                close_profile(profile_id, 3)
                delete_profile(profile_id, 10)
            except Exception:
                pass
            try:
                self.created_profiles.discard(profile_id)
            except Exception:
                pass
            self._delete_profile_path(profile_id)
            return None, None, None

        self._set_status(item_id, "LOGIN...")
        ok_login, err_login = login_scoopz(
            driver_path,
            remote,
            acc["uid"],
            acc["pass"],
            "",
            max_retries=3,
            keep_browser=True,
        )
        if not ok_login:
            status = self._format_login_error(err_login)
            self._set_status(item_id, status)
            self._log(f"[{acc['uid']}] {status}")
            self._record_failed(item_id, acc, status)
            try:
                close_profile(profile_id, 3)
                delete_profile(profile_id, 10)
            except Exception:
                pass
            try:
                self.created_profiles.discard(profile_id)
            except Exception:
                pass
            self._delete_profile_path(profile_id)
            return None, None, None

        self._set_status(item_id, "LOGIN OK")
        self._log(f"[{acc['uid']}] LOGIN OK")
        with self.active_drivers_lock:
            self.active_drivers[item_id] = {
                "profile_id": profile_id,
                "driver_path": driver_path,
                "remote": remote,
                "close_func": lambda: (close_profile(profile_id, 3), delete_profile(profile_id, 10)),
            }
        return driver_path, remote, profile_id

    def _cleanup_profile_session(self, item_id: str = None, profile_id: str = None) -> None:
        info = None
        if item_id is not None:
            with self.active_drivers_lock:
                info = self.active_drivers.pop(item_id, None)
        if info and not profile_id:
            profile_id = info.get("profile_id")
        try:
            if info and "close_func" in info:
                info["close_func"]()
        except Exception:
            pass
        if profile_id:
            try:
                close_profile(profile_id, 3)
                delete_profile(profile_id, 10)
            except Exception:
                pass
            try:
                self.created_profiles.discard(profile_id)
            except Exception:
                pass
            self._delete_profile_path(profile_id)
            self._track_profile_cleanup()

    def _extract_profile_path(self, data: dict) -> str:
        payload = data.get("data") if isinstance(data, dict) and isinstance(data.get("data"), dict) else data
        if not isinstance(payload, dict):
            return ""
        for key in (
            "profile_path",
            "profilePath",
            "profile_dir",
            "profileDir",
            "path",
            "folder",
            "local_path",
            "localPath",
        ):
            val = payload.get(key)
            if isinstance(val, str) and val.strip():
                return val.strip()
        return ""

    def _get_gpm_root(self) -> str:
        try:
            raw = self.entry_gpm_path.get().strip()
        except Exception:
            raw = ""
        if not raw:
            raw = self._default_gpm_root
        try:
            path = os.path.abspath(os.path.expanduser(raw))
        except Exception:
            path = raw
        if self._is_unsafe_gpm_root(path):
            fallback = self._default_gpm_root
            try:
                os.makedirs(fallback, exist_ok=True)
            except Exception:
                pass
            try:
                self.entry_gpm_path.delete(0, tk.END)
                self.entry_gpm_path.insert(0, fallback)
            except Exception:
                pass
            try:
                self._log(f"[GPM] Unsafe root blocked: {path} -> {fallback}")
            except Exception:
                pass
            return fallback
        return path

    def _is_unsafe_gpm_root(self, path: str) -> bool:
        try:
            norm = os.path.normcase(os.path.abspath(path))
        except Exception:
            norm = os.path.normcase(str(path or ""))
        blocked = set()
        for candidate in (_THIS_DIR, _BASE_DIR, os.getcwd()):
            if not candidate:
                continue
            try:
                blocked.add(os.path.normcase(os.path.abspath(candidate)))
            except Exception:
                blocked.add(os.path.normcase(candidate))
        return norm in blocked

    def _guess_profile_path(self) -> str:
        gpm_root = self._get_gpm_root()
        if not gpm_root or not os.path.isdir(gpm_root):
            return ""
        try:
            entries = []
            for name in os.listdir(gpm_root):
                if name.lower() == "profile_data.db":
                    continue
                path = os.path.join(gpm_root, name)
                if not os.path.isdir(path):
                    continue
                if path in self.profile_paths_used:
                    continue
                try:
                    mtime = os.path.getmtime(path)
                except Exception:
                    mtime = 0
                entries.append((mtime, path))
            if not entries:
                return ""
            entries.sort(key=lambda x: x[0], reverse=True)
            return entries[0][1]
        except Exception:
            return ""

    def _repair_gpm_clipboard_extension(self, profile_path: str) -> None:
        if not profile_path:
            return
        try:
            base = os.path.abspath(profile_path)
        except Exception:
            base = profile_path
        ext_dir = os.path.join(base, "Default", "GPMSoft", "Extensions", "clipboard-ext")
        manifest_path = os.path.join(ext_dir, "manifest.json")
        script_path = os.path.join(ext_dir, "contentscript.js")

        try:
            os.makedirs(ext_dir, exist_ok=True)
        except Exception:
            return

        manifest_ok = False
        if os.path.exists(manifest_path):
            try:
                with open(manifest_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    manifest_ok = True
            except Exception:
                manifest_ok = False

        if not manifest_ok:
            manifest = {
                "manifest_version": 3,
                "name": "GPM Clipboard",
                "version": "1.0.0",
                "description": "Auto-repaired placeholder extension.",
                "content_scripts": [
                    {
                        "matches": ["<all_urls>"],
                        "js": ["contentscript.js"],
                        "run_at": "document_start",
                    }
                ],
            }
            try:
                with open(manifest_path, "w", encoding="utf-8") as f:
                    json.dump(manifest, f, ensure_ascii=False, indent=2)
            except Exception:
                pass

        if not os.path.exists(script_path):
            try:
                with open(script_path, "w", encoding="utf-8") as f:
                    f.write("// Auto-repaired placeholder script for GPM clipboard extension.\n")
            except Exception:
                pass

        try:
            self._log(f"[GPM] clipboard-ext checked: {ext_dir}")
        except Exception:
            pass

    def _remember_profile_path(self, profile_id: str, data: dict) -> None:
        if not profile_id:
            return
        path = self._extract_profile_path(data or {})
        if not path:
            path = self._guess_profile_path()
            if not path:
                try:
                    self._log(f"[GPM] profile_path missing for {profile_id}")
                except Exception:
                    pass
                return
            try:
                self._log(f"[GPM] profile_path guessed for {profile_id}: {path}")
            except Exception:
                pass
        with self.profile_paths_lock:
            self.profile_paths[profile_id] = path
            self.profile_paths_used.add(path)
        try:
            self._repair_gpm_clipboard_extension(path)
        except Exception:
            pass
        try:
            self._log(f"[GPM] profile_path for {profile_id}: {path}")
        except Exception:
            pass

    def _delete_profile_path(self, profile_id: str) -> None:
        if not profile_id:
            return
        with self.profile_paths_lock:
            path = self.profile_paths.pop(profile_id, None)
            if path:
                self.profile_paths_used.discard(path)
        if not path:
            return
        try:
            abs_path = os.path.abspath(path)
            if os.path.isdir(abs_path):
                shutil.rmtree(abs_path, ignore_errors=True)
        except Exception:
            pass

    def _clear_all_gpm_profiles(self) -> None:
        """Delete all GPM profiles by calling GPM list API (background thread)."""
        try:
            popup = tk.Toplevel(self.root)
            popup.title("GPM Cleanup")
            popup.geometry("360x120")
            popup.resizable(False, False)
            popup.transient(self.root)
            label = tk.Label(popup, text="Loading GPM profiles...", font=("Segoe UI", 10))
            label.pack(expand=True, fill="both", padx=12, pady=20)
            self.root.update_idletasks()
            try:
                self.root.update_idletasks()
                x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 180
                y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 60
                popup.geometry(f"+{max(0, x)}+{max(0, y)}")
            except Exception:
                pass
        except Exception:
            popup = None
            label = None

        def _set_label(text: str) -> None:
            try:
                if label is not None:
                    label.configure(text=text)
                    self.root.update_idletasks()
            except Exception:
                pass

        def _worker() -> None:
            # For very large datasets, avoid loading all IDs into memory at once.
            # Strategy: always fetch page=1 in small chunks, delete that chunk, repeat until empty.
            per_page = 500
            max_workers = 8
            deleted_count = 0
            list_retry_limit = 5
            session = requests.Session()
            estimate_total = None
            overall_t0 = time.time()
            batch_index = 0

            def _extract_items(payload) -> list:
                items = []
                if isinstance(payload, dict):
                    data = payload.get("data")
                    if isinstance(data, dict):
                        items = data.get("data") or data.get("items") or []
                    elif isinstance(data, list):
                        items = data
                return items if isinstance(items, list) else []

            while True:
                items = []
                last_err = ""
                list_t0 = time.time()
                for attempt in range(1, list_retry_limit + 1):
                    try:
                        url = f"http://127.0.0.1:19995/api/v3/profiles?page=1&per_page={per_page}"
                        resp = session.get(url, timeout=45)
                        resp.raise_for_status()
                        payload = resp.json() if resp.content else {}
                        items = _extract_items(payload)
                        if estimate_total is None and isinstance(payload, dict):
                            pg = payload.get("pagination") or {}
                            if isinstance(pg, dict):
                                total_val = pg.get("total")
                                if isinstance(total_val, int) and total_val >= 0:
                                    estimate_total = total_val
                        break
                    except Exception as e:
                        last_err = str(e)
                        if attempt < list_retry_limit:
                            wait_s = min(8, attempt * 2)
                            self._log(f"[GPM] List retry {attempt}/{list_retry_limit} in {wait_s}s: {e}")
                            time.sleep(wait_s)
                        else:
                            self._log(f"[GPM] List profiles error: {e}")

                if not items and last_err:
                    self.root.after(0, _set_label, f"List profiles error: {last_err}")
                    return

                list_elapsed = max(0.001, time.time() - list_t0)
                if items:
                    list_rate = len(items) / list_elapsed
                    self._log(
                        f"[GPM] Batch {batch_index + 1}: list fetched {len(items)} profiles in {list_elapsed:.2f}s ({list_rate:.2f} profiles/s)"
                    )

                if not items:
                    if deleted_count == 0:
                        self._log("[GPM] No profile IDs found to delete")
                        self.root.after(0, _set_label, "No profiles found")
                    else:
                        self.root.after(0, _set_label, f"Done ({deleted_count})")
                        self._log(f"[GPM] Delete all profiles done ({deleted_count})")
                    break

                ids = []
                for item in items:
                    pid = (item.get("id") or item.get("profile_id") or "").strip()
                    if pid:
                        ids.append(pid)

                if not ids:
                    self.root.after(0, _set_label, f"Done ({deleted_count})")
                    self._log(f"[GPM] No valid profile IDs in current page, stop ({deleted_count})")
                    break

                if estimate_total is not None:
                    self.root.after(
                        0,
                        _set_label,
                        f"Deleting {deleted_count}/{estimate_total}... (batch {len(ids)})",
                    )
                else:
                    self.root.after(0, _set_label, f"Deleting {deleted_count}... (batch {len(ids)})")

                batch_index += 1
                batch_t0 = time.time()
                with ThreadPoolExecutor(max_workers=max_workers) as ex:
                    futures = [ex.submit(delete_profile, pid, 12) for pid in ids]
                    batch_done = 0
                    for _f in as_completed(futures):
                        batch_done += 1
                        deleted_count += 1
                        if batch_done % 100 == 0 or batch_done == len(ids):
                            if estimate_total is not None:
                                self.root.after(0, _set_label, f"Deleting {deleted_count}/{estimate_total}...")
                            else:
                                self.root.after(0, _set_label, f"Deleting {deleted_count}...")

                batch_elapsed = max(0.001, time.time() - batch_t0)
                batch_rate = batch_done / batch_elapsed if batch_done else 0.0
                overall_elapsed = max(0.001, time.time() - overall_t0)
                overall_rate = deleted_count / overall_elapsed if deleted_count else 0.0
                self._log(
                    f"[GPM] Batch {batch_index}: deleted {batch_done} profiles in {batch_elapsed:.2f}s "
                    f"({batch_rate:.2f} profiles/s) | total {deleted_count} in {overall_elapsed:.2f}s ({overall_rate:.2f} profiles/s)"
                )

                # Small pause to reduce pressure on local GPM API between batches.
                time.sleep(0.15)

            try:
                time.sleep(0.6)
            except Exception:
                pass
            try:
                if popup is not None:
                    self.root.after(0, popup.destroy)
            except Exception:
                pass

        threading.Thread(target=_worker, daemon=True).start()

    def _track_profile_cleanup(self) -> None:
        with self._gpm_cleanup_lock:
            self._gpm_cleanup_count += 1
            count = self._gpm_cleanup_count
        if count >= 10:
            self._cleanup_gpm_root_if_needed()

    def _cleanup_gpm_root_if_needed(self) -> None:
        with self._gpm_cleanup_lock:
            if self._gpm_cleanup_running:
                return
            self._gpm_cleanup_running = True
        try:
            self._cleanup_gpm_root(force=False)
        finally:
            with self._gpm_cleanup_lock:
                self._gpm_cleanup_count = 0
                self._gpm_cleanup_running = False

    def _cleanup_gpm_root(self, force: bool = False) -> None:
        gpm_root = self._get_gpm_root()
        if not gpm_root or not os.path.isdir(gpm_root):
            return
        try:
            if force:
                self._log(f"[GPM] Cleaning all profile folders in {gpm_root} (force)")
            else:
                self._log(f"[GPM] Cleaning profile folders in {gpm_root}")
        except Exception:
            pass
        try:
            for name in os.listdir(gpm_root):
                if name.lower() == "profile_data.db":
                    continue
                path = os.path.join(gpm_root, name)
                if not os.path.isdir(path):
                    continue
                if not force and path in self.profile_paths_used:
                    continue
                try:
                    shutil.rmtree(path, ignore_errors=True)
                except Exception:
                    pass
        except Exception:
            pass

    def _login_only_worker(self, item_id: str, acc: dict) -> None:
        if self.stop_event.is_set():
            return
        try:
            self._ensure_logged_in(item_id, acc)
            self._set_status(item_id, "LOGIN OK (HOLD)")
            self._log(f"[{acc['uid']}] LOGIN OK (HOLD) - will close on STOP")
        except Exception:
            pass

    def _fb_login_only_worker(self, item_id: str, acc: dict, win_pos: str = "", win_size: str = "") -> None:
        if self.stop_event.is_set():
            return
        profile_id = None
        try:
            self._set_fb_status(item_id, "CREATE...")
            ok_c = False
            data_c = {}
            msg_c = ""
            with self.create_lock:
                for attempt in range(3):
                    ok_c, data_c, msg_c = create_profile(acc["uid"], acc["proxy"], SCOOPZ_URL)
                    if ok_c:
                        break
                    wait_s = 5 + attempt * 3
                    self._set_fb_status(item_id, f"CREATE RETRY {attempt+1}/3")
                    self._log(f"[{acc['uid']}] CREATE ERR: {msg_c} | retry in {wait_s}s")
                    time.sleep(wait_s)
            if not ok_c:
                self._set_fb_status(item_id, f"CREATE ERR: {msg_c}")
                return

            if isinstance(data_c, dict):
                profile_id = (data_c.get("data") or {}).get("id") or data_c.get("id") or data_c.get("profile_id")
            if not profile_id:
                self._set_fb_status(item_id, "NO PROFILE ID")
                return
            self._set_fb_profile_info(item_id, profile_id=profile_id)
            self._remember_profile_path(profile_id, data_c)
            self.created_profiles.add(profile_id)

            self._set_fb_status(item_id, "START...")
            ok_s, data_s, msg_s = start_profile(profile_id, win_pos=win_pos, win_size=win_size)
            if not ok_s:
                if self._is_proxy_error(msg_s):
                    try:
                        close_profile(profile_id, 3)
                        delete_profile(profile_id, 10)
                    except Exception:
                        pass
                    try:
                        self.created_profiles.discard(profile_id)
                    except Exception:
                        pass
                    self._delete_profile_path(profile_id)
                    new_id, new_data_s, _err = self._retry_start_profile_with_new_proxy(
                        acc,
                        item_id,
                        "fb",
                        self.fb_tree,
                        lambda s: self._set_fb_status(item_id, s),
                        win_pos=win_pos,
                        win_size=win_size,
                        created_set="created_profiles",
                    )
                    if new_id and new_data_s:
                        profile_id = new_id
                        data_s = new_data_s
                        ok_s = True
                    else:
                        self._set_fb_status(item_id, f"START ERR: {msg_s}")
                        return
                else:
                    self._set_fb_status(item_id, f"START ERR: {msg_s}")
                    return

            with self.active_lock:
                self.active_profiles[item_id] = profile_id
            driver_path, remote = extract_driver_info(data_s)
            status = "STARTED" if driver_path and remote else "STARTED (no debug)"
            self._set_fb_status(item_id, status)

            if not (driver_path and remote):
                return

            self._set_fb_status(item_id, "LOGIN...")
            ok_login, err_login = login_scoopz(
                driver_path,
                remote,
                acc["uid"],
                acc["pass"],
                "",
                max_retries=3,
                keep_browser=True,
            )
            if not ok_login:
                status = self._format_login_error(err_login)
                self._set_fb_status(item_id, status)
                return

            self._set_fb_status(item_id, "LOGIN OK (HOLD)")
            self._log(f"[{acc['uid']}] FB LOGIN OK (HOLD) - will close on STOP")
        except Exception:
            pass

    def _fb_follow_only_worker(
        self,
        item_id: str,
        acc: dict,
        win_pos: str = "",
        win_size: str = "",
        perform_creator_fund_check: bool = False,
        force_creator_fund_check: bool = False,
        stats_source: str = "",
        stats_email: str = "",
        skip_follow_fetch: bool = False,
    ) -> None:
        if self.stop_event.is_set():
            return
        profile_id = None
        try:
            if stats_source and stats_email:
                self._set_stats_row_checking(stats_source, stats_email, True, "CREATE...")
            self._set_fb_status(item_id, "CREATE...")
            ok_c = False
            data_c = {}
            msg_c = ""
            with self.create_lock:
                for attempt in range(3):
                    ok_c, data_c, msg_c = create_profile(acc["uid"], acc["proxy"], SCOOPZ_URL)
                    if ok_c:
                        break
                    wait_s = 5 + attempt * 3
                    self._set_fb_status(item_id, f"CREATE RETRY {attempt+1}/3")
                    self._log(f"[{acc['uid']}] CREATE ERR: {msg_c} | retry in {wait_s}s")
                    time.sleep(wait_s)
            if not ok_c:
                self._set_fb_status(item_id, f"CREATE ERR: {msg_c}")
                return

            if isinstance(data_c, dict):
                profile_id = (data_c.get("data") or {}).get("id") or data_c.get("id") or data_c.get("profile_id")
            if not profile_id:
                self._set_fb_status(item_id, "NO PROFILE ID")
                return
            self._set_fb_profile_info(item_id, profile_id=profile_id)
            self._remember_profile_path(profile_id, data_c)
            self.created_profiles.add(profile_id)

            self._set_fb_status(item_id, "START...")
            ok_s, data_s, msg_s = start_profile(profile_id, win_pos=win_pos, win_size=win_size)
            if not ok_s:
                if self._is_proxy_error(msg_s):
                    try:
                        close_profile(profile_id, 3)
                        delete_profile(profile_id, 10)
                    except Exception:
                        pass
                    try:
                        self.created_profiles.discard(profile_id)
                    except Exception:
                        pass
                    self._delete_profile_path(profile_id)
                    new_id, new_data_s, _err = self._retry_start_profile_with_new_proxy(
                        acc,
                        item_id,
                        "fb",
                        self.fb_tree,
                        lambda s: self._set_fb_status(item_id, s),
                        win_pos=win_pos,
                        win_size=win_size,
                        created_set="created_profiles",
                    )
                    if new_id and new_data_s:
                        profile_id = new_id
                        data_s = new_data_s
                        ok_s = True
                    else:
                        self._set_fb_status(item_id, f"START ERR: {msg_s}")
                        return
                else:
                    self._set_fb_status(item_id, f"START ERR: {msg_s}")
                    return

            with self.active_lock:
                self.active_profiles[item_id] = profile_id
            driver_path, remote = extract_driver_info(data_s)
            status = "STARTED" if driver_path and remote else "STARTED (no debug)"
            self._set_fb_status(item_id, status)
            if not (driver_path and remote):
                return

            if stats_source and stats_email:
                self._set_stats_row_checking(stats_source, stats_email, True, "LOGIN...")
            self._set_fb_status(item_id, "LOGIN...")
            ok_login, err_login = login_scoopz(
                driver_path,
                remote,
                acc["uid"],
                acc["pass"],
                "",
                max_retries=3,
                keep_browser=True,
            )
            if not ok_login:
                status = self._format_login_error(err_login)
                self._set_fb_status(item_id, status)
                return

            if skip_follow_fetch:
                if perform_creator_fund_check:
                    if stats_source and stats_email:
                        self._set_stats_row_checking(stats_source, stats_email, True, "CHECK APPLY...")
                    self._maybe_check_creator_fund(
                        driver_path,
                        remote,
                        acc,
                        acc.get("followers"),
                        "FB",
                        force_check=force_creator_fund_check,
                    )
                self._set_fb_status(item_id, "CHECK DONE")
                return

            followers = None
            profile_url = ""
            posts = None
            for attempt in range(3):
                followers, profile_url, posts = fetch_followers(driver_path, remote, self._log, acc.get("uid", ""))
                if followers is not None or posts is not None:
                    break
                time.sleep(2 + attempt)

            fnum = self._followers_to_int(followers if followers is not None else acc.get("followers"))
            if stats_source and stats_email:
                self._set_stats_row_checking(
                    stats_source,
                    stats_email,
                    True,
                    "ELIGIBLE" if fnum >= 1000 else "NOT_ELIGIBLE",
                )

            if followers is not None or posts is not None:
                self._set_fb_profile_info(item_id, profile_url, followers, posts)
                if perform_creator_fund_check:
                    if stats_source and stats_email:
                        self._set_stats_row_checking(stats_source, stats_email, True, "CHECK APPLY...")
                    self._maybe_check_creator_fund(
                        driver_path,
                        remote,
                        acc,
                        followers,
                        "FB",
                        force_check=force_creator_fund_check,
                    )
                self._set_fb_status(item_id, "FOLLOW OK")
            else:
                if perform_creator_fund_check:
                    if stats_source and stats_email:
                        self._set_stats_row_checking(stats_source, stats_email, True, "CHECK APPLY...")
                    self._maybe_check_creator_fund(
                        driver_path,
                        remote,
                        acc,
                        followers,
                        "FB",
                        force_check=force_creator_fund_check,
                    )
                self._set_fb_status(item_id, "FOLLOW ERR")
        finally:
            if profile_id:
                self._cleanup_profile_session(item_id, profile_id)
            if stats_source and stats_email:
                self._set_stats_row_checking(stats_source, stats_email, False)

    def _upload_only_worker(self, item_id: str, acc: dict) -> None:
        if self.stop_event.is_set():
            return
        profile_id = None
        try:
            driver_path, remote, profile_id = self._ensure_logged_in(item_id, acc)
            if not driver_path or not remote:
                return
            if SKIP_DOWNLOAD_UPLOAD:
                try:
                    followers = None
                    profile_url = ""
                    posts = None
                    for attempt in range(3):
                        followers, profile_url, posts = fetch_followers(driver_path, remote, self._log, acc.get("uid", ""))
                        if followers is not None or posts is not None:
                            break
                        time.sleep(2 + attempt)
                    if followers is not None or posts is not None:
                        if followers is not None:
                            self._log(f"[{acc['uid']}] FOLLOWERS: {followers}")
                        if posts is not None:
                            self._log(f"[{acc['uid']}] POSTS: {posts}")
                        self._set_profile_info(item_id, profile_url, followers, posts)
                        self._set_status(item_id, "FOLLOW OK")
                    else:
                        self._set_status(item_id, "FOLLOW ERR")
                        self._log(f"[{acc['uid']}] FOLLOW ERR")
                except Exception as e:
                    self._log(f"[{acc['uid']}] FOLLOW ERR: {e}")
                return
            download_started = threading.Event()

            def _idle_watchdog():
                if download_started.wait(timeout=30):
                    return
                if self.stop_event.is_set():
                    return
                ok_next, row = get_next_unuploaded(acc["uid"])
                if not ok_next:
                    self._set_status(item_id, "NO VIDEO")
                    self._log(f"[{acc['uid']}] NO VIDEO: {row.get('msg', 'No URL in CSV')}")
                    return
                self._set_status(item_id, "LOGIN OK (IDLE)")
                self._log(f"[{acc['uid']}] LOGIN OK (IDLE) - no download started")

            threading.Thread(target=_idle_watchdog, daemon=True).start()
            success_count = 0
            safety_guard = 0
            max_videos = 1
            while success_count < max_videos:
                if self.stop_event.is_set():
                    return
                safety_guard += 1
                if safety_guard > max_videos * 5:
                    self._log(f"[{acc['uid']}] Too many skips, stop loop")
                    return

                ok_next, row = get_next_unuploaded(acc["uid"])
                if not ok_next:
                    self._set_status(item_id, "NO VIDEO")
                    self._log(f"[{acc['uid']}] NO VIDEO: {row.get('msg', 'No URL in CSV')}")
                    return
                row_url = (row.get("url") or "").strip()
                row_id = (row.get("video_id") or "").strip()
                if not row_url:
                    if row_id.startswith("http"):
                        row_url = row_id
                    elif row_id:
                        row_url = f"https://www.youtube.com/shorts/{row_id}"
                if not row_url:
                    self._log(f"[{acc['uid']}] No URL in CSV row")
                    return

                self._set_status(item_id, "DOWNLOAD...")
                download_started.set()
                self._log(f"[{acc['uid']}] NEXT VIDEO: {row_id} | {row_url}")
                ok_dl, path_or_err, vid_id, title = download_one(
                    acc["uid"],
                    row_url,
                    self._log_progress,
                    cookie_path=COOKIES_FILE,
                    fallback_cookie_path=COOKIES_FILE_FALLBACK,
                    timeout_s=30,
                )
                mark_id = vid_id or row_id
                if not ok_dl:
                    err_text = str(path_or_err)
                    lower = err_text.lower()
                    is_skipped = (
                        "video skipped" in lower
                        or "video unavailable" in lower
                        or "video is unavailable" in lower
                        or "private video" in lower
                        or "watch video on youtube" in lower
                        or "sign in if you've been granted access" in lower
                        or "sign in to confirm your age" in lower
                        or "age-restricted" in lower
                        or "age restricted" in lower
                    )
                    if is_skipped:
                        self._log(f"[{acc['uid']}] VIDEO UNAVAILABLE - AUTO SKIP")
                        try:
                            mark_uploaded(acc["uid"], mark_id)
                        except Exception:
                            pass
                        continue
                    self._log(f"[{acc['uid']}] DOWNLOAD ERR: {err_text}")
                    if "timeout" in lower or "timed out" in lower:
                        self._record_failed(item_id, acc, f"DOWNLOAD ERR: {err_text}")
                    return

                caption = self._build_caption(title, 1000)
                token = self._enqueue_upload_turn()
                if not self._wait_upload_turn(token):
                    return
                try:
                    ok_p = False
                    drv = None
                    up_status = ""
                    up_msg = ""
                    retry_reopen = {"caption_error", "dialog_error", "timeout", "unexpected_error", "error"}
                    for attempt in range(3):
                        ok_p, drv, up_status, up_msg = upload_prepare(
                            driver_path,
                            remote,
                            path_or_err,
                            caption,
                            lambda: self.stop_event.is_set(),
                            self._log,
                            acc.get("uid", ""),
                            max_total_s=360,
                            file_dialog_semaphore=None,
                        )
                        if ok_p:
                            break
                        if up_status in retry_reopen and attempt < 2:
                            if up_status == "caption_error":
                                caption = self._next_caption_after_error(caption, 1000)
                                self._log(f"[{acc['uid']}] Caption error -> switched fallback caption for retry")
                            wait_s = 2 + attempt
                            self._log(f"[{acc['uid']}] Upload page retry {attempt+1}/2 in {wait_s}s (status={up_status})")
                            time.sleep(wait_s)
                            continue
                        break
                finally:
                    self._release_upload_turn(token)
                if not ok_p:
                    status_text = f"UPLOAD LOI: {up_status}" if up_status in ("select_not_found", "select_click_error") else f"UPLOAD ERR: {up_msg or up_status}"
                    self._set_status(item_id, status_text)
                    self._log(f"[{acc['uid']}] {status_text}")
                    if up_status in ("select_not_found", "select_click_error"):
                        self._record_failed(item_id, acc, f"UPLOAD {up_status}")
                    elif up_status == "timeout":
                        self._record_failed(item_id, acc, f"UPLOAD ERR: {up_msg or up_status}")
                    return

                self._set_status(item_id, "POSTING...")
                st, msg, purl, foll, posts = upload_post_async(
                    drv,
                    self._log,
                    acc_email=acc.get("uid", ""),
                    max_total_s=420,
                    post_button_semaphore=self.post_button_semaphore,
                )
                if st == "success":
                    try:
                        mark_uploaded(acc["uid"], mark_id)
                    except Exception:
                        pass
                    self._set_profile_info(item_id, purl, foll, posts)
                    self._set_status(item_id, "UPLOAD OK")
                    self._log(f"[{acc['uid']}] UPLOAD OK")
                    self._delete_uploaded_video(path_or_err, acc["uid"])
                    success_count += 1
                    try:
                        followers = None
                        profile_url = ""
                        posts = None
                        for attempt in range(3):
                            followers, profile_url, posts = fetch_followers(driver_path, remote, self._log, acc.get("uid", ""))
                            if followers is not None or posts is not None:
                                break
                            time.sleep(2 + attempt)
                        if followers is not None or posts is not None:
                            if followers is not None:
                                self._log(f"[{acc['uid']}] FOLLOWERS: {followers}")
                            if posts is not None:
                                self._log(f"[{acc['uid']}] POSTS: {posts}")
                            self._set_profile_info(item_id, profile_url, followers, posts)
                        else:
                            self._log(f"[{acc['uid']}] FOLLOW ERR: empty")
                    except Exception as e:
                        self._log(f"[{acc['uid']}] FOLLOW ERR: {e}")
                    time.sleep(10.0)
                else:
                    err_text = msg or st
                    status_text = "UPLOAD LOI" if "Select video not found" in (err_text or "") else f"UPLOAD ERR: {err_text}"
                    self._set_status(item_id, status_text)
                    self._log(f"[{acc['uid']}] UPLOAD ERR: {err_text}")
                    if st == "timeout":
                        self._record_failed(item_id, acc, f"POST ERR: {err_text}")
                    return
        finally:
            if profile_id:
                self._cleanup_profile_session(item_id, profile_id)

    def _follow_only_worker(
        self,
        item_id: str,
        acc: dict,
        perform_creator_fund_check: bool = False,
        force_creator_fund_check: bool = False,
        stats_source: str = "",
        stats_email: str = "",
        win_pos: str | None = None,
        win_size: str | None = None,
        skip_follow_fetch: bool = False,
    ) -> None:
        if self.stop_event.is_set():
            return
        profile_id = None
        try:
            if stats_source and stats_email:
                self._set_stats_row_checking(stats_source, stats_email, True, "LOGIN...")
            driver_path, remote, profile_id = self._ensure_logged_in(item_id, acc, win_pos=win_pos, win_size=win_size)
            if not driver_path or not remote:
                return

            if skip_follow_fetch:
                if perform_creator_fund_check:
                    if stats_source and stats_email:
                        self._set_stats_row_checking(stats_source, stats_email, True, "CHECK APPLY...")
                    self._maybe_check_creator_fund(
                        driver_path,
                        remote,
                        acc,
                        acc.get("followers"),
                        "YTB",
                        force_check=force_creator_fund_check,
                    )
                self._set_status(item_id, "CHECK DONE")
                return

            followers = None
            profile_url = ""
            posts = None
            for attempt in range(3):
                followers, profile_url, posts = fetch_followers(driver_path, remote, self._log, acc.get("uid", ""))
                if followers is not None or posts is not None:
                    break
                time.sleep(2 + attempt)

            fnum = self._followers_to_int(followers if followers is not None else acc.get("followers"))
            if stats_source and stats_email:
                self._set_stats_row_checking(
                    stats_source,
                    stats_email,
                    True,
                    "ELIGIBLE" if fnum >= 1000 else "NOT_ELIGIBLE",
                )

            if followers is not None or posts is not None:
                if followers is not None:
                    self._log(f"[{acc['uid']}] FOLLOWERS: {followers}")
                if posts is not None:
                    self._log(f"[{acc['uid']}] POSTS: {posts}")
                self._set_profile_info(item_id, profile_url, followers, posts)
                if perform_creator_fund_check:
                    if stats_source and stats_email:
                        self._set_stats_row_checking(stats_source, stats_email, True, "CHECK APPLY...")
                    self._maybe_check_creator_fund(
                        driver_path,
                        remote,
                        acc,
                        followers,
                        "YTB",
                        force_check=force_creator_fund_check,
                    )
                self._set_status(item_id, "FOLLOW OK")
            else:
                if perform_creator_fund_check:
                    if stats_source and stats_email:
                        self._set_stats_row_checking(stats_source, stats_email, True, "CHECK APPLY...")
                    self._maybe_check_creator_fund(
                        driver_path,
                        remote,
                        acc,
                        followers,
                        "YTB",
                        force_check=force_creator_fund_check,
                    )
                self._set_status(item_id, "FOLLOW ERR")
                self._log(f"[{acc['uid']}] FOLLOW ERR")
        finally:
            if profile_id:
                self._cleanup_profile_session(item_id, profile_id)
            if stats_source and stats_email:
                self._set_stats_row_checking(stats_source, stats_email, False)

    def _profile_open_worker(self, item_id: str, acc: dict, win_pos: str, win_size: str) -> None:
        sem = self.profile_semaphore
        if sem:
            sem.acquire()
        profile_id = None
        started = False
        email = (acc.get("uid") or "").strip()
        try:
            if self.stop_event.is_set():
                return
            started = True
            yt_url = (acc.get("youtube") or "").strip()
            if not yt_url:
                self._set_profile_status(item_id, "YTB ERR: Thieu link")
                self._log(f"[{acc.get('uid','')}] YTB ERR: Thieu link YouTube")
                self._record_profile_failed(item_id, acc, "YTB ERR: Thieu link")
                return
            self._ensure_video_folder(acc.get("uid", ""))
            cached = self._load_profile_assets(acc.get("uid", ""))
            name = (cached.get("name") or "").strip()
            username = (cached.get("username") or "").strip()
            avatar_path = (cached.get("avatar_path") or "").strip()
            if name and username and avatar_path and os.path.exists(avatar_path):
                self._set_profile_status(item_id, "YTB CACHED")
                self._log(f"[{acc['uid']}] YTB CACHED")
            else:
                self._set_profile_status(item_id, "YTB FETCH...")
                self._log(f"[{acc['uid']}] YTB FETCH START: {yt_url}")
                fetched_name, fetched_username, avatar_path = fetch_youtube_profile_assets_local(yt_url, self._log)
                if not avatar_path:
                    self._set_profile_status(item_id, "YTB ERR: Empty")
                    self._log(f"[{acc['uid']}] YTB ERR: Empty")
                    self._record_profile_failed(item_id, acc, "YTB ERR: Empty")
                    return
                name = fetched_name
                username = fetched_username
                self._save_profile_assets(acc["uid"], fetched_name, fetched_username, avatar_path)
                self._set_profile_status(item_id, "YTB OK")
                self._log(f"[{acc['uid']}] YTB OK")
            self._log(f"[{acc['uid']}] START PROFILE")
            self._set_profile_status(item_id, "CREATE...")
            ok_c = False
            data_c = {}
            msg_c = ""
            with self.create_lock:
                for attempt in range(3):
                    ok_c, data_c, msg_c = create_profile(acc["uid"], acc["proxy"], SCOOPZ_URL)
                    if ok_c:
                        break
                    wait_s = 5 + attempt * 3
                    self._set_profile_status(item_id, f"CREATE RETRY {attempt+1}/3")
                    self._log(f"[{acc['uid']}] CREATE ERR: {msg_c} | retry in {wait_s}s")
                    time.sleep(wait_s)
            if not ok_c:
                self._set_profile_status(item_id, f"CREATE ERR: {msg_c}")
                self._log(f"[{acc['uid']}] CREATE ERR: {msg_c}")
                self._record_profile_failed(item_id, acc, f"CREATE ERR: {msg_c}")
                return

            if isinstance(data_c, dict):
                profile_id = (data_c.get("data") or {}).get("id") or data_c.get("id") or data_c.get("profile_id")
            if not profile_id:
                self._set_profile_status(item_id, "NO PROFILE ID")
                self._log(f"[{acc['uid']}] NO PROFILE ID")
                self._record_profile_failed(item_id, acc, "NO PROFILE ID")
                return
            self._remember_profile_path(profile_id, data_c)
            self.profile_created_profiles.add(profile_id)

            self._set_profile_status(item_id, "START...")
            ok_s, data_s, msg_s = start_profile(profile_id, win_pos=win_pos, win_size=win_size)
            if not ok_s:
                if self._is_proxy_error(msg_s):
                    try:
                        close_profile(profile_id, 3)
                        delete_profile(profile_id, 10)
                    except Exception:
                        pass
                    try:
                        self.profile_created_profiles.discard(profile_id)
                    except Exception:
                        pass
                    self._delete_profile_path(profile_id)
                    new_id, new_data_s, _err = self._retry_start_profile_with_new_proxy(
                        acc,
                        item_id,
                        "profile",
                        self.profile_tree,
                        lambda s: self._set_profile_status(item_id, s),
                        win_pos=win_pos,
                        win_size=win_size,
                        created_set="profile_created_profiles",
                    )
                    if new_id and new_data_s:
                        profile_id = new_id
                        data_s = new_data_s
                        ok_s = True
                    else:
                        self._set_profile_status(item_id, f"START ERR: {msg_s}")
                        self._log(f"[{acc['uid']}] START ERR: {msg_s}")
                        self._record_profile_failed(item_id, acc, f"START ERR: {msg_s}")
                        return
                else:
                    self._set_profile_status(item_id, f"START ERR: {msg_s}")
                    self._log(f"[{acc['uid']}] START ERR: {msg_s}")
                    self._record_profile_failed(item_id, acc, f"START ERR: {msg_s}")
                    return

            driver_path, remote = extract_driver_info(data_s)
            status = "STARTED" if driver_path and remote else "STARTED (no debug)"
            self._set_profile_status(item_id, status)
            self._log(f"[{acc['uid']}] START OK")

            if not driver_path or not remote:
                self._record_profile_failed(item_id, acc, "STARTED (no debug)")
                return

            self._set_profile_status(item_id, "LOGIN...")
            self._log(f"[{acc['uid']}] LOGIN START")
            ok_login, err_login = login_scoopz(
                driver_path,
                remote,
                acc["uid"],
                acc["pass"],
                "",
                max_retries=3,
                keep_browser=True,
            )
            if not ok_login:
                status = self._format_login_error(err_login)
                self._set_profile_status(item_id, status)
                self._log(f"[{acc['uid']}] {status}")
                self._record_profile_failed(item_id, acc, status)
                return

            if self.stop_event.is_set():
                return

            self._set_profile_status(item_id, "LOGIN OK")
            self._log(f"[{acc['uid']}] LOGIN OK (PROFILE)")

            self._set_profile_status(item_id, "OPEN PROFILE...")
            ok_p = False
            err_p = ""
            for attempt in range(1, 4):
                if self.stop_event.is_set():
                    return
                self._set_profile_status(item_id, f"OPEN PROFILE... ({attempt}/3)")
                ok_p, err_p = open_profile_in_scoopz(
                    driver_path,
                    remote,
                    avatar_path,
                    name,
                    username,
                    logger=self._log,
                    max_retries=3,
                )
                if ok_p:
                    break
                retryable = (
                    "cannot connect to chrome" in (err_p or "").lower()
                    or "profile link not found" in (err_p or "").lower()
                    or "profile page load timeout" in (err_p or "").lower()
                )
                if not retryable:
                    break
                wait_s = 2 + attempt * 2
                self._log(f"[{acc['uid']}] PROFILE RETRY {attempt}/3 in {wait_s}s: {err_p}")
                time.sleep(wait_s)
            if not ok_p:
                self._set_profile_status(item_id, f"PROFILE ERR: {err_p}")
                self._log(f"[{acc['uid']}] PROFILE ERR: {err_p}")
                self._record_profile_failed(item_id, acc, f"PROFILE ERR: {err_p}")
                return
            self._set_profile_status(item_id, "PROFILE OPENED")
            self._log(f"[{acc['uid']}] PROFILE OPENED")

            yt_url = (acc.get("youtube") or "").strip()
            if not yt_url:
                self._set_profile_status(item_id, "YTB ERR: Thieu link")
                self._log(f"[{acc.get('uid','')}] YTB ERR: Thieu link YouTube")
                self._record_profile_failed(item_id, acc, "YTB ERR: Thieu link")
                return
            self._ensure_video_folder(acc.get("uid", ""))
            cached = self._load_profile_assets(acc.get("uid", ""))
            name = (cached.get("name") or "").strip()
            username = (cached.get("username") or "").strip()
            avatar_path = (cached.get("avatar_path") or "").strip()
            if name and username and avatar_path and os.path.exists(avatar_path):
                self._set_profile_status(item_id, "YTB CACHED")
                self._log(f"[{acc['uid']}] YTB CACHED")
            else:
                self._set_profile_status(item_id, "YTB FETCH...")
                self._log(f"[{acc['uid']}] YTB FETCH START: {yt_url}")
                fetched_name, fetched_username, avatar_path = fetch_youtube_profile_assets_local(yt_url, self._log)
                if not avatar_path:
                    self._set_profile_status(item_id, "YTB ERR: Empty")
                    self._log(f"[{acc['uid']}] YTB ERR: Empty")
                    self._record_profile_failed(item_id, acc, "YTB ERR: Empty")
                    return
                name = fetched_name
                username = fetched_username
                self._save_profile_assets(acc["uid"], fetched_name, fetched_username, avatar_path)
                self._set_profile_status(item_id, "YTB OK")
                self._log(f"[{acc['uid']}] YTB OK")

            with self.profile_active_drivers_lock:
                self.profile_active_drivers[item_id] = {
                    "profile_id": profile_id,
                    "driver_path": driver_path,
                    "remote": remote,
                    "close_func": lambda: (close_profile(profile_id, 3), delete_profile(profile_id, 10)),
                }
            self._set_profile_status(item_id, "DONE")
            self._log(f"[{acc['uid']}] PROFILE DONE")
        finally:
            if started:
                self._mark_run_done("profile", email)
            if profile_id:
                try:
                    close_profile(profile_id, 3)
                except Exception:
                    pass
                try:
                    delete_profile(profile_id, 10)
                except Exception:
                    pass
                with self.profile_active_drivers_lock:
                    self.profile_active_drivers.pop(item_id, None)
                try:
                    self.profile_created_profiles.discard(profile_id)
                except Exception:
                    pass
                self._delete_profile_path(profile_id)
                self._track_profile_cleanup()
            if sem:
                sem.release()

    def stop_jobs(self) -> None:
        self.stop_event.set()
        self._runtime_pause()
        self._set_busy(False)
        self._fixed_threads = None
        self._reset_cycle_count()
        self._stop_next_cycle_countdown()
        try:
            self._resume_pending["upload"] = self._collect_pending_emails(self.tree, {"UPLOAD OK", "DONE"})
            self._resume_pending["profile"] = self._collect_pending_emails(self.profile_tree, {"DONE"})
            self._resume_pending["fb"] = self._collect_pending_emails(self.fb_tree, {"UPLOAD OK", "DONE"})
            self._resume_pending["fb_profile"] = self._collect_pending_emails(self.fb_profile_tree, {"DONE"})
        except Exception:
            pass
        if self._repeat_after_id:
            try:
                self.root.after_cancel(self._repeat_after_id)
            except Exception:
                pass
            self._repeat_after_id = None
        # Keep failed account history and logs for review
        if self.executor is not None:
            try:
                self.executor.shutdown(wait=False, cancel_futures=True)
            except Exception:
                pass
            self.executor = None
        
        # Clean up active drivers
        with self.active_drivers_lock:
            for driver_info in self.active_drivers.values():
                try:
                    if driver_info and "close_func" in driver_info:
                        driver_info["close_func"]()
                except Exception:
                    pass
            self.active_drivers.clear()
        with self.profile_active_drivers_lock:
            for driver_info in self.profile_active_drivers.values():
                try:
                    if driver_info and "close_func" in driver_info:
                        driver_info["close_func"]()
                except Exception:
                    pass
            self.profile_active_drivers.clear()
        
        with self.active_lock:
            ids = list(self.active_profiles.values())
            self.active_profiles.clear()
        ids.extend(list(self.created_profiles))
        self.created_profiles.clear()
        ids.extend(list(self.profile_created_profiles))
        self.profile_created_profiles.clear()
        ids = list({pid for pid in ids if pid})
        for pid in ids:
            threading.Thread(
                target=lambda p=pid: self._cleanup_profile_session(None, p),
                daemon=True,
            ).start()
        try:
            self._cleanup_gpm_root(force=True)
        except Exception:
            pass

    def reload_app(self) -> None:
        if getattr(self, "_reloading", False):
            return
        self._reloading = True
        try:
            try:
                self._log("[RELOAD] Restarting app process...")
            except Exception:
                pass
            self.stop_jobs()

            # Launch a new process first, then close current one.
            if getattr(sys, "frozen", False):
                cmd = [sys.executable]
                cwd = os.path.dirname(sys.executable) or _THIS_DIR
            else:
                cmd = [sys.executable, os.path.abspath(__file__)]
                cwd = _THIS_DIR

            subprocess.Popen(cmd, cwd=cwd)

            try:
                self._log(f"[RELOAD] Spawned new process: {' '.join(cmd)}")
            except Exception:
                pass
            try:
                self.root.after(200, self.root.destroy)
            except Exception:
                self.root.destroy()
        except Exception as e:
            try:
                self._log(f"[RELOAD] ERR: {e}")
            except Exception:
                pass
            try:
                messagebox.showerror("Reload Error", f"Restart failed: {e}")
            except Exception:
                pass
        finally:
            self._reloading = False

    def _worker_one(self, item_id: str, acc: dict, win_pos: str, win_size: str, max_videos: int) -> None:
        email = (acc.get("uid") or "").strip()
        started = False
        try:
            if self.stop_event.is_set():
                return
            started = True
    
            def _restart_profile() -> tuple:
                try:
                    self._set_status(item_id, "RESTART...")
                    close_profile(profile_id, 3)
                    ok_s, data_s, msg_s = start_profile(profile_id, win_pos=win_pos, win_size=win_size)
                    if not ok_s:
                        self._set_status(item_id, f"RESTART ERR: {msg_s}")
                        self._log(f"[{acc['uid']}] RESTART ERR: {msg_s}")
                        return None, None
                    drv_path, dbg_addr = extract_driver_info(data_s)
                    if not drv_path or not dbg_addr:
                        self._set_status(item_id, "RESTART NO DEBUG", profile_id=profile_id)
                        return None, None
                    self._set_status(item_id, "RELOGIN...")
                    ok_login, err_login = login_scoopz(
                        drv_path,
                        dbg_addr,
                        acc["uid"],
                        acc["pass"],
                        "",
                        max_retries=2,
                        keep_browser=True,
                    )
                    if not ok_login:
                        status = self._format_login_error(err_login)
                        if status == "SAI PASS":
                            self._set_status(item_id, status)
                            self._log(f"[{acc['uid']}] {status}")
                        else:
                            self._set_status(item_id, f"RELOGIN ERR: {err_login}")
                            self._log(f"[{acc['uid']}] RELOGIN ERR: {err_login}")
                        return None, None
                    self._set_status(item_id, "RESTART OK", profile_id=profile_id)
                    return drv_path, dbg_addr
                except Exception as e:
                    self._set_status(item_id, f"RESTART ERR: {e}")
                    self._log(f"[{acc['uid']}] RESTART ERR: {e}")
                    return None, None
    
            def _extract_video_id(text: str) -> str:
                val = (text or "").strip()
                if not val:
                    return ""
                if "shorts/" in val:
                    return val.split("shorts/", 1)[1].split("?", 1)[0].strip("/")
                if "watch?v=" in val:
                    return val.split("watch?v=", 1)[1].split("&", 1)[0]
                return ""
    
            self._log(f"[{acc['uid']}] START")
            self._set_status(item_id, "CREATE...")
            ok_c = False
            data_c = {}
            msg_c = ""
            with self.create_lock:
                for attempt in range(3):
                    ok_c, data_c, msg_c = create_profile(acc["uid"], acc["proxy"], SCOOPZ_URL)
                    if ok_c:
                        break
                    wait_s = 5 + attempt * 3
                    self._set_status(item_id, f"CREATE RETRY {attempt+1}/3")
                    self._log(f"[{acc['uid']}] CREATE ERR: {msg_c} | retry in {wait_s}s")
                    time.sleep(wait_s)
            if not ok_c:
                self._set_status(item_id, f"CREATE ERR: {msg_c}")
                self._log(f"[{acc['uid']}] CREATE ERR: {msg_c}")
                self._record_failed(item_id, acc, f"CREATE ERR: {msg_c}")
                return
    
            profile_id = None
            if isinstance(data_c, dict):
                profile_id = (data_c.get("data") or {}).get("id") or data_c.get("id") or data_c.get("profile_id")
            if not profile_id:
                self._set_status(item_id, "NO PROFILE ID")
                self._log(f"[{acc['uid']}] NO PROFILE ID")
                self._record_failed(item_id, acc, "NO PROFILE ID")
                return
            self._remember_profile_path(profile_id, data_c)
            self.created_profiles.add(profile_id)
    
            self._set_status(item_id, "START...", profile_id=profile_id)
            ok_s, data_s, msg_s = start_profile(profile_id, win_pos=win_pos, win_size=win_size)
            if not ok_s:
                if self._is_proxy_error(msg_s):
                    swapped = self._replace_proxy_for_account(acc, item_id, "upload", self.tree)
                    if swapped:
                        self._set_status(item_id, f"START ERR: {msg_s} (proxy replaced)")
                        self._log(f"[{acc['uid']}] START ERR: {msg_s} (proxy replaced)")
                        self._record_failed(item_id, acc, f"START ERR: {msg_s} (proxy replaced)")
                        return
                self._set_status(item_id, f"START ERR: {msg_s}")
                self._log(f"[{acc['uid']}] START ERR: {msg_s}")
                self._record_failed(item_id, acc, f"START ERR: {msg_s}")
                return
    
            with self.active_lock:
                self.active_profiles[item_id] = profile_id
            driver_path, remote = extract_driver_info(data_s)
            status = "STARTED" if driver_path and remote else "STARTED (no debug)"
            self._set_status(item_id, status, profile_id=profile_id)
            self._log(f"[{acc['uid']}] START OK")
    
            if driver_path and remote:
                self._set_status(item_id, "LOGIN...")
                self._log(f"[{acc['uid']}] LOGIN START")
                ok_login, err_login = login_scoopz(
                    driver_path,
                    remote,
                    acc["uid"],
                    acc["pass"],
                    "",
                    max_retries=3,
                    keep_browser=True,
                )
                if ok_login:
                    self._set_status(item_id, "LOGIN OK")
                    self._log(f"[{acc['uid']}] LOGIN OK")
                    if SKIP_DOWNLOAD_UPLOAD:
                        try:
                            followers = None
                            profile_url = ""
                            posts = None
                            for attempt in range(3):
                                followers, profile_url, posts = fetch_followers(driver_path, remote, self._log, acc.get("uid", ""))
                                if followers is not None or posts is not None:
                                    break
                                time.sleep(2 + attempt)
                            if followers is not None or posts is not None:
                                if followers is not None:
                                    self._log(f"[{acc['uid']}] FOLLOWERS: {followers}")
                                if posts is not None:
                                    self._log(f"[{acc['uid']}] POSTS: {posts}")
                                self._set_profile_info(item_id, profile_url, followers, posts)
                                self._set_status(item_id, "FOLLOW OK")
                            else:
                                self._set_status(item_id, "FOLLOW ERR")
                                self._log(f"[{acc['uid']}] FOLLOW ERR")
                        except Exception as e:
                            self._log(f"[{acc['uid']}] FOLLOW ERR: {e}")
                        return
                    download_started = threading.Event()
                    def _idle_watchdog():
                        if download_started.wait(timeout=30):
                            return
                        if self.stop_event.is_set():
                            return
                        ok_next, row = get_next_unuploaded(acc["uid"])
                        if not ok_next:
                            self._set_status(item_id, "NO VIDEO")
                            self._log(f"[{acc['uid']}] NO VIDEO: {row.get('msg', 'No URL in CSV')}")
                            return
                        self._set_status(item_id, "LOGIN OK (IDLE)")
                        self._log(f"[{acc['uid']}] LOGIN OK (IDLE) - no download started")
                    threading.Thread(target=_idle_watchdog, daemon=True).start()
                    success_count = 0
                    safety_guard = 0
                    restart_attempts = 0
                    while success_count < max_videos:
                        if self.stop_event.is_set():
                            break
                        safety_guard += 1
                        if safety_guard > max_videos * 5:
                            self._log(f"[{acc['uid']}] Too many skips, stop loop")
                            break
    
                        # Smart delay before next video
                        self.operation_delayer.delay_before_download(acc["uid"], self._log_progress)
    
                        ok_next, row = get_next_unuploaded(acc["uid"])
                        if not ok_next:
                            self._set_status(item_id, "NO VIDEO")
                            self._log(f"[{acc['uid']}] NO VIDEO: {row.get('msg', 'No URL in CSV')}")
                            break
                        row_url = (row.get("url") or "").strip()
                        row_id = (row.get("video_id") or "").strip()
                        if not row_url:
                            if row_id.startswith("http"):
                                row_url = row_id
                            elif row_id:
                                row_url = f"https://www.youtube.com/shorts/{row_id}"
                        if not row_url:
                            self._log(f"[{acc['uid']}] No URL in CSV row")
                            break
                        download_started.set()
                        self._log(f"[{acc['uid']}] NEXT VIDEO: {row_id} | {row_url}")
                        self._set_status(item_id, f"DOWNLOAD {success_count+1}/{max_videos}...")
                        self._log(f"[{acc['uid']}] DOWNLOAD START: {row_url}")
                        retry_dl = 0
                        path_or_err = ""
                        vid_id = ""
                        title = ""
                        ok_dl = False
                        skip_current = False
                        skip_account = False
                        download_start_ts = time.time()
                        download_timed_out = threading.Event()
                        watchdog = self._start_download_watchdog(
                            acc["uid"],
                            "DOWNLOAD",
                            max_seconds=600,
                            on_timeout=lambda _e: download_timed_out.set(),
                        )
                        while True:
                            ok_dl, path_or_err, vid_id, title = download_one(
                                acc["uid"],
                                row_url,
                                self._log_progress,
                                cookie_path=COOKIES_FILE,
                                fallback_cookie_path=COOKIES_FILE_FALLBACK,
                                timeout_s=30,
                            )
                            if ok_dl:
                                break
                            err_text = str(path_or_err)
                            lower = err_text.lower()
                            is_timeout = "timeout" in lower or "timed out" in lower
                            is_restricted = (
                                "video restricted" in lower
                                or "members-only" in lower
                                or "members only" in lower
                                or "join this channel" in lower
                                or "premium only" in lower
                                or "membership required" in lower
                            )
                            is_skipped = (
                                "video skipped" in lower
                                or "video unavailable" in lower
                                or "video is unavailable" in lower
                                or "private video" in lower
                                or "watch video on youtube" in lower
                                or "sign in if you've been granted access" in lower
                                or "sign in to confirm your age" in lower
                                or "age-restricted" in lower
                                or "age restricted" in lower
                            )
    
                            if is_skipped:
                                self._log(f"[{acc['uid']}] VIDEO UNAVAILABLE - AUTO SKIP")
                                # Mark as uploaded so it's skipped in future runs
                                mark_id = vid_id or row_id or _extract_video_id(row_url)
                                try:
                                    mark_uploaded(acc["uid"], mark_id)
                                except Exception as e:
                                    self._log(f"[{acc['uid']}] Could not mark video: {e}")
                                skip_current = True
                                break
                            if is_timeout:
                                if retry_dl < 1:
                                    retry_dl += 1
                                    self._log(f"[{acc['uid']}] DOWNLOAD TIMEOUT - RETRY 1/1")
                                    continue
                                self._log(f"[{acc['uid']}] DOWNLOAD TIMEOUT - SKIP ACCOUNT")
                                self._set_status(item_id, "DOWNLOAD TIMEOUT")
                                self._record_failed(item_id, acc, "DOWNLOAD TIMEOUT")
                                skip_account = True
                                break

                            if is_restricted and retry_dl < 1:
                                retry_dl += 1
                                self._log(f"[{acc['uid']}] DOWNLOAD RETRY (restricted): {row_url}")
                                self.operation_delayer.delay_on_error(acc['uid'], "restricted_video", self._log_progress)
                                continue
                            if is_restricted and retry_dl == 1:
                                # Relogin disabled - skip to next video
                                self._log(f"[{acc['uid']}] VIDEO RESTRICTED - SKIP (no relogin): {err_text}")
                                self.error_logger.log_download_error(acc['uid'], row_url, f"Video restricted - {err_text}")
                                break
                            if is_restricted and retry_dl >= 2:
                                self._log(f"[{acc['uid']}] VIDEO RESTRICTED - SKIP: {err_text}")
                                self.error_logger.log_download_error(acc['uid'], row_url, f"Video restricted - {err_text}")
                                break
                            break
    
                        mark_id = vid_id or row_id or _extract_video_id(row_url)
                        try:
                            watchdog.set()
                        except Exception:
                            pass
                        elapsed = int(time.time() - download_start_ts)
                        if ok_dl:
                            self._log(f"[{acc['uid']}] DOWNLOAD END OK after {elapsed}s")
                        else:
                            self._log(f"[{acc['uid']}] DOWNLOAD END ERR after {elapsed}s: {path_or_err}")
                        if download_timed_out.is_set():
                            self._set_status(item_id, "DOWNLOAD TIMEOUT")
                            try:
                                self.error_logger.log_download_error(acc["uid"], row_url, "DOWNLOAD TIMEOUT (watchdog 600s)")
                            except Exception:
                                pass
                            self._record_failed(item_id, acc, "DOWNLOAD TIMEOUT")
                            skip_account = True

                        if skip_account:
                            break
                        if skip_current:
                            continue
                        if ok_dl:
                            title = self._ensure_title(title, 1000)
                            if title:
                                try:
                                    title = self._format_fb_title_case1(title)
                                    title = self._limit_caption(title, 1000)
                                except Exception:
                                    pass
                            if vid_id and title:
                                try:
                                    update_title_if_empty(acc["uid"], vid_id, title)
                                except Exception:
                                    pass
                            self._set_status(item_id, f"DOWNLOAD OK {success_count+1}/{max_videos}")
                            self._log(f"[{acc['uid']}] DOWNLOAD OK")
                            caption = self._build_caption(title, 1000)
                            ok_p, drv, up_status, up_msg, caption = self._prepare_upload_with_retry(
                                driver_path,
                                remote,
                                path_or_err,
                                caption,
                                acc.get("uid", ""),
                                caption_limit=1000,
                            )

                            if not ok_p and up_status in ("select_not_found", "select_click_error"):
                                # Upload select not found - add to retry queue and break
                                self._set_status(item_id, f"UPLOAD LOI: {up_status}")
                                self._log(f"[{acc['uid']}] Upload {up_status} - retry this account")
                                self._record_failed(item_id, acc, f"UPLOAD {up_status}")
                                break
    
                            if not ok_p:
                                if up_status == "account_blocked" or "Could not detect Uploading/Uploaded status" in (up_msg or ""):
                                    self._set_status(item_id, "ACCOUNT BLOCKED")
                                    self._log(f"[{acc['uid']}] ACCOUNT BLOCKED - skip retry")
                                    self._record_failed(item_id, acc, "ACCOUNT BLOCKED")
                                    break
                                else:
                                    self._set_status(item_id, f"UPLOAD ERR: {up_msg or up_status}")
                                    self._log(f"[{acc['uid']}] UPLOAD ERR: {up_msg or up_status}")
                                    self.error_logger.log_upload_error(acc['uid'], path_or_err, up_msg or up_status)
                                    self._record_failed(item_id, acc, f"UPLOAD ERR: {up_msg or up_status}")
                                    # Other upload errors - skip without retry
                                    break
                            else:
                                self._set_status(item_id, f"POSTING {success_count+1}/{max_videos}...")
                                st, msg, purl, foll, posts = self._post_uploaded_video(drv, acc.get("uid", ""))
                                if st == "success":
                                    try:
                                        mark_uploaded(acc["uid"], mark_id)
                                    except Exception:
                                        pass
                                    self._set_profile_info(item_id, purl, foll, posts)
                                    self._set_status(item_id, "UPLOAD OK")
                                    self._log(f"[{acc['uid']}] UPLOAD OK")
                                    self._delete_uploaded_video(path_or_err, acc["uid"])
                                    self.error_logger.log_success(
                                        acc['uid'],
                                        "UPLOAD",
                                        f"Video {success_count+1}/{max_videos} posted successfully",
                                    )
                                    success_count += 1
                                else:
                                    err_text = msg or st
                                    status_text = "UPLOAD LOI" if "Select video not found" in (err_text or "") else f"UPLOAD ERR: {err_text}"
                                    self._set_status(item_id, status_text)
                                    self._log(f"[{acc['uid']}] UPLOAD ERR: {err_text}")
                                    self.error_logger.log_upload_error(acc['uid'], path_or_err, err_text)
                                    self._record_failed(item_id, acc, f"UPLOAD ERR: {err_text}")
                                    break
                        else:
                            err_text = str(path_or_err)
                            self._set_status(item_id, f"DOWNLOAD ERR: {err_text}")
                            self._log(f"[{acc['uid']}] DOWNLOAD ERR: {err_text}")
                            self.error_logger.log_download_error(acc['uid'], row_url, err_text)
                            lower = err_text.lower()
                            if (
                                "video unavailable" in lower
                                or "removed by the uploader" in lower
                                or "members-only" in lower
                                or "members only" in lower
                                or "join this channel" in lower
                                or "video skipped" in lower
                                or "private video" in lower
                                or "sign in if you've been granted access" in lower
                            ):
                                try:
                                    mark_uploaded(acc["uid"], mark_id)
                                except Exception:
                                    pass
                                continue
                            # Relogin disabled - skip video
                            break
                else:
                    status = self._format_login_error(err_login)
                    self._set_status(item_id, status)
                    self._log(f"[{acc['uid']}] {status}")
                    # Track failed account for retry
                    self._record_failed(item_id, acc, status)
                try:
                    if profile_id:
                        close_profile(profile_id, 3)
                        delete_profile(profile_id, 10)
                except Exception:
                    pass
                try:
                    if profile_id:
                        self.created_profiles.discard(profile_id)
                except Exception:
                    pass
                if profile_id:
                    self._delete_profile_path(profile_id)
                    self._track_profile_cleanup()
            try:
                with self.active_lock:
                    self.active_profiles.pop(item_id, None)
            except Exception:
                pass
        finally:
            if started:
                self._mark_run_done("upload", email)

if __name__ == "__main__":
    # Single-instance guard for frozen app (prevents multiple windows)
    if getattr(sys, "frozen", False):
        try:
            import msvcrt
            import tempfile
            import ctypes

            lock_path = os.path.join(tempfile.gettempdir(), "ScoopzTool.lock")
            _lock_file = open(lock_path, "a+")
            try:
                msvcrt.locking(_lock_file.fileno(), msvcrt.LK_NBLCK, 1)
            except Exception:
                ctypes.windll.user32.MessageBoxW(
                    0,
                    "ScoopzTool is already running.",
                    "ScoopzTool",
                    0x00000040,
                )
                sys.exit(0)
        except Exception:
            pass

    root = tk.Tk()
    App(root)
    root.mainloop()
